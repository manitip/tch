# app.py
# FastAPI API + планировщик отчётов + aiogram bot (единое приложение)
# По ТЗ: SQLite, allowlist (users.json), роли, Telegram WebApp auth (initData), расчёты как Excel.

from __future__ import annotations

import asyncio
import calendar
import dataclasses
import datetime as dt
import hashlib
import hmac
import importlib.util
import io
import json
import math
import os
import re
import secrets
import sqlite3
import time
import traceback
import urllib.parse
import shutil
import tempfile
import zipfile
from contextlib import asynccontextmanager
from typing import Any, Awaitable, Callable, Dict, List, Optional, Set, Tuple

from fastapi import (
    BackgroundTasks,
    Body,
    Depends,
    FastAPI,
    File,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel, ConfigDict, Field

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from zoneinfo import ZoneInfo

# aiogram v3
from aiogram import Bot, Dispatcher, Router, F
from aiogram.filters import Command
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.types import (
    Message,
    CallbackQuery,
    BufferedInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    WebAppInfo,
    MenuButtonDefault,
    MenuButtonWebApp,
)

# Optional for Excel export
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    openpyxl = None

from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

ATTACHMENTS_DIR = BASE_DIR / "uploads" / "receipts"
UPLOADS_DIR = BASE_DIR / "uploads"
BACKUPS_DIR = BASE_DIR / "backups"
ALLOWED_ATTACHMENT_MIME_TYPES = {
    "image/jpeg",
    "image/png",
    "image/webp",
    "application/pdf",
}
MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_ATTACHMENTS_PER_EXPENSE = 10
MAX_BACKUP_UPLOAD_BYTES = 200 * 1024 * 1024
ACCOUNTS = ("main", "praise", "alpha")

# ---------------------------
# Config
# ---------------------------

@dataclasses.dataclass
class Config:
    BOT_TOKEN: str
    APP_URL: str
    WEBAPP_URL: str
    DB_PATH: str
    USERS_JSON_PATH: str
    SESSION_SECRET: str
    TZ: str

    def tzinfo(self) -> ZoneInfo:
        try:
            return ZoneInfo(self.TZ)
        except Exception:
            return ZoneInfo("Europe/Warsaw")


def load_config() -> Config:
    bot_token = os.getenv("BOT_TOKEN", "").strip()
    if not bot_token:
        raise RuntimeError("BOT_TOKEN is required in .env / environment")

    base_dir = Path(__file__).resolve().parent
    db_path = os.getenv("DB_PATH", "db.sqlite3").strip()
    users_json_path = os.getenv("USERS_JSON", "users.json").strip()
    if not os.path.isabs(db_path):
        db_path = str(base_dir / db_path)
    if not os.path.isabs(users_json_path):
        users_json_path = str(base_dir / users_json_path)


    return Config(
        BOT_TOKEN=bot_token,
        APP_URL=os.getenv("APP_URL", "http://localhost:8000").strip(),
        WEBAPP_URL=os.getenv("WEBAPP_URL", "http://localhost:8000/webapp").strip(),
        DB_PATH=db_path,
        USERS_JSON_PATH=users_json_path,
        SESSION_SECRET=os.getenv("SESSION_SECRET", secrets.token_urlsafe(32)).strip(),
        TZ=os.getenv("TIMEZONE", "Europe/Warsaw").strip(),
    )


CFG = load_config()

def require_https_webapp_url(url: str) -> Optional[str]:
    if not url:
        return None
    try:
        parsed = urllib.parse.urlparse(url)
    except Exception:
        return None
    if parsed.scheme.lower() != "https":
        return None
    return url

def cashapp_webapp_url() -> Optional[str]:
    primary = require_https_webapp_url(f"{CFG.APP_URL.rstrip('/')}/cashapp")
    if primary:
        return primary
    webapp_url = require_https_webapp_url(CFG.WEBAPP_URL)
    if not webapp_url:
        return None
    parsed = urllib.parse.urlparse(webapp_url)
    cashapp_url = parsed._replace(path="/cashapp", params="", query="", fragment="")
    return urllib.parse.urlunparse(cashapp_url)



# ---------------------------
# DB helpers (sqlite3)
# ---------------------------

def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(CFG.DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn


def db_exec(sql: str, params: Tuple[Any, ...] = ()) -> None:
    with db_connect() as conn:
        conn.execute(sql, params)
        conn.commit()


def db_exec_returning_id(sql: str, params: Tuple[Any, ...] = ()) -> int:
    with db_connect() as conn:
        cur = conn.execute(sql, params)
        conn.commit()
        return int(cur.lastrowid)


def db_fetchone(sql: str, params: Tuple[Any, ...] = ()) -> Optional[sqlite3.Row]:
    with db_connect() as conn:
        cur = conn.execute(sql, params)
        return cur.fetchone()


def db_fetchall(sql: str, params: Tuple[Any, ...] = ()) -> List[sqlite3.Row]:
    with db_connect() as conn:
        cur = conn.execute(sql, params)
        return cur.fetchall()


def table_has_column(table: str, col: str) -> bool:
    rows = db_fetchall(f"PRAGMA table_info({table});")
    return any(r["name"] == col for r in rows)

def services_unique_has_account() -> bool:
    if not db_fetchone("SELECT name FROM sqlite_master WHERE type='table' AND name='services';"):
        return False
    rows = db_fetchall("PRAGMA index_list(services);")
    for r in rows:
        if int(r["unique"] or 0) != 1:
            continue
        idx_name = r["name"]
        cols = [c["name"] for c in db_fetchall(f"PRAGMA index_info({idx_name});")]
        if cols == ["month_id", "service_date", "account", "income_type"]:
            return True
    return False


def init_db() -> None:
    # users
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            name TEXT,
            role TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );
        """
    )
    # months
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS months (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            monthly_min_needed REAL NOT NULL DEFAULT 0,
            start_balance REAL NOT NULL DEFAULT 0,
            sundays_override INTEGER NULL,
            is_closed INTEGER NOT NULL DEFAULT 0,
            closed_at TEXT NULL,
            closed_by_user_id INTEGER NULL,            
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(year, month),
            FOREIGN KEY (closed_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        );
        """
    )
    # services
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month_id INTEGER NOT NULL,
            service_date TEXT NOT NULL,
            idx INTEGER NOT NULL,
            cashless REAL NOT NULL DEFAULT 0,
            cash REAL NOT NULL DEFAULT 0,
            total REAL NOT NULL DEFAULT 0,
            weekly_min_needed REAL NOT NULL DEFAULT 0,
            mnsps_status TEXT NOT NULL DEFAULT 'Не собрана',
            pvs_ratio REAL NOT NULL DEFAULT 0,
            income_type TEXT NOT NULL DEFAULT 'donation',
            account TEXT NOT NULL DEFAULT 'main',            
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(month_id, service_date, account, income_type),
            FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE
        );
        """
    )
    # expenses
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month_id INTEGER NOT NULL,
            expense_date TEXT NOT NULL,
            category TEXT NOT NULL,
            title TEXT NOT NULL,
            qty REAL NOT NULL DEFAULT 1,
            unit_amount REAL NOT NULL DEFAULT 0,
            total REAL NOT NULL DEFAULT 0,
            comment TEXT,
            is_system INTEGER NOT NULL DEFAULT 0,
            account TEXT NOT NULL DEFAULT 'main',           
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE
        );
        """
    )
    # cashflow (наличные подписи)
    import cashflow_models

    with db_connect() as conn:
        cashflow_models.init_cashflow_db(conn)
    # categories
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """
    )
    # month budgets
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS month_budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            limit_amount REAL NOT NULL,
            warn_threshold REAL NOT NULL DEFAULT 0.9,
            include_system INTEGER NOT NULL DEFAULT 1,
            created_by_user_id INTEGER NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(month_id, category_id),
            FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE,
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
        );
        """
    )
    # tags
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            name_norm TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """
    )
    db_exec("CREATE INDEX IF NOT EXISTS idx_tags_name_norm ON tags(name_norm);")
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS expense_tags (
            expense_id INTEGER NOT NULL,
            tag_id INTEGER NOT NULL,
            PRIMARY KEY (expense_id, tag_id),
            FOREIGN KEY (expense_id) REFERENCES expenses(id) ON DELETE CASCADE,
            FOREIGN KEY (tag_id) REFERENCES tags(id) ON DELETE CASCADE
        );
        """
    )
    db_exec("CREATE INDEX IF NOT EXISTS idx_expense_tags_tag_id ON expense_tags(tag_id);")
    db_exec("CREATE INDEX IF NOT EXISTS idx_expense_tags_expense_id ON expense_tags(expense_id);")
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS category_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER NOT NULL,
            alias TEXT NOT NULL,
            alias_norm TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(alias_norm),
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
        );
        """
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_category_aliases_category_id ON category_aliases(category_id);"
    )
    # drafts
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kind TEXT NOT NULL,
            month_id INTEGER NOT NULL,
            created_by_user_id INTEGER NOT NULL,
            payload_json TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE,
            FOREIGN KEY (created_by_user_id) REFERENCES users(id) ON DELETE CASCADE
        );
        """
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_drafts_month_user ON drafts(month_id, created_by_user_id, status);"
    )
    # settings (single row is fine)
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_chat_id INTEGER,
            sunday_report_time TEXT NOT NULL DEFAULT '18:00',
            month_report_time TEXT NOT NULL DEFAULT '21:00',
            timezone TEXT NOT NULL DEFAULT 'Europe/Warsaw',
            ui_theme TEXT NOT NULL DEFAULT 'auto',
            daily_expenses_enabled INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """
    )
    # bot subscribers (anyone who started bot can receive reports)
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS bot_subscribers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """
    )
    # audit_log␊
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id INTEGER,
            before_json TEXT,
            after_json TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );
        """
    )
    # attachments
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id INTEGER NOT NULL,
            orig_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            mime TEXT NOT NULL,
            size_bytes INTEGER NOT NULL,
            sha256 TEXT NOT NULL,
            created_by_user_id INTEGER NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY (created_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        );
        """
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_attach_entity ON attachments(entity_type, entity_id);"
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_attach_created ON attachments(created_at);"
    )

    # system logs
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS system_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,
            source TEXT NOT NULL,
            message TEXT NOT NULL,
            details_json TEXT NULL,
            trace TEXT NULL,
            created_at TEXT NOT NULL
        );
        """
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_syslog_time ON system_logs(level, created_at);"
    )
    # job runs
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS job_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT NOT NULL,
            status TEXT NOT NULL,
            started_at TEXT NOT NULL,
            finished_at TEXT NOT NULL,
            duration_ms INTEGER NOT NULL,
            error TEXT NULL
        );
        """
    )
    # message deliveries
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS message_deliveries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kind TEXT NOT NULL,
            recipient_id INTEGER NOT NULL,
            status TEXT NOT NULL,
            error TEXT NULL,
            created_at TEXT NOT NULL
        );
        """
    )

    # diagnostics runs
    db_exec(
        """
        CREATE TABLE IF NOT EXISTS diagnostic_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            created_by_user_id INTEGER NULL,
            suite TEXT NOT NULL,
            mode TEXT NOT NULL,
            status TEXT NOT NULL,
            started_at TEXT NULL,
            finished_at TEXT NULL,
            duration_ms INTEGER NULL,
            options_json TEXT NOT NULL,
            summary_json TEXT NULL,
            FOREIGN KEY (created_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        );
        """
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_diag_runs_created ON diagnostic_runs(created_at DESC);"
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_diag_runs_status ON diagnostic_runs(status, created_at DESC);"
    )

    db_exec(
        """
        CREATE TABLE IF NOT EXISTS diagnostic_steps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL,
            key TEXT NOT NULL,
            title TEXT NOT NULL,
            severity TEXT NOT NULL,
            status TEXT NOT NULL,
            started_at TEXT NULL,
            finished_at TEXT NULL,
            duration_ms INTEGER NULL,
            message TEXT NULL,
            details_json TEXT NULL,
            FOREIGN KEY (run_id) REFERENCES diagnostic_runs(id) ON DELETE CASCADE
        );
        """
    )
    db_exec(
        "CREATE INDEX IF NOT EXISTS idx_diag_steps_run ON diagnostic_steps(run_id, id);"
    )

    # Ensure settings row exists (id=1)
    row = db_fetchone("SELECT * FROM settings ORDER BY id LIMIT 1;")
    if not row:
        now = iso_now(CFG.tzinfo())
        db_exec(
            """
            INSERT INTO settings (
                report_chat_id, sunday_report_time, month_report_time,
                timezone, ui_theme, daily_expenses_enabled, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?);
            """,
            (None, "18:00", "21:00", CFG.TZ, "auto", 0, now, now),
        )

    if not table_has_column("services", "income_type"):
        db_exec("ALTER TABLE services ADD COLUMN income_type TEXT NOT NULL DEFAULT 'donation';")
    db_exec("UPDATE services SET income_type='donation' WHERE income_type IS NULL OR income_type='';")

    if not table_has_column("services", "account") or not services_unique_has_account():
        has_account = table_has_column("services", "account")
        db_exec(
            """
            CREATE TABLE services_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_id INTEGER NOT NULL,
                service_date TEXT NOT NULL,
                idx INTEGER NOT NULL,
                cashless REAL NOT NULL DEFAULT 0,
                cash REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL DEFAULT 0,
                weekly_min_needed REAL NOT NULL DEFAULT 0,
                mnsps_status TEXT NOT NULL DEFAULT 'Не собрана',
                pvs_ratio REAL NOT NULL DEFAULT 0,
                income_type TEXT NOT NULL DEFAULT 'donation',
                account TEXT NOT NULL DEFAULT 'main',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(month_id, service_date, account, income_type),
                FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE
            );
            """
        )
        account_expr = "COALESCE(account, 'main')" if has_account else "'main'"
        db_exec(
            f"""
            INSERT INTO services_new (
                id, month_id, service_date, idx, cashless, cash, total,
                weekly_min_needed, mnsps_status, pvs_ratio, income_type, account,
                created_at, updated_at
            )
            SELECT
                id, month_id, service_date, idx, cashless, cash, total,
                weekly_min_needed, mnsps_status, pvs_ratio,
                COALESCE(income_type, 'donation'),
                {account_expr},
                created_at, updated_at
            FROM services;
            """
        )
        db_exec("DROP TABLE services;")
        db_exec("ALTER TABLE services_new RENAME TO services;")
    if table_has_column("services", "account"):
        db_exec("UPDATE services SET account='main' WHERE account IS NULL OR account='';")

    if not table_has_column("expenses", "account"):
        db_exec("ALTER TABLE expenses ADD COLUMN account TEXT NOT NULL DEFAULT 'main';")
    if table_has_column("expenses", "account"):
        db_exec("UPDATE expenses SET account='main' WHERE account IS NULL OR account='';")

    if not table_has_column("months", "closed_at"):
        db_exec("ALTER TABLE months ADD COLUMN closed_at TEXT NULL;")
    if not table_has_column("months", "closed_by_user_id"):
        db_exec("ALTER TABLE months ADD COLUMN closed_by_user_id INTEGER NULL;")
    if not table_has_column("months", "is_closed"):
        db_exec("ALTER TABLE months ADD COLUMN is_closed INTEGER NOT NULL DEFAULT 0;")
        db_exec("UPDATE months SET is_closed=1 WHERE closed_at IS NOT NULL;")

    ensure_categories_from_expenses()


def iso_now(tz: ZoneInfo) -> str:
    return dt.datetime.now(tz=tz).replace(microsecond=0).isoformat()


def iso_date(d: dt.date) -> str:
    return d.isoformat()


def parse_iso_date(s: str) -> dt.date:
    return dt.date.fromisoformat(s)

def normalize_account(value: Optional[str]) -> str:
    s = str(value or "").strip().lower()
    if not s:
        return "main"
    if s not in ACCOUNTS:
        raise HTTPException(status_code=400, detail="Invalid account")
    return s


def normalize_alias(alias: str) -> str:
    s = str(alias or "").strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"[.,;:/\\\\]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_tag_name(tag: str) -> str:
    s = str(tag or "").strip().lower()
    if not s:
        return ""
    s = s.replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_tag_list(tags: Optional[List[str]]) -> List[str]:
    if not tags:
        return []
    cleaned: List[str] = []
    seen: set[str] = set()
    for tag in tags:
        name = str(tag or "").strip()
        if not name:
            continue
        norm = normalize_tag_name(name)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        cleaned.append(name)
    return cleaned




def ensure_categories_from_expenses() -> None:
    if not db_fetchone("SELECT name FROM sqlite_master WHERE type='table' AND name='categories';"):
        return
    now = iso_now(CFG.tzinfo())
    rows = db_fetchall("SELECT DISTINCT category FROM expenses;")
    for r in rows:
        name = str(r["category"] or "").strip()
        if not name:
            continue
        existing = db_fetchone("SELECT id FROM categories WHERE name=?;", (name,))
        if not existing:
            db_exec(
                """
                INSERT INTO categories (name, is_active, sort_order, created_at, updated_at)
                VALUES (?, 1, 0, ?, ?);
                """,
                (name, now, now),
            )


def resolve_category(name: str, role: str, user_id: Optional[int] = None) -> str:
    raw = str(name or "").strip() or "Прочее"
    alias_norm = normalize_alias(raw)
    if alias_norm:
        row = db_fetchone(
            """
            SELECT c.name
            FROM category_aliases a
            JOIN categories c ON c.id=a.category_id
            WHERE a.alias_norm=?;
            """,
            (alias_norm,),
        )
        if row:
            return str(row["name"])

    existing = db_fetchone("SELECT * FROM categories WHERE name=?;", (raw,))
    if existing:
        return str(existing["name"])

    if role in ("admin", "accountant"):
        now = iso_now(CFG.tzinfo())
        new_id = db_exec_returning_id(
            """
            INSERT INTO categories (name, is_active, sort_order, created_at, updated_at)
            VALUES (?, 1, 0, ?, ?);
            """,
            (raw, now, now),
        )
        after = db_fetchone("SELECT * FROM categories WHERE id=?;", (new_id,))
        log_audit(user_id, "CREATE", "category", int(new_id), None, dict(after) if after else None)
        return raw

    return raw

def get_tags_payload() -> List[Dict[str, Any]]:
    tags = db_fetchall("SELECT * FROM tags ORDER BY name COLLATE NOCASE;")
    return [dict(t) for t in tags]


def get_expense_tag_names(expense_id: int) -> List[str]:
    rows = db_fetchall(
        """
        SELECT t.name
        FROM expense_tags et
        JOIN tags t ON t.id=et.tag_id
        WHERE et.expense_id=?
        ORDER BY t.name COLLATE NOCASE;
        """,
        (expense_id,),
    )
    return [str(r["name"]) for r in rows]


def fetch_expense_tags(expense_ids: List[int]) -> Dict[int, List[str]]:
    if not expense_ids:
        return {}
    placeholders = ",".join("?" for _ in expense_ids)
    rows = db_fetchall(
        f"""
        SELECT et.expense_id, t.name
        FROM expense_tags et
        JOIN tags t ON t.id=et.tag_id
        WHERE et.expense_id IN ({placeholders})
        ORDER BY t.name COLLATE NOCASE;
        """,
        tuple(expense_ids),
    )
    tag_map: Dict[int, List[str]] = {}
    for r in rows:
        tag_map.setdefault(int(r["expense_id"]), []).append(str(r["name"]))
    return tag_map


def set_expense_tags(
    expense_id: int,
    tag_names: Optional[List[str]],
    role: str,
    user_id: Optional[int],
) -> List[str]:
    tags = normalize_tag_list(tag_names)
    with db_connect() as conn:
        conn.execute("DELETE FROM expense_tags WHERE expense_id=?;", (expense_id,))
        if not tags:
            conn.commit()
            return []

        resolved_ids: List[int] = []
        for name in tags:
            norm = normalize_tag_name(name)
            row = conn.execute("SELECT id, name FROM tags WHERE name_norm=?;", (norm,)).fetchone()
            if row:
                resolved_ids.append(int(row["id"]))
                continue
            if role not in ("admin", "accountant"):
                raise HTTPException(status_code=403, detail="Insufficient role for tag creation")
            now = iso_now(CFG.tzinfo())
            cur = conn.execute(
                """
                INSERT INTO tags (name, name_norm, created_at, updated_at)
                VALUES (?, ?, ?, ?);
                """,
                (name, norm, now, now),
            )
            new_id = int(cur.lastrowid)
            resolved_ids.append(new_id)
            after = conn.execute("SELECT * FROM tags WHERE id=?;", (new_id,)).fetchone()
            log_audit(user_id, "CREATE", "tag", new_id, None, dict(after) if after else None)

        conn.executemany(
            "INSERT INTO expense_tags (expense_id, tag_id) VALUES (?, ?);",
            [(expense_id, tag_id) for tag_id in resolved_ids],
        )
        conn.commit()
    return tags



def get_categories_payload() -> List[Dict[str, Any]]:
    categories = db_fetchall(
        "SELECT * FROM categories ORDER BY sort_order ASC, name COLLATE NOCASE;"
    )
    aliases = db_fetchall(
        "SELECT * FROM category_aliases ORDER BY created_at ASC, id ASC;"
    )
    counts = db_fetchall("SELECT category, COUNT(*) AS cnt FROM expenses GROUP BY category;")
    count_map = {str(r["category"]): int(r["cnt"]) for r in counts}

    alias_map: Dict[int, List[Dict[str, Any]]] = {}
    for a in aliases:
        alias_map.setdefault(int(a["category_id"]), []).append(dict(a))

    items = []
    for c in categories:
        item = dict(c)
        item["aliases"] = alias_map.get(int(c["id"]), [])
        item["expense_count"] = count_map.get(str(c["name"]), 0)
        items.append(item)
    return items

def budget_status_from_usage(usage: Optional[float], warn_threshold: float) -> str:
    if usage is None:
        return "OK"
    if usage < warn_threshold:
        return "OK"
    if usage < 1:
        return "WARN"
    return "OVER"


def get_budget_fact(month_id: int, category_name: str, include_system: bool) -> float:
    row = db_fetchone(
        """
        SELECT COALESCE(SUM(total), 0) AS s
        FROM expenses
        WHERE month_id=? AND category=? AND account='main' AND (?=1 OR is_system=0);
        """,
        (month_id, category_name, 1 if include_system else 0),
    )
    return float(row["s"] if row and row["s"] is not None else 0.0)


def get_month_budget_rows(month_id: int) -> List[Dict[str, Any]]:
    rows = db_fetchall(
        """
        SELECT mb.*, c.name AS category_name, c.sort_order AS category_sort
        FROM month_budgets mb
        JOIN categories c ON c.id=mb.category_id
        WHERE mb.month_id=?
        ORDER BY c.sort_order ASC, c.name COLLATE NOCASE;
        """,
        (month_id,),
    )
    items: List[Dict[str, Any]] = []
    for row in rows:
        limit_amount = float(row["limit_amount"])
        warn_threshold = float(row["warn_threshold"] or 0.9)
        include_system = bool(int(row["include_system"] or 0) == 1)
        category_name = str(row["category_name"])
        fact = get_budget_fact(month_id, category_name, include_system)
        usage: Optional[float] = None
        if limit_amount > 0:
            usage = fact / limit_amount
        delta = limit_amount - fact
        status = budget_status_from_usage(usage, warn_threshold)
        item = dict(row)
        item["category_name"] = category_name
        item["include_system"] = 1 if include_system else 0
        item["fact"] = round(fact, 2)
        item["usage"] = round(usage, 3) if usage is not None else None
        item["status"] = status
        item["delta"] = round(delta, 2)
        items.append(item)
    return items


def get_budget_warning_for_category(month_id: int, category_name: str) -> Optional[Dict[str, Any]]:
    row = db_fetchone(
        """
        SELECT mb.*, c.name AS category_name
        FROM month_budgets mb
        JOIN categories c ON c.id=mb.category_id
        WHERE mb.month_id=? AND c.name=?;
        """,
        (month_id, category_name),
    )
    if not row:
        return None
    limit_amount = float(row["limit_amount"])
    warn_threshold = float(row["warn_threshold"] or 0.9)
    include_system = bool(int(row["include_system"] or 0) == 1)
    fact = get_budget_fact(month_id, category_name, include_system)
    usage: Optional[float] = None
    if limit_amount > 0:
        usage = fact / limit_amount
    status = budget_status_from_usage(usage, warn_threshold)
    if status != "OVER":
        return None
    return {
        "type": "budget_over",
        "category": category_name,
        "fact": round(fact, 2),
        "limit": round(limit_amount, 2),
        "usage": round(usage, 3) if usage is not None else None,
    }



def is_month_closed(month_id: int) -> bool:
    if table_has_column("months", "is_closed"):
        row = db_fetchone("SELECT is_closed FROM months WHERE id=?;", (month_id,))
        return bool(row and int(row["is_closed"] or 0) == 1)
    if table_has_column("months", "closed_at"):
        row = db_fetchone("SELECT closed_at FROM months WHERE id=?;", (month_id,))
        return bool(row and row["closed_at"])
    return False


def ensure_month_open(month_id: int) -> None:
    if is_month_closed(month_id):
        raise HTTPException(status_code=423, detail="Month is closed")


def normalize_expense_draft_payload(raw: Dict[str, Any], tz: ZoneInfo) -> Dict[str, Any]:
    today = dt.datetime.now(tz=tz).date()
    raw_date = raw.get("expense_date")
    if isinstance(raw_date, dt.date):
        expense_date = raw_date.isoformat()
    elif raw_date:
        try:
            expense_date = dt.date.fromisoformat(str(raw_date)).isoformat()
        except ValueError:
            expense_date = today.isoformat()
    else:
        expense_date = today.isoformat()

    def normalize_float(value: Any, default: float) -> float:
        if value in (None, "", "null"):
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    qty = normalize_float(raw.get("qty"), 1.0)
    unit_amount = normalize_float(raw.get("unit_amount"), 0.0)
    category = str(raw.get("category") or "").strip() or "Прочее"
    title = str(raw.get("title") or "").strip() or "Расход"
    comment = raw.get("comment")
    if comment is not None:
        comment = str(comment).strip()
        if not comment:
            comment = None
    tags = raw.get("tags")
    if tags is not None and not isinstance(tags, list):
        tags = [str(tags)]
    tags = normalize_tag_list(tags)
    account = normalize_account(raw.get("account"))

    payload = {
        "expense_date": expense_date,
        "category": category,
        "title": title,
        "qty": qty,
        "unit_amount": unit_amount,
        "comment": comment,
        "account": account,
    }
    if tags:
        payload["tags"] = tags
    return payload


def draft_payload_summary(payload: Dict[str, Any]) -> Dict[str, Any]:
    qty = float(payload.get("qty") or 0)
    unit_amount = float(payload.get("unit_amount") or 0)
    total = round(qty * unit_amount, 2)
    return {
        "title": payload.get("title"),
        "category": payload.get("category"),
        "expense_date": payload.get("expense_date"),
        "total": total,
    }


def get_expense_or_404(expense_id: int) -> sqlite3.Row:
    row = db_fetchone("SELECT * FROM expenses WHERE id=?;", (expense_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Expense not found")
    return row


def get_draft_or_404(draft_id: int) -> sqlite3.Row:
    row = db_fetchone("SELECT * FROM drafts WHERE id=?;", (draft_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Draft not found")
    return row


def get_attachment_or_404(attachment_id: int) -> sqlite3.Row:
    row = db_fetchone("SELECT * FROM attachments WHERE id=?;", (attachment_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    return row


def attachment_storage_dir(entity_id: int, created_at: str) -> Path:
    try:
        created_dt = dt.datetime.fromisoformat(created_at)
    except ValueError:
        created_dt = dt.datetime.now(tz=CFG.tzinfo())
    ym = f"{created_dt.year}-{created_dt.month:02d}"
    return ATTACHMENTS_DIR / ym / str(entity_id)


def attachment_extension(orig_filename: str, mime: str) -> str:
    ext = Path(orig_filename or "").suffix.lower()
    if ext:
        return ext
    return {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "application/pdf": ".pdf",
    }.get(mime, "")


# ---------------------------
# Allowlist (users.json) + sync to DB
# ---------------------------

ALLOWLIST_CACHE: Dict[int, Dict[str, Any]] = {}
ALLOWLIST_MTIME: Optional[float] = None


def load_allowlist() -> Dict[int, Dict[str, Any]]:
    """
    users.json:
    [
      {"telegram_id": 123, "name": "Иван", "role": "admin", "active": true},
      ...
    ]
    """
    path = CFG.USERS_JSON_PATH
    if not os.path.exists(path):
        # пустой allowlist = никого не пускаем
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    out: Dict[int, Dict[str, Any]] = {}
    for item in data:
        try:
            tid = int(item["telegram_id"])
            out[tid] = {
                "telegram_id": tid,
                "name": str(item.get("name", "")).strip(),
                "role": str(item.get("role", "viewer")).strip(),
                "active": bool(item.get("active", True)),
            }
        except Exception:
            continue
    return out


def sync_allowlist_to_db(allow: Dict[int, Dict[str, Any]]) -> None:
    now = iso_now(CFG.tzinfo())
    for tid, u in allow.items():
        existing = db_fetchone("SELECT * FROM users WHERE telegram_id=?;", (tid,))
        if not existing:
            db_exec(
                """
                INSERT INTO users (telegram_id, name, role, active, created_at)
                VALUES (?, ?, ?, ?, ?);
                """,
                (tid, u.get("name"), u.get("role", "viewer"), 1 if u.get("active", True) else 0, now),
            )
        else:
            db_exec(
                """
                UPDATE users SET name=?, role=?, active=?
                WHERE telegram_id=?;
                """,
                (u.get("name"), u.get("role", existing["role"]), 1 if u.get("active", True) else 0, tid),
            )

def refresh_allowlist_if_needed() -> Dict[int, Dict[str, Any]]:
    global ALLOWLIST_CACHE, ALLOWLIST_MTIME
    path = CFG.USERS_JSON_PATH
    try:
        mtime = os.path.getmtime(path)
    except FileNotFoundError:
        mtime = None
    if mtime != ALLOWLIST_MTIME:
        allow = load_allowlist()
        sync_allowlist_to_db(allow)
        ALLOWLIST_CACHE = allow
        ALLOWLIST_MTIME = mtime
        return ALLOWLIST_CACHE
    return ALLOWLIST_CACHE

def create_cashflow_collect_request_if_needed(
    *,
    account: str,
    cash_amount: float,
    cashless_amount: float,
    income_type: str,
    month_id: int,
    service_date: str,
    created_by_telegram_id: Optional[int],
) -> Optional[int]:
    if cash_amount <= 0:
        return None
    account_norm = normalize_account(account)
    if account_norm not in ("main", "praise", "alpha"):
        return None
    if not created_by_telegram_id:
        return None
    import cashflow_models as cf

    payload = {
        "month_id": int(month_id),
        "service_date": str(service_date),
        "cashless": float(cashless_amount),
        "income_type": str(income_type),
    }
    payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)

    cfg_base = cf.load_cashflow_config(BASE_DIR)
    cfg = cf.CashflowConfig(
        base_dir=cfg_base.base_dir,
        db_path=Path(CFG.DB_PATH),
        users_json_path=Path(CFG.USERS_JSON_PATH),
        uploads_dir=cfg_base.uploads_dir,
        timezone=cfg_base.timezone,
    )
    with db_connect() as conn:
        cf.init_cashflow_db(conn)
        existing = conn.execute(
            """
            SELECT id
            FROM cash_requests
            WHERE account=? AND op_type='collect'
              AND status IN ('PENDING_SIGNERS','PENDING_ADMIN')
              AND amount=? AND created_by_telegram_id=?
              AND source_kind='service' AND source_payload=?              
            ORDER BY id DESC
            LIMIT 1;
            """,
            (account_norm, float(cash_amount), int(created_by_telegram_id), payload_json),
        ).fetchone()
        if existing:
            return int(existing["id"])

        return cf.create_cash_request(
            conn,
            cfg,
            account=account_norm,
            op_type="collect",
            amount=float(cash_amount),
            created_by_telegram_id=int(created_by_telegram_id),
            source_kind="service",
            source_id=None,
            source_payload=payload,
        )

def finalize_cashflow_collect_request(request_id: int) -> Optional[int]:
    req = db_fetchone("SELECT * FROM cash_requests WHERE id=?;", (int(request_id),))
    if not req:
        return None
    if hasattr(req, "keys"):
        req = dict(req)
    if str(req["status"]) != "FINAL" or str(req["op_type"]) != "collect":
        return None
    if str(req.get("source_kind") or "") != "service":
        return None
    if req.get("source_id"):
        return int(req["source_id"])
    payload_raw = req.get("source_payload")
    if not payload_raw:
        return None
    try:
        payload = json.loads(payload_raw)
    except Exception:
        return None
    service_date = str(payload.get("service_date") or "").strip()
    if not service_date:
        return None
    try:
        parse_iso_date(service_date)
    except Exception:
        return None
    month_id_raw = payload.get("month_id")
    if month_id_raw is None:
        return None
    month_id = int(month_id_raw)
    cashless = float(payload.get("cashless") or 0.0)
    income_type = str(payload.get("income_type") or "donation").strip().lower()
    if income_type not in ("donation", "other"):
        income_type = "donation"
    account = normalize_account(req["account"])
    cash = float(req["amount"])
    total = round(cashless + cash, 2)

    actor_tid = req.get("created_by_telegram_id") or req.get("admin_telegram_id")
    user_row = None
    if actor_tid:
        user_row = db_fetchone("SELECT id FROM users WHERE telegram_id=?;", (int(actor_tid),))
    user_id = int(user_row["id"]) if user_row else None

    before = db_fetchone(
        """
        SELECT * FROM services
        WHERE month_id=? AND service_date=? AND account=? AND income_type=?;
        """,
        (month_id, service_date, account, income_type),
    )

    now = iso_now(CFG.tzinfo())
    if before:
        db_exec(
            """
            UPDATE services
            SET cashless=?, cash=?, total=?, income_type=?, account=?, updated_at=?
            WHERE id=?;
            """,
            (cashless, cash, total, income_type, account, now, before["id"]),
        )
        after = db_fetchone("SELECT * FROM services WHERE id=?;", (before["id"],))
        log_audit(user_id, "UPDATE", "service", int(before["id"]), dict(before), dict(after) if after else None)
        service_id = int(before["id"])
    else:
        service_id = db_exec_returning_id(
            """
            INSERT INTO services (
                month_id, service_date, idx, cashless, cash, total,
                weekly_min_needed, mnsps_status, pvs_ratio, income_type, account,
                created_at, updated_at
            ) VALUES (?, ?, 0, ?, ?, ?, 0, 'Не собрана', 0, ?, ?, ?, ?);
            """,
            (month_id, service_date, cashless, cash, total, income_type, account, now, now),

        )
        after = db_fetchone("SELECT * FROM services WHERE id=?;", (service_id,))
        log_audit(user_id, "CREATE", "service", int(service_id), None, dict(after) if after else None)

    db_exec(
        "UPDATE cash_requests SET source_id=?, updated_at=? WHERE id=?;",
        (service_id, now, int(request_id)),
    )
    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=user_id)
    return service_id

def register_bot_subscriber(telegram_id: int) -> None:
    now = iso_now(CFG.tzinfo())
    existing = db_fetchone("SELECT * FROM bot_subscribers WHERE telegram_id=?;", (telegram_id,))
    if not existing:
        db_exec(
            """
            INSERT INTO bot_subscribers (telegram_id, active, created_at, updated_at)
            VALUES (?, 1, ?, ?);
            """,
            (telegram_id, now, now),
        )
        return
    db_exec(
        """
        UPDATE bot_subscribers SET active=1, updated_at=?
        WHERE telegram_id=?;
        """,
        (now, telegram_id),
    )

def list_report_recipients(settings_row: Optional[sqlite3.Row] = None) -> List[int]:
    s = settings_row or get_settings()
    ids: set[int] = set()
    chat_id = s["report_chat_id"]
    if chat_id:
        ids.add(int(chat_id))
    allow = refresh_allowlist_if_needed()
    allowed_ids = {tid for tid, u in allow.items() if u.get("active") is True}
    if allowed_ids:
        placeholders = ",".join("?" for _ in allowed_ids)
        rows = db_fetchall(
            f"""
            SELECT telegram_id
            FROM bot_subscribers
            WHERE active=1
              AND telegram_id IN ({placeholders});
            """,
            tuple(allowed_ids),
        )
    else:
        rows = []
    ids.update(int(r["telegram_id"]) for r in rows)
    return sorted(ids)
# ---------------------------
# Session token (HMAC signed JSON) - self-contained (no external JWT deps)
# ---------------------------


def _b64url_encode(raw: bytes) -> str:
    import base64
    return base64.urlsafe_b64encode(raw).decode("utf-8").rstrip("=")


def _b64url_decode(s: str) -> bytes:
    import base64
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)


def make_session_token(payload: Dict[str, Any], secret: str) -> str:
    payload_bytes = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    body = _b64url_encode(payload_bytes)
    sig = hmac.new(secret.encode("utf-8"), body.encode("utf-8"), hashlib.sha256).digest()
    return f"{body}.{_b64url_encode(sig)}"


def verify_session_token(token: str, secret: str) -> Dict[str, Any]:
    try:
        body, sig = token.split(".", 1)
    except ValueError:
        raise HTTPException(status_code=401, detail="Invalid token format")

    expected = hmac.new(secret.encode("utf-8"), body.encode("utf-8"), hashlib.sha256).digest()
    try:
        provided_sig = _b64url_decode(sig)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token signature")

    if not hmac.compare_digest(provided_sig, expected):
        raise HTTPException(status_code=401, detail="Invalid token signature")

    try:
        payload = json.loads(_b64url_decode(body).decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token payload")

    if not isinstance(payload, dict):
        raise HTTPException(status_code=401, detail="Invalid token payload")

    exp = int(payload.get("exp", 0))
    if exp and int(time.time()) > exp:
        raise HTTPException(status_code=401, detail="Token expired")
    return payload


# ---------------------------
# Telegram WebApp initData validation
# ---------------------------

def validate_telegram_init_data(init_data: str, bot_token: str, max_age_sec: int = 7 * 24 * 3600) -> Dict[str, Any]:
    """
    Telegram WebApp initData verification:
    - Parse query string
    - Exclude 'hash'
    - data_check_string = '\n'.join(sorted([f"{k}={v}"]))
    - secret_key = HMAC_SHA256(key=b"WebAppData", msg=bot_token)
    - check_hash = HMAC_SHA256(key=secret_key, msg=data_check_string).hexdigest()
    """
    if not init_data:
        raise HTTPException(status_code=400, detail="initData is required")

    parsed = dict(urllib.parse.parse_qsl(init_data, keep_blank_values=True))
    received_hash = parsed.get("hash")
    if not received_hash:
        raise HTTPException(status_code=401, detail="Missing hash in initData")

    # auth_date freshness (optional but recommended)
    try:
        auth_date = int(parsed.get("auth_date", "0"))
        if auth_date and (int(time.time()) - auth_date) > max_age_sec:
            raise HTTPException(status_code=401, detail="initData is too old")
    except ValueError:
        pass

    data_pairs = []
    for k, v in parsed.items():
        if k == "hash":
            continue
        data_pairs.append(f"{k}={v}")
    data_pairs.sort()
    data_check_string = "\n".join(data_pairs)

    secret_key = hmac.new(b"WebAppData", bot_token.encode("utf-8"), hashlib.sha256).digest()
    computed_hash = hmac.new(secret_key, data_check_string.encode("utf-8"), hashlib.sha256).hexdigest()

    if not hmac.compare_digest(computed_hash, received_hash):
        raise HTTPException(status_code=401, detail="initData validation failed")

    # Extract user (JSON string)
    user_raw = parsed.get("user")
    if not user_raw:
        raise HTTPException(status_code=401, detail="No user in initData")

    try:
        user_obj = json.loads(user_raw)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid user payload in initData")

    return user_obj


# ---------------------------
# Domain calculations (Excel logic)
# ---------------------------

def count_sundays_in_month(year: int, month: int) -> int:
    c = calendar.Calendar(firstweekday=0)  # Monday is 0; doesn't matter
    sundays = 0
    for d in c.itermonthdates(year, month):
        if d.month == month and d.weekday() == 6:
            sundays += 1
    return sundays


def last_day_of_month(year: int, month: int) -> dt.date:
    last = calendar.monthrange(year, month)[1]
    return dt.date(year, month, last)


def get_settings() -> sqlite3.Row:
    row = db_fetchone("SELECT * FROM settings ORDER BY id LIMIT 1;")
    if not row:
        raise HTTPException(status_code=500, detail="Settings not initialized")
    return row


def get_month_by_id(month_id: int) -> sqlite3.Row:
    m = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    if not m:
        raise HTTPException(status_code=404, detail="Month not found")
    return m


def get_or_create_month(year: int, month: int) -> sqlite3.Row:
    m = db_fetchone("SELECT * FROM months WHERE year=? AND month=?;", (year, month))
    if m:
        return m

    # start_balance from previous month fact_balance
    tz = CFG.tzinfo()
    now = iso_now(tz)

    prev_y, prev_m = year, month - 1
    if prev_m == 0:
        prev_m = 12
        prev_y -= 1

    prev = db_fetchone("SELECT * FROM months WHERE year=? AND month=?;", (prev_y, prev_m))
    start_balance = 0.0
    if prev:
        prev_summary = compute_month_summary(prev["id"], ensure_tithe=True)
        start_balance = float(prev_summary["fact_balance"])

    new_id = db_exec_returning_id(
        """
        INSERT INTO months (year, month, monthly_min_needed, start_balance, sundays_override, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?);
        """,
        (year, month, 0.0, float(start_balance), None, now, now),
    )
    m2 = db_fetchone("SELECT * FROM months WHERE id=?;", (new_id,))
    assert m2 is not None
    return m2


def calc_weekly_min_needed(month_row: sqlite3.Row) -> float:
    year, month = int(month_row["year"]), int(month_row["month"])
    override = month_row["sundays_override"]
    n = int(override) if override else count_sundays_in_month(year, month)
    if n <= 0:
        n = 4
    monthly_min_needed = float(month_row["monthly_min_needed"] or 0.0)
    return round(monthly_min_needed / n, 2) if monthly_min_needed else 0.0


def recalc_services_for_month(month_id: int) -> None:
    m = get_month_by_id(month_id)
    weekly_min = calc_weekly_min_needed(m)
    tz = CFG.tzinfo()
    now = iso_now(tz)

    services = db_fetchall("SELECT * FROM services WHERE month_id=? ORDER BY service_date ASC;", (month_id,))
    for s in services:
        cashless = float(s["cashless"] or 0.0)
        cash = float(s["cash"] or 0.0)
        total = round(cashless + cash, 2)
        income_type = str(s["income_type"] or "donation")
        account = str(s["account"] or "main")
        if account != "main":
            status = "Доп. счет"
            pvs = 0.0
            weekly_min_for_row = 0.0
        elif income_type == "donation":
            status = "Собрана" if (weekly_min and total >= weekly_min) else "Не собрана"
            pvs = (total / weekly_min) if weekly_min else 0.0
            weekly_min_for_row = weekly_min
        else:
            status = "Иной доход"
            pvs = 0.0
            weekly_min_for_row = 0.0
        db_exec(
            """
            UPDATE services
            SET total=?, weekly_min_needed=?, mnsps_status=?, pvs_ratio=?, updated_at=?
            WHERE id=?;
            """,
            (total, weekly_min_for_row, status, pvs, now, s["id"]),
        )

    # Update idx sequentially by date
    services2 = db_fetchall(
        """
        SELECT id FROM services
        WHERE month_id=? AND account='main' AND income_type='donation'
        ORDER BY service_date ASC;
        """,
        (month_id,),
    )
    for i, row in enumerate(services2, start=1):
        db_exec("UPDATE services SET idx=? WHERE id=?;", (i, row["id"]))
    db_exec(
        """
        UPDATE services
        SET idx=0
        WHERE month_id=?
          AND (account!='main' OR income_type!='donation' OR income_type IS NULL);
        """,
        (month_id,),
    )


def ensure_tithe_expense(month_id: int, user_id: Optional[int] = None) -> None:
    """
    Auto expense:
      category="Десятина"
      title="10% Объединение"
      qty=1
      unit_amount=tithe_amount
      total=tithe_amount
      is_system=true
      date=last day of month
    """
    m = get_month_by_id(month_id)
    year, month = int(m["year"]), int(m["month"])
    tithe_date = last_day_of_month(year, month)
    income_sum = float(
        db_fetchone(
            """
            SELECT COALESCE(SUM(total),0) AS s
            FROM services
            WHERE month_id=?
              AND account='main'            
              AND (income_type='donation' OR income_type IS NULL)
            """,
            (month_id,),
        )["s"]
    )
    tithe_amount = round(income_sum * 0.10, 2)

    existing = db_fetchone(
        """
        SELECT * FROM expenses
        WHERE month_id=? AND is_system=1 AND title='10% Объединение'
        LIMIT 1;
        """,
        (month_id,),
    )
    tz = CFG.tzinfo()
    now = iso_now(tz)
    category = resolve_category("Десятина", "admin", user_id)

    if existing:
        before = dict(existing)
        db_exec(
            """
            UPDATE expenses
            SET expense_date=?, category=?, qty=1, unit_amount=?, total=?, account='main', updated_at=?          
            WHERE id=?;
            """,
            (iso_date(tithe_date), category, tithe_amount, tithe_amount, now, existing["id"]),
        )
        after = dict(db_fetchone("SELECT * FROM expenses WHERE id=?;", (existing["id"],)))
        log_audit(user_id, "UPSERT_SYSTEM_TITHE", "expense", int(existing["id"]), before, after)
    else:
        new_id = db_exec_returning_id(
            """
            INSERT INTO expenses (
                month_id, expense_date, category, title, qty, unit_amount, total, comment,
                is_system, account, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 1, ?, ?, NULL, 1, 'main', ?, ?);
            """,
            (month_id, iso_date(tithe_date), category, "10% Объединение", tithe_amount, tithe_amount, now, now),
        )
        after = dict(db_fetchone("SELECT * FROM expenses WHERE id=?;", (new_id,)))
        log_audit(user_id, "CREATE_SYSTEM_TITHE", "expense", int(new_id), None, after)


def compute_month_summary(month_id: int, ensure_tithe: bool = True) -> Dict[str, Any]:
    m = get_month_by_id(month_id)
    # Services should be recalculated in case weekly_min_needed changed
    recalc_services_for_month(month_id)

    if ensure_tithe:
        ensure_tithe_expense(month_id, user_id=None)

    income_sum = float(
        db_fetchone(
            "SELECT COALESCE(SUM(total),0) AS s FROM services WHERE month_id=? AND account='main';",
            (month_id,),
        )["s"]
    )
    expenses_sum = float(
        db_fetchone(
            "SELECT COALESCE(SUM(total),0) AS s FROM expenses WHERE month_id=? AND account='main';",
            (month_id,),
        )["s"]
    )

    month_balance = round(income_sum - expenses_sum, 2)
    start_balance = float(m["start_balance"] or 0.0)
    fact_balance = round(start_balance + income_sum - expenses_sum, 2)

    monthly_min_needed = float(m["monthly_min_needed"] or 0.0)
    if income_sum > monthly_min_needed:
        sddr = round(income_sum - monthly_min_needed, 2)
    else:
        sddr = 0.0

    monthly_completion = 0.0
    if monthly_min_needed > 0:
        monthly_completion = min(income_sum / monthly_min_needed, 1.0)

    # psdpm
    year, month = int(m["year"]), int(m["month"])
    prev_y, prev_m = year, month - 1
    if prev_m == 0:
        prev_m = 12
        prev_y -= 1
    prev = db_fetchone("SELECT id FROM months WHERE year=? AND month=?;", (prev_y, prev_m))
    psdpm = None
    if prev:
        prev_income = float(
            db_fetchone(
                "SELECT COALESCE(SUM(total),0) AS s FROM services WHERE month_id=? AND account='main';",
                (prev["id"],),
            )["s"]
        )
        if prev_income > 0:
            psdpm = (income_sum - prev_income) / prev_income
        else:
            psdpm = None

    # avg_sunday: avg services.total where total>0
    row = db_fetchone(
        """
        SELECT AVG(total) AS a
        FROM services
        WHERE month_id=?
          AND account='main'
          AND total>0
          AND (income_type='donation' OR income_type IS NULL);
        """,
        (month_id,),
    )
    avg_sunday = float(row["a"]) if row and row["a"] is not None else 0.0

    # counts / weekly min
    override = m["sundays_override"]
    sundays_count = int(override) if override else count_sundays_in_month(year, month)
    weekly_min = calc_weekly_min_needed(m)

    is_closed = bool(int(m["is_closed"] or 0) == 1) if table_has_column("months", "is_closed") else bool(
        m["closed_at"] if table_has_column("months", "closed_at") else False
    )
    closed_at = m["closed_at"] if table_has_column("months", "closed_at") else None
    closed_by = None
    if table_has_column("months", "closed_by_user_id"):
        closed_by_id = m["closed_by_user_id"]
        if closed_by_id:
            row = db_fetchone(
                "SELECT id, telegram_id, name FROM users WHERE id=?;",
                (int(closed_by_id),),
            )
            if row:
                closed_by = {
                    "id": int(row["id"]),
                    "telegram_id": int(row["telegram_id"]),
                    "name": row["name"],
                }

    subaccounts: Dict[str, Dict[str, float]] = {}
    for account in ("praise", "alpha"):
        income = float(
            db_fetchone(
                "SELECT COALESCE(SUM(total),0) AS s FROM services WHERE account=?;",
                (account,),
            )["s"]
        )
        expenses = float(
            db_fetchone(
                "SELECT COALESCE(SUM(total),0) AS s FROM expenses WHERE account=?;",
                (account,),
            )["s"]
        )
        subaccounts[account] = {"balance": round(income - expenses, 2)}


    return {
        "month": {
            "id": m["id"],
            "year": year,
            "month": month,
            "is_closed": is_closed,
            "closed_at": closed_at,
            "closed_by": closed_by,
        },
        "monthly_min_needed": round(monthly_min_needed, 2),
        "sundays_count": sundays_count,
        "weekly_min_needed": weekly_min,

        "month_income_sum": round(income_sum, 2),
        "month_expenses_sum": round(expenses_sum, 2),

        "month_balance": month_balance,
        "start_balance": round(start_balance, 2),
        "fact_balance": fact_balance,

        "sddr": round(sddr, 2),
        "monthly_completion": float(monthly_completion),  # 0..1

        "psdpm": psdpm,  # float or None
        "avg_sunday": round(avg_sunday, 2),
        "subaccounts": subaccounts,
    }


def compute_year_analytics(year: int) -> Dict[str, Any]:
    months = db_fetchall("SELECT * FROM months WHERE year=? ORDER BY month ASC;", (year,))
    month_map = {int(m["month"]): m for m in months}

    month_items: List[Dict[str, Any]] = []
    totals = {
        "income": 0.0,
        "expenses": 0.0,
        "balance": 0.0,
        "min_needed": 0.0,
        "months_count": 0,
    }
    good_months: List[Dict[str, Any]] = []
    bad_months: List[Dict[str, Any]] = []

    for m in range(1, 13):
        row = month_map.get(m)
        if not row:
            month_items.append(
                {
                    "month": m,
                    "has_data": False,
                    "income": 0.0,
                    "expenses": 0.0,
                    "balance": 0.0,
                    "min_needed": 0.0,
                    "completion": 0.0,
                }
            )
            continue

        summary = compute_month_summary(int(row["id"]), ensure_tithe=True)
        income = float(summary["month_income_sum"])
        expenses = float(summary["month_expenses_sum"])
        balance = float(summary["month_balance"])
        min_needed = float(summary["monthly_min_needed"])
        completion = float(summary["monthly_completion"])

        totals["income"] += income
        totals["expenses"] += expenses
        totals["balance"] += balance
        totals["min_needed"] += min_needed
        totals["months_count"] += 1

        is_good = (min_needed > 0 and income >= min_needed) or (min_needed == 0 and balance >= 0)
        is_bad = (min_needed > 0 and income < min_needed) or balance < 0

        month_item = {
            "month": m,
            "month_id": int(row["id"]),
            "has_data": True,
            "income": round(income, 2),
            "expenses": round(expenses, 2),
            "balance": round(balance, 2),
            "min_needed": round(min_needed, 2),
            "completion": completion,
        }
        month_items.append(month_item)

        if is_good:
            good_months.append(
                {
                    "month": m,
                    "income": round(income, 2),
                    "balance": round(balance, 2),
                    "completion": completion,
                }
            )
        elif is_bad:
            bad_months.append(
                {
                    "month": m,
                    "income": round(income, 2),
                    "balance": round(balance, 2),
                    "completion": completion,
                }
            )

    totals["income"] = round(totals["income"], 2)
    totals["expenses"] = round(totals["expenses"], 2)
    totals["balance"] = round(totals["balance"], 2)
    totals["min_needed"] = round(totals["min_needed"], 2)
    totals["completion"] = (
        round(totals["income"] / totals["min_needed"], 4) if totals["min_needed"] > 0 else 0.0
    )

    prev_year = year - 1
    prev_totals = {
        "income": 0.0,
        "expenses": 0.0,
        "balance": 0.0,
        "min_needed": 0.0,
        "months_count": 0,
    }
    prev_months = db_fetchall("SELECT id FROM months WHERE year=?;", (prev_year,))
    for row in prev_months:
        summary = compute_month_summary(int(row["id"]), ensure_tithe=True)
        prev_totals["income"] += float(summary["month_income_sum"])
        prev_totals["expenses"] += float(summary["month_expenses_sum"])
        prev_totals["balance"] += float(summary["month_balance"])
        prev_totals["min_needed"] += float(summary["monthly_min_needed"])
        prev_totals["months_count"] += 1

    prev_totals["income"] = round(prev_totals["income"], 2)
    prev_totals["expenses"] = round(prev_totals["expenses"], 2)
    prev_totals["balance"] = round(prev_totals["balance"], 2)
    prev_totals["min_needed"] = round(prev_totals["min_needed"], 2)
    prev_totals["completion"] = (
        round(prev_totals["income"] / prev_totals["min_needed"], 4)
        if prev_totals["min_needed"] > 0
        else 0.0
    )

    def ratio(cur: float, prev: float) -> Optional[float]:
        if prev == 0:
            return None
        return (cur - prev) / prev

    yoy = {
        "income": ratio(totals["income"], prev_totals["income"]),
        "expenses": ratio(totals["expenses"], prev_totals["expenses"]),
        "balance": ratio(totals["balance"], prev_totals["balance"]),
    }

    return {
        "year": year,
        "months": month_items,
        "totals": totals,
        "prev_year": {"year": prev_year, "totals": prev_totals},
        "yoy": yoy,
        "good_months": good_months,
        "bad_months": bad_months,
    }

def compute_period_bounds(
    period_type: str,
    year: int,
    month: Optional[int] = None,
    quarter: Optional[int] = None,
) -> Tuple[dt.date, dt.date]:
    if period_type == "month":
        if not month:
            raise HTTPException(status_code=400, detail="month is required for type=month")
        m = int(month)
        if m < 1 or m > 12:
            raise HTTPException(status_code=400, detail="month must be 1..12")
        start = dt.date(year, m, 1)
        end = dt.date(year, m, calendar.monthrange(year, m)[1])
        return start, end
    if period_type == "quarter":
        if not quarter:
            raise HTTPException(status_code=400, detail="quarter is required for type=quarter")
        q = int(quarter)
        if q < 1 or q > 4:
            raise HTTPException(status_code=400, detail="quarter must be 1..4")
        start_month = (q - 1) * 3 + 1
        end_month = start_month + 2
        start = dt.date(year, start_month, 1)
        end = dt.date(year, end_month, calendar.monthrange(year, end_month)[1])
        return start, end
    if period_type == "year":
        start = dt.date(year, 1, 1)
        end = dt.date(year, 12, 31)
        return start, end
    raise HTTPException(status_code=400, detail="type must be month, quarter, or year")


def previous_period(period_type: str, year: int, month: Optional[int], quarter: Optional[int]) -> Tuple[dt.date, dt.date]:
    if period_type == "month":
        m = int(month or 1)
        y = int(year)
        if m == 1:
            y -= 1
            m = 12
        else:
            m -= 1
        return compute_period_bounds("month", y, month=m)
    if period_type == "quarter":
        q = int(quarter or 1)
        y = int(year)
        if q == 1:
            y -= 1
            q = 4
        else:
            q -= 1
        return compute_period_bounds("quarter", y, quarter=q)
    return compute_period_bounds("year", int(year) - 1)


def compute_period_totals(start: dt.date, end: dt.date) -> Dict[str, float]:
    income = float(
        db_fetchone(
            """
            SELECT COALESCE(SUM(total),0) AS s
            FROM services
            WHERE service_date BETWEEN ? AND ? AND account='main';
            """,
            (iso_date(start), iso_date(end)),
        )["s"]
    )
    expenses = float(
        db_fetchone(
            """
            SELECT COALESCE(SUM(total),0) AS s
            FROM expenses
            WHERE expense_date BETWEEN ? AND ? AND account='main';
            """,
            (iso_date(start), iso_date(end)),
        )["s"]
    )
    net = income - expenses
    return {
        "income": round(income, 2),
        "expenses": round(expenses, 2),
        "net": round(net, 2),
    }


def compute_period_analytics(
    period_type: str,
    year: int,
    month: Optional[int],
    quarter: Optional[int],
    include_top_categories: bool,
    top_limit: int = 5,
) -> Dict[str, Any]:
    start, end = compute_period_bounds(period_type, year, month=month, quarter=quarter)
    prev_start, prev_end = previous_period(period_type, year, month=month, quarter=quarter)

    totals = compute_period_totals(start, end)
    prev_totals = compute_period_totals(prev_start, prev_end)

    def ratio(cur: float, prev: float) -> Optional[float]:
        if prev <= 0:
            return None
        return (cur - prev) / prev

    delta_income = totals["income"] - prev_totals["income"]
    delta_expenses = totals["expenses"] - prev_totals["expenses"]
    delta_net = totals["net"] - prev_totals["net"]

    delta = {
        "income_abs": round(delta_income, 2),
        "income_pct": ratio(totals["income"], prev_totals["income"]),
        "expenses_abs": round(delta_expenses, 2),
        "expenses_pct": ratio(totals["expenses"], prev_totals["expenses"]),
        "net_abs": round(delta_net, 2),
        "net_pct": ratio(totals["net"], prev_totals["net"]),
    }

    top_expenses: List[Dict[str, Any]] = []
    if include_top_categories:
        rows = db_fetchall(
            """
            SELECT category, COALESCE(SUM(total),0) AS s
            FROM expenses
            WHERE expense_date BETWEEN ? AND ? AND account='main'
            GROUP BY category
            ORDER BY s DESC, category ASC
            LIMIT ?;
            """,
            (iso_date(start), iso_date(end), int(top_limit)),
        )
        top_expenses = [{"category": r["category"], "sum": round(float(r["s"]), 2)} for r in rows]

    return {
        "period": {"type": period_type, "start": iso_date(start), "end": iso_date(end)},
        "totals": totals,
        "prev": {
            "start": iso_date(prev_start),
            "end": iso_date(prev_end),
            "income": prev_totals["income"],
            "expenses": prev_totals["expenses"],
            "net": prev_totals["net"],
        },
        "delta": delta,
        "top_expenses_by_category": top_expenses,
    }


# ---------------------------
# Audit
# ---------------------------

def log_audit(
    user_id: Optional[int],
    action: str,
    entity_type: str,
    entity_id: Optional[int],
    before: Optional[Dict[str, Any]],
    after: Optional[Dict[str, Any]],
) -> None:
    now = iso_now(CFG.tzinfo())
    db_exec(
        """
        INSERT INTO audit_log (user_id, action, entity_type, entity_id, before_json, after_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?);
        """,
        (
            user_id,
            action,
            entity_type,
            entity_id,
            json.dumps(before, ensure_ascii=False) if before is not None else None,
            json.dumps(after, ensure_ascii=False) if after is not None else None,
            now,
        ),
    )

# ---------------------------
# Monitoring logs
# ---------------------------

def log_system_log(
    level: str,
    source: str,
    message: str,
    details: Optional[Dict[str, Any]] = None,
    trace: Optional[str] = None,
) -> None:
    now = iso_now(CFG.tzinfo())
    db_exec(
        """
        INSERT INTO system_logs (level, source, message, details_json, trace, created_at)
        VALUES (?, ?, ?, ?, ?, ?);
        """,
        (
            level,
            source,
            message,
            json.dumps(details, ensure_ascii=False) if details is not None else None,
            trace,
            now,
        ),
    )


def log_job_run(
    job_id: str,
    status: str,
    started_at: str,
    finished_at: str,
    duration_ms: int,
    error: Optional[str],
) -> None:
    db_exec(
        """
        INSERT INTO job_runs (job_id, status, started_at, finished_at, duration_ms, error)
        VALUES (?, ?, ?, ?, ?, ?);
        """,
        (job_id, status, started_at, finished_at, int(duration_ms), error),
    )


def log_message_delivery(
    kind: str,
    recipient_id: int,
    status: str,
    error: Optional[str],
) -> None:
    now = iso_now(CFG.tzinfo())
    db_exec(
        """
        INSERT INTO message_deliveries (kind, recipient_id, status, error, created_at)
        VALUES (?, ?, ?, ?, ?);
        """,
        (kind, int(recipient_id), status, error, now),
    )


# ---------------------------
# API Auth + roles
# ---------------------------

class AuthTelegramIn(BaseModel):
    initData: str = Field(..., description="Telegram WebApp initData string")


class AuthOut(BaseModel):
    token: str
    user: Dict[str, Any]


def get_bearer_token(request: Request) -> str:
    h = request.headers.get("Authorization", "").strip()
    if not h.lower().startswith("bearer "):
        token = request.query_params.get("token", "").strip()
        if token:
            return token
        raise HTTPException(status_code=401, detail="Missing Authorization Bearer token")
    return h.split(" ", 1)[1].strip()


def get_current_user(request: Request) -> sqlite3.Row:
    token = get_bearer_token(request)
    payload = verify_session_token(token, CFG.SESSION_SECRET)
    telegram_id = int(payload.get("telegram_id", 0))
    if not telegram_id:
        raise HTTPException(status_code=401, detail="Invalid token payload")

    # обновляем allowlist, чтобы роли в БД не отставали от users.json
    refresh_allowlist_if_needed()
    u = db_fetchone("SELECT * FROM users WHERE telegram_id=?;", (telegram_id,))
    if not u or int(u["active"]) != 1:
        raise HTTPException(status_code=403, detail="User not allowed / inactive")
    return u


def require_role(*allowed_roles: str):
    def _dep(u: sqlite3.Row = Depends(get_current_user)) -> sqlite3.Row:
        role = str(u["role"])
        if role not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient role")
        return u
    return _dep


# ---------------------------
# API Models (requests)
# ---------------------------

class MonthCreateIn(BaseModel):
    year: int
    month: int
    monthly_min_needed: float = 0.0
    start_balance: Optional[float] = None
    sundays_override: Optional[int] = Field(None, description="Admin override: 4 or 5 (optional)")


class MonthUpdateIn(BaseModel):
    monthly_min_needed: Optional[float] = None
    start_balance: Optional[float] = None
    sundays_override: Optional[int] = None


class ServiceIn(BaseModel):
    service_date: dt.date
    cashless: float = 0.0
    cash: float = 0.0
    income_type: str = "donation"
    account: str = "main"


class ExpenseIn(BaseModel):
    expense_date: dt.date
    category: str
    title: str
    qty: float = 1.0
    unit_amount: float = 0.0
    comment: Optional[str] = None
    tags: Optional[List[str]] = None
    account: str = "main"

class MonthBudgetIn(BaseModel):
    category_id: int
    limit_amount: float
    warn_threshold: Optional[float] = 0.9
    include_system: Optional[bool] = True


class SettingsUpdateIn(BaseModel):
    report_chat_id: Optional[int] = None
    sunday_report_time: Optional[str] = None  # "HH:MM"
    month_report_time: Optional[str] = None   # "HH:MM"
    timezone: Optional[str] = None
    ui_theme: Optional[str] = None
    daily_expenses_enabled: Optional[bool] = None


class CategoryCreateIn(BaseModel):
    name: str


class CategoryUpdateIn(BaseModel):
    name: Optional[str] = None
    is_active: Optional[bool] = None
    sort_order: Optional[int] = None


class CategoryAliasCreateIn(BaseModel):
    alias: str


class CategoryMergeIn(BaseModel):
    target_id: int
    source_ids: List[int]


class CategoryRenameMassIn(BaseModel):
    from_: str = Field(..., alias="from")
    to: str

    model_config = ConfigDict(validate_by_name=True)


# ---------------------------
# Telegram Bot (aiogram)
# ---------------------------

bot: Optional[Bot] = None
dp: Optional[Dispatcher] = None
router = Router()

# Pending confirmations (in-memory)
PENDING: Dict[int, Dict[str, Any]] = {}  # telegram_id -> payload


def is_allowed_telegram_user(telegram_id: int) -> bool:
    allow = refresh_allowlist_if_needed()
    u = allow.get(int(telegram_id))
    return bool(u and u.get("active") is True)


def get_user_role_from_db(telegram_id: int) -> str:
    row = db_fetchone("SELECT role, active FROM users WHERE telegram_id=?;", (telegram_id,))
    if not row or int(row["active"]) != 1:
        return "none"
    return str(row["role"])


def get_persistent_menu_url(role: str) -> Optional[str]:
    if role == "cash_signer":
        return cashapp_webapp_url()
    if role in ("admin", "accountant", "viewer"):
        return require_https_webapp_url(CFG.WEBAPP_URL)
    return None


async def configure_persistent_menu(chat_id: int, role: str) -> None:
    if not bot:
        return
    url = get_persistent_menu_url(role)
    try:
        if url:
            await bot.set_chat_menu_button(
                chat_id=chat_id,
                menu_button=MenuButtonWebApp(text="Бухгалтерия", web_app=WebAppInfo(url=url)),
            )
        else:
            await bot.set_chat_menu_button(chat_id=chat_id, menu_button=MenuButtonDefault())
    except Exception:
        pass


def main_menu_kb(role: str) -> InlineKeyboardMarkup:
    # Для администратора и подписанта используем только постоянную кнопку
    # в нижнем меню Telegram, без дублирования кнопок в чате.
    if role in ("admin", "cash_signer"):
        return InlineKeyboardMarkup(inline_keyboard=[])

    buttons = [
        [
            InlineKeyboardButton(text="Быстрый ввод пожертвования", callback_data="quick:donation"),
            InlineKeyboardButton(text="Быстрый ввод расхода", callback_data="quick:expense"),
        ],
        [InlineKeyboardButton(text="Отчёты", callback_data="menu:reports")],
    ]
    if role == "admin":
        buttons.append([InlineKeyboardButton(text="Настройки", callback_data="menu:settings")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def confirm_kb(kind: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Сохранить", callback_data=f"confirm:save:{kind}"),
                InlineKeyboardButton(text="❌ Отмена", callback_data=f"confirm:cancel:{kind}"),
            ]
        ]
    )


def reports_kb() -> InlineKeyboardMarkup:
    webapp_url = require_https_webapp_url(CFG.WEBAPP_URL)
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Отчёт за текущее воскресенье", callback_data="report:sunday")],
            [InlineKeyboardButton(text="Расходы за текущий месяц", callback_data="report:month_expenses")],
            [InlineKeyboardButton(text="Итоги месяца", callback_data="report:month_summary")],
            *(
                [[InlineKeyboardButton(text="Открыть дашборд", web_app=WebAppInfo(url=webapp_url))]]
                if webapp_url
                else []
            ),
        ]
    )

    def webapp_url_with_screen(screen: str) -> Optional[str]:
        url = require_https_webapp_url(CFG.WEBAPP_URL)
        if not url:
            return None
        try:
            parsed = urllib.parse.urlparse(url)
            query = urllib.parse.parse_qs(parsed.query)
            query["screen"] = [screen]
            new_query = urllib.parse.urlencode(query, doseq=True)
            return urllib.parse.urlunparse(parsed._replace(query=new_query))
        except Exception:
            sep = "&" if "?" in url else "?"
            return f"{url}{sep}screen={screen}"

def parse_quick_input(text: str) -> Optional[Dict[str, Any]]:
    """
    Форматы из ТЗ:
      "пож 8500 4800"   -> cashless=8500, cash=4800
      "расход 2500 зал" -> unit_amount=2500, category="зал", title="зал"
    """
    t = text.strip()
    low = t.lower()

    if low.startswith("пож"):
        parts = t.split()
        if len(parts) < 2:
            return None
        cashless = float(parts[1].replace(",", "."))
        cash = float(parts[2].replace(",", ".")) if len(parts) >= 3 else 0.0
        return {"kind": "donation", "cashless": cashless, "cash": cash}

    if low.startswith("расход"):
        parts = t.split()
        if len(parts) < 2:
            return None
        amount = float(parts[1].replace(",", "."))
        tail = " ".join(parts[2:]).strip() if len(parts) >= 3 else "Прочее"
        category = tail if tail else "Прочее"
        title = tail if tail else "Расход"
        return {"kind": "expense", "unit_amount": amount, "category": category, "title": title}

    return None


def last_sunday(today: dt.date) -> dt.date:
    # Sunday is weekday=6
    delta = (today.weekday() - 6) % 7
    return today - dt.timedelta(days=delta)


def format_telegram_exception(exc: Exception) -> str:
    msg = str(exc)
    lower = msg.lower()
    hint = ""
    if "chat not found" in lower:
        hint = "Чат не найден. Проверьте chat_id и добавьте бота в чат/канал или нажмите /start."
    elif "bot was blocked by the user" in lower:
        hint = "Бот заблокирован пользователем. Нужно разблокировать и снова нажать /start."
    elif "not enough rights" in lower or "administrator rights" in lower:
        hint = "У бота нет прав в чате. Проверьте, что бот добавлен и имеет доступ к отправке сообщений."
    return f"{msg}. {hint}".strip()


async def bot_send_safe(
    chat_id: int,
    text: str,
    reply_markup: Optional[InlineKeyboardMarkup] = None,
) -> Tuple[bool, Optional[str]]:
    if not bot:
        return False, "Bot is not initialized"
    try:
        await bot.send_message(chat_id=chat_id, text=text, reply_markup=reply_markup)
        return True, None
    except Exception as exc:
        # avoid crashing scheduler/bot
        msg = format_telegram_exception(exc)
        print("Failed to send message:", msg)
        return False, msg


async def bot_send_or_http_error(
    chat_id: int,
    text: str,
    reply_markup: Optional[InlineKeyboardMarkup] = None,
) -> None:
    if not bot:
        raise HTTPException(status_code=503, detail="Bot is not initialized")
    try:
        await bot.send_message(chat_id=chat_id, text=text, reply_markup=reply_markup)
    except TelegramForbiddenError as exc:
        raise HTTPException(status_code=403, detail=format_telegram_exception(exc))
    except TelegramBadRequest as exc:
        raise HTTPException(status_code=400, detail=format_telegram_exception(exc))
    except Exception as exc:
        raise HTTPException(status_code=502, detail=format_telegram_exception(exc))


async def ensure_report_chat_reachable(chat_id: int) -> None:
    if not bot:
        raise HTTPException(status_code=503, detail="Bot is not initialized")
    try:
        await bot.get_chat(chat_id)
    except TelegramForbiddenError as exc:
        raise HTTPException(status_code=403, detail=format_telegram_exception(exc))
    except TelegramBadRequest as exc:
        raise HTTPException(status_code=400, detail=format_telegram_exception(exc))
    except Exception as exc:
        raise HTTPException(status_code=502, detail=format_telegram_exception(exc))


async def send_report_to_recipients(
        text: str,
        reply_markup: Optional[InlineKeyboardMarkup],
        recipients: List[int],
        raise_on_error: bool,
        kind: str = "report",
) -> None:
    if not recipients:
        if raise_on_error:
            raise HTTPException(status_code=400, detail="No report recipients configured")
        return
    errors: List[str] = []
    for chat_id in recipients:
        safe_markup = reply_markup if is_allowed_telegram_user(chat_id) else None
        try:
            if raise_on_error:
                await bot_send_or_http_error(chat_id, text, safe_markup)
                log_message_delivery(kind, chat_id, "success", None)
            else:
                ok, err = await bot_send_safe(chat_id, text, safe_markup)
                if ok:
                    log_message_delivery(kind, chat_id, "success", None)
                else:
                    log_message_delivery(kind, chat_id, "fail", err)
        except HTTPException as exc:
            log_message_delivery(kind, chat_id, "fail", str(exc.detail))
            errors.append(f"{chat_id}: {exc.detail}")
    if errors and raise_on_error:
        raise HTTPException(status_code=502, detail="; ".join(errors))


async def send_report_png_to_recipients(
        png_bytes: bytes,
        filename: str,
        caption: str,
        recipients: List[int],
        raise_on_error: bool,
        kind: str = "report_png",
) -> None:
    if not recipients:
        if raise_on_error:
            raise HTTPException(status_code=400, detail="No report recipients configured")
        return
    if not bot:
        if raise_on_error:
            raise HTTPException(status_code=503, detail="Bot is not initialized")
        return

    errors: List[str] = []
    for chat_id in recipients:
        try:
            payload = BufferedInputFile(png_bytes, filename=filename)
            await bot.send_document(chat_id=chat_id, document=payload, caption=caption)
            log_message_delivery(kind, chat_id, "success", None)
        except Exception as exc:
            msg = format_telegram_exception(exc)
            log_message_delivery(kind, chat_id, "fail", msg)
            if raise_on_error:
                errors.append(f"{chat_id}: {msg}")
    if errors and raise_on_error:
        raise HTTPException(status_code=502, detail="; ".join(errors))


@router.message(Command("start"))
async def on_start(m: Message):
    tid = m.from_user.id if m.from_user else 0
    if not tid or not is_allowed_telegram_user(tid):
        await m.answer(
            "Доступ запрещён. Ваш Telegram ID не в allowlist.\n"
            f"Ваш Telegram ID: {tid}\n"
            "Добавьте его в users.json и повторите /start."
        )
        return
    if tid:
        register_bot_subscriber(tid)

    role = get_user_role_from_db(tid)
    await configure_persistent_menu(m.chat.id, role)
    await m.answer(
        "Меню бухгалтерии:",
        reply_markup=main_menu_kb(role),
    )
    webapp_url = require_https_webapp_url(CFG.WEBAPP_URL)
    cashapp_url = cashapp_webapp_url()
    if not webapp_url or (role == "cash_signer" and not cashapp_url):
        await m.answer(
            "Внимание: WebApp-кнопки доступны только по HTTPS.\n"
            "Настройте публичный HTTPS-домен и задайте APP_URL/WEBAPP_URL в .env."
        )


@router.message(F.text)
async def on_text(m: Message):
    if not m.from_user or not m.text:
        return
    tid = m.from_user.id
    if not is_allowed_telegram_user(tid):
        return

    parsed = parse_quick_input(m.text)
    if not parsed:
        return

    tz = CFG.tzinfo()
    today = dt.datetime.now(tz).date()

    if parsed["kind"] == "donation":
        s_date = last_sunday(today)
        PENDING[tid] = {
            "kind": "donation",
            "service_date": s_date.isoformat(),
            "cashless": float(parsed["cashless"]),
            "cash": float(parsed["cash"]),
        }
        await m.answer(
            f"Проверить пожертвование:\n"
            f"Дата: {s_date.strftime('%d.%m.%Y')}\n"
            f"Безнал: {parsed['cashless']:.2f}\n"
            f"Наличные: {parsed['cash']:.2f}\n"
            f"Итого: {(parsed['cashless']+parsed['cash']):.2f}",
            reply_markup=confirm_kb("donation"),
        )
        return

    if parsed["kind"] == "expense":
        e_date = today
        PENDING[tid] = {
            "kind": "expense",
            "expense_date": e_date.isoformat(),
            "category": parsed["category"],
            "title": parsed["title"],
            "qty": 1.0,
            "unit_amount": float(parsed["unit_amount"]),
            "comment": None,
        }
        await m.answer(
            f"Проверить расход:\n"
            f"Дата: {e_date.strftime('%d.%m.%Y')}\n"
            f"Категория: {parsed['category']}\n"
            f"Название: {parsed['title']}\n"
            f"Сумма: {parsed['unit_amount']:.2f}",
            reply_markup=confirm_kb("expense"),
        )
        return


@router.callback_query(F.data.startswith("quick:"))
async def on_quick(cq: CallbackQuery):
    if not cq.from_user:
        return
    tid = cq.from_user.id
    if not is_allowed_telegram_user(tid):
        await cq.answer("Нет доступа", show_alert=True)
        return

    kind = cq.data.split(":", 1)[1]
    if kind == "donation":
        await cq.message.answer("Отправьте сообщением: `пож 8500 4800` (безнал нал)", parse_mode="Markdown")
        await cq.answer()
        return
    if kind == "expense":
        await cq.message.answer("Отправьте сообщением: `расход 2500 зал`", parse_mode="Markdown")
        await cq.answer()
        return

    await cq.answer()


@router.callback_query(F.data.startswith("menu:reports"))
async def on_reports_menu(cq: CallbackQuery):
    if not cq.from_user:
        return
    tid = cq.from_user.id
    if not is_allowed_telegram_user(tid):
        await cq.answer("Нет доступа", show_alert=True)
        return
        await cq.message.answer("Отчёты:", reply_markup=reports_kb())
        await cq.answer()

    @router.callback_query(F.data.startswith("menu:settings"))
    async def on_settings_menu(cq: CallbackQuery):
        if not cq.from_user:
            return
        tid = cq.from_user.id
        if not is_allowed_telegram_user(tid):
            await cq.answer("Нет доступа", show_alert=True)
            return

        role = get_user_role_from_db(tid)
        if role != "admin":
            await cq.answer("Нет доступа", show_alert=True)
            return

        settings_url = webapp_url_with_screen("settings")
        inline_keyboard = []
        if settings_url:
            inline_keyboard.append(
                [InlineKeyboardButton(text="Открыть настройки (WebApp)",
                                      web_app=WebAppInfo(url=settings_url))]
            )
        await cq.message.answer(
            "Настройки:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=inline_keyboard),
        )
        await cq.answer()

    @router.callback_query(F.data.startswith("report:"))
    async def on_report_actions(cq: CallbackQuery):
        if not cq.from_user:
            return
        tid = cq.from_user.id
        if not is_allowed_telegram_user(tid):
            await cq.answer("Нет доступа", show_alert=True)
            return


    role = get_user_role_from_db(tid)
    if role not in ("admin", "accountant", "viewer"):
        await cq.answer("Нет роли", show_alert=True)
        return

    tz = CFG.tzinfo()
    today = dt.datetime.now(tz).date()
    action = cq.data.split(":", 1)[1]

    if action == "sunday":
        text, kb = build_sunday_report_text(today)
        await cq.message.answer(text, reply_markup=kb)
        await cq.answer()
        return

    if action == "month_expenses":
        text, kb = build_month_expenses_report_text(today)
        await cq.message.answer(text, reply_markup=kb)
        await cq.answer()
        return

    if action == "month_summary":
        m = get_or_create_month(today.year, today.month)
        summary = compute_month_summary(int(m["id"]), ensure_tithe=True)
        text = format_month_summary_text(summary)
        await cq.message.answer(text)
        await cq.answer()
        return

    await cq.answer()


@router.callback_query(F.data.startswith("confirm:"))
async def on_confirm(cq: CallbackQuery):
    if not cq.from_user:
        return
    tid = cq.from_user.id
    if not is_allowed_telegram_user(tid):
        await cq.answer("Нет доступа", show_alert=True)
        return

    parts = cq.data.split(":")
    if len(parts) < 3:
        await cq.answer()
        return
    action = parts[1]  # save / cancel
    kind = parts[2]

    pending = PENDING.get(tid)
    if not pending or pending.get("kind") != kind:
        await cq.answer("Нет данных для сохранения", show_alert=True)
        return

    if action == "cancel":
        PENDING.pop(tid, None)
        await cq.message.answer("Отменено.")
        await cq.answer()
        return

    # save
    user_row = db_fetchone("SELECT * FROM users WHERE telegram_id=?;", (tid,))
    if not user_row or int(user_row["active"]) != 1:
        await cq.answer("Пользователь не активен", show_alert=True)
        return
    user_id = int(user_row["id"])
    role = str(user_row["role"])
    if role not in ("admin", "accountant"):
        await cq.answer("Недостаточно прав для сохранения", show_alert=True)
        return

    tz = CFG.tzinfo()
    now = iso_now(tz)

    if kind == "donation":
        s_date = parse_iso_date(pending["service_date"])
        m = get_or_create_month(s_date.year, s_date.month)
        month_id = int(m["id"])

        before = db_fetchone(
            """
            SELECT * FROM services
            WHERE month_id=? AND service_date=? AND account='main' AND income_type='donation';
            """,
            (month_id, pending["service_date"]),
        )
        cashless = float(pending["cashless"])
        cash = float(pending["cash"])
        total = round(cashless + cash, 2)
        if cash > 0:
            request_id = create_cashflow_collect_request_if_needed(
                account="main",
                cash_amount=cash,
                cashless_amount=cashless,
                income_type="donation",
                month_id=month_id,
                service_date=pending["service_date"],
                created_by_telegram_id=int(tid),
            )
            if request_id:
                PENDING.pop(tid, None)
                await cq.message.answer(
                    f"Создан запрос на подтверждение наличных №{request_id}. "
                    "Внесение возможно после подписей."
                )
                await cq.answer()
                return

        # upsert service
        if before:
            db_exec(
                """
                UPDATE services
                SET cashless=?, cash=?, total=?, account='main', updated_at=?
                WHERE id=?;
                """,
                (cashless, cash, total, now, before["id"]),
            )
            after = db_fetchone("SELECT * FROM services WHERE id=?;", (before["id"],))
            log_audit(user_id, "UPDATE", "service", int(before["id"]), dict(before), dict(after) if after else None)
        else:
            new_id = db_exec_returning_id(
                """
                INSERT INTO services (
                    month_id, service_date, idx, cashless, cash, total,
                    weekly_min_needed, mnsps_status, pvs_ratio,
                    created_at, updated_at
                ) VALUES (?, ?, 1, ?, ?, ?, 0, 'Не собрана', 0, 'main', 'donation', ?, ?);
                """,
                (month_id, pending["service_date"], cashless, cash, total, now, now),
            )
            after = db_fetchone("SELECT * FROM services WHERE id=?;", (new_id,))
            log_audit(user_id, "CREATE", "service", int(new_id), None, dict(after) if after else None)

        # Recalc + tithe
        recalc_services_for_month(month_id)
        ensure_tithe_expense(month_id, user_id=user_id)

        PENDING.pop(tid, None)
        await cq.message.answer("✅ Пожертвование сохранено.")
        await cq.answer()
        return

    if kind == "expense":
        e_date = parse_iso_date(pending["expense_date"])
        m = get_or_create_month(e_date.year, e_date.month)
        month_id = int(m["id"])
        category = resolve_category(pending["category"], role, user_id)

        new_id = db_exec_returning_id(
            """
            INSERT INTO expenses (
                month_id, expense_date, category, title, qty, unit_amount, total, comment,
                is_system, account, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 'main', ?, ?);
            """,
            (
                month_id,
                pending["expense_date"],
                category,
                pending["title"],
                float(pending.get("qty", 1.0)),
                float(pending["unit_amount"]),
                round(float(pending.get("qty", 1.0)) * float(pending["unit_amount"]), 2),
                pending.get("comment"),
                now,
                now,
            ),
        )
        after = db_fetchone("SELECT * FROM expenses WHERE id=?;", (new_id,))
        log_audit(user_id, "CREATE", "expense", int(new_id), None, dict(after) if after else None)

        # expenses changed -> recompute summary (tithe depends on income, но пусть живёт; ensure anyway)
        ensure_tithe_expense(month_id, user_id=user_id)

        PENDING.pop(tid, None)
        await cq.message.answer("✅ Расход сохранён.")
        await cq.answer()
        return

    await cq.answer()


# ---------------------------
# Report builders (Telegram text)
# ---------------------------

def fmt_money(x: float) -> str:
    # 41 502.16 style
    s = f"{x:,.2f}"
    s = s.replace(",", " ").replace(".00", ".00")
    return s


def fmt_money_commas(x: float) -> str:
    s = f"{x:,.2f}"
    return s.replace(".00", "")


def fmt_percent_1(x: float) -> str:
    return f"{x * 100:.1f}%"


def build_sunday_report_text(today: dt.date) -> Tuple[str, Optional[InlineKeyboardMarkup]]:
    tz = CFG.tzinfo()
    s_date = last_sunday(today)
    m = get_or_create_month(s_date.year, s_date.month)
    month_id = int(m["id"])

    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=None)

    service = db_fetchone(
        """
        SELECT * FROM services
        WHERE month_id=? AND service_date=? AND (income_type='donation' OR income_type IS NULL);
        """,
        (month_id, s_date.isoformat()),
    )
    cashless = float(service["cashless"]) if service else 0.0
    cash = float(service["cash"]) if service else 0.0
    total = float(service["total"]) if service else 0.0
    weekly_min = float(service["weekly_min_needed"]) if service else calc_weekly_min_needed(m)
    status = str(service["mnsps_status"]) if service else ("Собрана" if (weekly_min and total >= weekly_min) else "Не собрана")
    pvs = float(service["pvs_ratio"]) if service else ((total / weekly_min) if weekly_min else 0.0)

    summary = compute_month_summary(month_id, ensure_tithe=True)

    title = f"<b>Отчёт по пожертвованиям — {s_date.strftime('%d.%m.%Y')}</b>"
    block1 = (
        f"\n\n<b>Пожертвования</b>\n"
        f"• Безнал: <b>{fmt_money(cashless)}</b>\n"
        f"• Наличные: <b>{fmt_money(cash)}</b>\n"
        f"• Итого: <b>{fmt_money(total)}</b>"
    )
    block2 = (
        f"\n\n<b>МНСП на воскресенье</b>\n"
        f"• Минимум: <b>{fmt_money(weekly_min)}</b>\n"
        f"• Статус: <b>{status}</b>\n"
        f"• ПВС: <b>{fmt_percent_1(pvs)}</b>"
    )
    sddr = float(summary["sddr"])
    sddr_text = fmt_money(sddr) if sddr > 0 else "Нет суммы"

    block3 = (
        f"\n\n<b>Месяц на текущую дату</b>\n"
        f"• Итого доход: <b>{fmt_money(float(summary['month_income_sum']))}</b>\n"
        f"• Выполнение МНСП: <b>{fmt_percent_1(float(summary['monthly_completion']))}</b>\n"
        f"• СДДР: <b>{sddr_text}</b>"
    )

    return title + block1 + block2 + block3, None


def build_month_expenses_report_text(today: dt.date) -> Tuple[str, Optional[InlineKeyboardMarkup]]:
    m = get_or_create_month(today.year, today.month)
    month_id = int(m["id"])

    ensure_tithe_expense(month_id, user_id=None)
    summary = compute_month_summary(month_id, ensure_tithe=True)

    rows = db_fetchall(
        """
        SELECT category, SUM(total) AS s
        FROM expenses
        WHERE month_id=? AND account='main'
        GROUP BY category
        ORDER BY s DESC
        LIMIT 3;
        """,
        (month_id,),
    )
    top = "\n".join([f"• {r['category']}: <b>{fmt_money(float(r['s']))}</b>" for r in rows]) or "—"

    last5 = db_fetchall(
        """
        SELECT expense_date, title, category, total
        FROM expenses
        WHERE month_id=? AND account='main'
        ORDER BY expense_date DESC, id DESC
        LIMIT 5;
        """,
        (month_id,),
    )
    last_lines = []
    for r in last5:
        d = parse_iso_date(str(r["expense_date"]))
        last_lines.append(f"• {d.strftime('%d.%m')}: {r['title']} — <b>{fmt_money(float(r['total']))}</b> ({r['category']})")
    last_block = "\n".join(last_lines) or "—"

    month_name = calendar.month_name[int(m["month"])]
    title = f"<b>Расходы — {month_name} {int(m['year'])}</b>"

    body = (
        f"\n\n<b>Сумма расходов</b>\n"
        f"• Итого: <b>{fmt_money(float(summary['month_expenses_sum']))}</b>\n\n"
        f"<b>Топ-категории</b>\n{top}\n\n"
        f"<b>Последние 5 расходов</b>\n{last_block}\n\n"
        f"<b>Балансы</b>\n"
        f"• Баланс месяца: <b>{fmt_money(float(summary['month_balance']))}</b>\n"
        f"• Факт. баланс: <b>{fmt_money(float(summary['fact_balance']))}</b>"
    )

    return title + body, None


def format_month_summary_text(summary: Dict[str, Any]) -> str:
    y = summary["month"]["year"]
    m = summary["month"]["month"]
    month_name = calendar.month_name[int(m)]

    psdpm = summary["psdpm"]
    psdpm_text = f"{psdpm*100:.1f}%" if isinstance(psdpm, (int, float)) else "—"
    sddr = float(summary["sddr"])
    sddr_text = fmt_money(sddr) if sddr > 0 else "Нет суммы"

    return (
        f"<b>Итоги месяца — {month_name} {y}</b>\n\n"
        f"<b>Ключевые показатели</b>\n"
        f"• Доход: <b>{fmt_money(float(summary['month_income_sum']))}</b>\n"
        f"• Расход: <b>{fmt_money(float(summary['month_expenses_sum']))}</b>\n"
        f"• Баланс: <b>{fmt_money(float(summary['month_balance']))}</b>\n"
        f"• Факт. баланс: <b>{fmt_money(float(summary['fact_balance']))}</b>\n\n"
        f"<b>МНСП</b>\n"
        f"• МНСП месяц: <b>{fmt_money(float(summary['monthly_min_needed']))}</b>\n"
        f"• Выполнение: <b>{fmt_percent_1(float(summary['monthly_completion']))}</b>\n"
        f"• СДДР: <b>{sddr_text}</b>\n\n"
        f"<b>Сравнение</b>\n"
        f"• ПСДПМ: <b>{psdpm_text}</b>\n"
        f"• Среднее воскресенье: <b>{fmt_money(float(summary['avg_sunday']))}</b>"
    )


# ---------------------------
# Scheduler jobs
# ---------------------------

scheduler = AsyncIOScheduler()


def parse_hhmm(s: str, default_h: int, default_m: int) -> Tuple[int, int]:
    try:
        hh, mm = s.strip().split(":")
        h = int(hh)
        m = int(mm)
        if 0 <= h <= 23 and 0 <= m <= 59:
            return h, m
    except Exception:
        pass
    return default_h, default_m


def reschedule_jobs() -> None:
    # remove previous jobs
    for job_id in ("job_sunday_report", "job_daily_expenses", "job_backup_daily"):
        try:
            scheduler.remove_job(job_id)
        except Exception:
            pass

    s = get_settings()
    tz_name = str(s["timezone"] or CFG.TZ)
    tzinfo = ZoneInfo(tz_name)

    sunday_time = str(s["sunday_report_time"] or "18:00")
    h, m = parse_hhmm(sunday_time, 18, 0)

    # Every Sunday
    scheduler.add_job(
        func=lambda: asyncio.create_task(run_sunday_report_job()),
        trigger=CronTrigger(day_of_week="sun", hour=h, minute=m, timezone=tzinfo),
        id="job_sunday_report",
        replace_existing=True,
        misfire_grace_time=3600,
    )

    daily_enabled = int(s["daily_expenses_enabled"] or 0) == 1
    if daily_enabled:
        daily_time = str(s["month_report_time"] or "21:00")
        dh, dm = parse_hhmm(daily_time, 21, 0)
        scheduler.add_job(
            func=lambda: asyncio.create_task(run_daily_expenses_job()),
            trigger=CronTrigger(hour=dh, minute=dm, timezone=tzinfo),
            id="job_daily_expenses",
            replace_existing=True,
            misfire_grace_time=3600,
        )

    scheduler.add_job(
        func=lambda: asyncio.create_task(run_backup_job()),
        trigger=CronTrigger(hour=3, minute=30, timezone=tzinfo),
        id="job_backup_daily",
        replace_existing=True,
        misfire_grace_time=3600,
    )

async def run_job_with_logging(job_id: str, coro: Awaitable[None]) -> None:
    tzinfo = CFG.tzinfo()
    started_at = iso_now(tzinfo)
    started_ts = time.perf_counter()
    status = "success"
    error: Optional[str] = None
    try:
        await coro
    except Exception as exc:
        status = "fail"
        error = str(exc)
        trace = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        try:
            log_system_log(
                "ERROR",
                "scheduler",
                f"Job {job_id} failed",
                details={"error": str(exc)},
                trace=trace,
            )
        except Exception:
            pass
    finally:
        finished_at = iso_now(tzinfo)
        duration_ms = int((time.perf_counter() - started_ts) * 1000)
        try:
            log_job_run(job_id, status, started_at, finished_at, duration_ms, error)
        except Exception:
            pass




async def send_sunday_reports_bundle(
        today: dt.date,
        recipients: List[int],
        raise_on_error: bool,
) -> None:
    text, kb = build_sunday_report_text(today)
    await send_report_to_recipients(text, kb, recipients, raise_on_error=raise_on_error, kind="report")

    month_row = get_or_create_month(today.year, today.month)
    png_data, filename, month_meta = build_month_report_png(int(month_row["id"]), preset="landscape", pixel_ratio=2, dpi=192)
    caption = f"PNG-отчёт за {RU_MONTHS[int(month_meta['month']) - 1]} {int(month_meta['year'])}"
    await send_report_png_to_recipients(
        png_data,
        filename,
        caption,
        recipients,
        raise_on_error=raise_on_error,
        kind="report_png",
    )

async def run_sunday_report_job() -> None:
    async def _job() -> None:
        s = get_settings()
        recipients = list_report_recipients(s)
        if not recipients:
            return
        tzinfo = ZoneInfo(str(s["timezone"] or CFG.TZ))
        today = dt.datetime.now(tzinfo).date()
        await send_sunday_reports_bundle(today, recipients, raise_on_error=False)

    await run_job_with_logging("sunday_report", _job())

async def run_daily_expenses_job() -> None:
    async def _job() -> None:
        s = get_settings()
        recipients = list_report_recipients(s)
        if not recipients:
            return
        tzinfo = ZoneInfo(str(s["timezone"] or CFG.TZ))
        today = dt.datetime.now(tzinfo).date()
        text, kb = build_month_expenses_report_text(today)
        await send_report_to_recipients(text, kb, recipients, raise_on_error=False, kind="report")

    await run_job_with_logging("daily_expenses", _job())


# ---------------------------
# Backups
# ---------------------------

BACKUP_PATTERN = re.compile(r"^(db|full)_(\d{8}_\d{4})\.(sqlite3|zip)$")


def ensure_backups_dir() -> Path:
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    return BACKUPS_DIR


def backups_tzinfo() -> ZoneInfo:
    s = get_settings()
    return ZoneInfo(str(s["timezone"] or CFG.TZ))


def backup_timestamp(tzinfo: ZoneInfo) -> str:
    return dt.datetime.now(tzinfo).strftime("%Y%m%d_%H%M")


def parse_backup_timestamp(ts: str, tzinfo: ZoneInfo) -> dt.datetime:
    try:
        return dt.datetime.strptime(ts, "%Y%m%d_%H%M").replace(tzinfo=tzinfo)
    except Exception:
        return dt.datetime.now(tzinfo)


def list_backups(tzinfo: ZoneInfo) -> List[Dict[str, Any]]:
    ensure_backups_dir()
    items: List[Dict[str, Any]] = []
    for path in BACKUPS_DIR.iterdir():
        if not path.is_file():
            continue
        m = BACKUP_PATTERN.match(path.name)
        if not m:
            continue
        backup_type = m.group(1)
        ts = m.group(2)
        created_at = parse_backup_timestamp(ts, tzinfo)
        stat = path.stat()
        items.append(
            {
                "name": path.name,
                "type": backup_type,
                "size_bytes": stat.st_size,
                "created_at": created_at.isoformat(),
            }
        )
    items.sort(key=lambda x: x["created_at"], reverse=True)
    return items


def enforce_backup_retention(tzinfo: ZoneInfo) -> None:
    ensure_backups_dir()
    entries = []
    for path in BACKUPS_DIR.iterdir():
        if not path.is_file():
            continue
        m = BACKUP_PATTERN.match(path.name)
        if not m:
            continue
        backup_type = m.group(1)
        ts = m.group(2)
        created_at = parse_backup_timestamp(ts, tzinfo)
        entries.append({"path": path, "type": backup_type, "created_at": created_at})

    by_type: Dict[str, List[Dict[str, Any]]] = {"db": [], "full": []}
    for entry in entries:
        if entry["type"] in by_type:
            by_type[entry["type"]].append(entry)

    for backup_type, group in by_type.items():
        group.sort(key=lambda x: x["created_at"], reverse=True)
        for stale in group[14:]:
            try:
                stale["path"].unlink()
            except Exception:
                pass

    # total cap 30
    remaining = []
    for path in BACKUPS_DIR.iterdir():
        if not path.is_file():
            continue
        m = BACKUP_PATTERN.match(path.name)
        if not m:
            continue
        ts = m.group(2)
        created_at = parse_backup_timestamp(ts, tzinfo)
        remaining.append({"path": path, "created_at": created_at})

    remaining.sort(key=lambda x: x["created_at"])
    while len(remaining) > 30:
        stale = remaining.pop(0)
        try:
            stale["path"].unlink()
        except Exception:
            pass


def create_db_backup(tzinfo: ZoneInfo) -> Path:
    ensure_backups_dir()
    stamp = backup_timestamp(tzinfo)
    temp_path = BACKUPS_DIR / f".tmp_db_{stamp}.sqlite3"
    dest_path = BACKUPS_DIR / f"db_{stamp}.sqlite3"
    shutil.copy2(CFG.DB_PATH, temp_path)
    os.replace(temp_path, dest_path)
    enforce_backup_retention(tzinfo)
    return dest_path


def create_full_backup(tzinfo: ZoneInfo) -> Path:
    ensure_backups_dir()
    stamp = backup_timestamp(tzinfo)
    temp_db = BACKUPS_DIR / f".tmp_full_{stamp}.sqlite3"
    temp_zip = BACKUPS_DIR / f".tmp_full_{stamp}.zip"
    dest_zip = BACKUPS_DIR / f"full_{stamp}.zip"

    shutil.copy2(CFG.DB_PATH, temp_db)
    with zipfile.ZipFile(temp_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(temp_db, arcname="db.sqlite3")
        if UPLOADS_DIR.exists():
            for path in UPLOADS_DIR.rglob("*"):
                if path.is_file():
                    arcname = path.relative_to(BASE_DIR)
                    zf.write(path, arcname=str(arcname))
    try:
        temp_db.unlink()
    except Exception:
        pass
    os.replace(temp_zip, dest_zip)
    enforce_backup_retention(tzinfo)
    return dest_zip


def validate_sqlite_file(path: Path) -> bool:
    try:
        conn = sqlite3.connect(str(path))
        conn.execute("SELECT name FROM sqlite_master LIMIT 1;")
        conn.close()
        return True
    except Exception:
        return False


def safe_extract_zip(zf: zipfile.ZipFile, dest: Path) -> None:
    for info in zf.infolist():
        p = Path(info.filename)
        if p.is_absolute() or ".." in p.parts:
            raise HTTPException(status_code=400, detail="Invalid archive paths")
    zf.extractall(dest)


async def run_backup_job() -> None:
    try:
        tzinfo = backups_tzinfo()
        await asyncio.to_thread(create_full_backup, tzinfo)
    except Exception:
        traceback.print_exc()


# ---------------------------
# FastAPI app
# ---------------------------

@asynccontextmanager
async def lifespan(app: FastAPI):
    # init DB
    init_db()

    # load allowlist and sync to db
    allow = load_allowlist()
    sync_allowlist_to_db(allow)
    app.state.allowlist = allow
    global ALLOWLIST_CACHE, ALLOWLIST_MTIME
    ALLOWLIST_CACHE = allow
    try:
        ALLOWLIST_MTIME = os.path.getmtime(CFG.USERS_JSON_PATH)
    except FileNotFoundError:
        ALLOWLIST_MTIME = None

    # init bot + dp
    global bot, dp
    from aiogram.client.default import DefaultBotProperties

    bot = Bot(
        token=CFG.BOT_TOKEN,
        default=DefaultBotProperties(parse_mode="HTML")
    )
    import cashflow_bot

    cashflow_bot.set_bot(bot)


    dp = Dispatcher()
    dp.include_router(router)

    # scheduler
    if not scheduler.running:
        scheduler.start()
    reschedule_jobs()

    # start polling as background task
    polling_task = asyncio.create_task(dp.start_polling(bot))  # type: ignore[arg-type]

    try:
        yield
    finally:
        try:
            polling_task.cancel()
        except Exception:
            pass
        try:
            await bot.session.close()  # type: ignore[union-attr]
        except Exception:
            pass
        try:
            scheduler.shutdown(wait=False)
        except Exception:
            pass


APP = FastAPI(title="Church Accounting Bot", version="1.0.0", lifespan=lifespan)


def _build_cors_allow_origins() -> List[str]:
    origins: List[str] = ["http://localhost", "http://localhost:8000"]
    for candidate in (CFG.APP_URL, CFG.WEBAPP_URL):
        value = (candidate or "").strip()
        if not value or value == "*":
            continue
        if value not in origins:
            origins.append(value)
    return origins

APP.add_middleware(
    CORSMiddleware,
    allow_origins=_build_cors_allow_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from cashflow_routes import (
    router as cashflow_router,
    my_requests as cashflow_my_requests,
    withdraw_act as cashflow_withdraw_act,
    withdraw_act_xlsx as cashflow_withdraw_act_xlsx,
)

APP.include_router(cashflow_router)

def _ensure_cashflow_routes() -> None:
    existing_paths = {route.path for route in APP.routes}
    if "/api/cashflow/requests/my" not in existing_paths:
        APP.add_api_route(
            "/api/cashflow/requests/my",
            cashflow_my_requests,
            methods=["GET"],
        )
    if "/api/cashflow/withdraw-act" not in existing_paths:
        APP.add_api_route(
            "/api/cashflow/withdraw-act",
            cashflow_withdraw_act,
            methods=["GET"],
        )
    if "/api/cashflow/withdraw-act.xlsx" not in existing_paths:
        APP.add_api_route(
            "/api/cashflow/withdraw-act.xlsx",
            cashflow_withdraw_act_xlsx,
            methods=["GET"],
        )

_ensure_cashflow_routes()


@APP.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception) -> JSONResponse:
    details = {"path": request.url.path, "method": request.method}
    if request.client:
        details["client"] = request.client.host
    trace = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    try:
        log_system_log("ERROR", "api", str(exc), details=details, trace=trace)
    except Exception:
        pass
    return JSONResponse(status_code=500, content={"detail": "Internal server error"})



@APP.get("/health")
def health():
    return {"ok": True}


# (опционально) отдача webapp.html
@APP.get("/webapp")
def webapp():
    path = os.path.join(os.path.dirname(__file__), "webapp.html")
    if not os.path.exists(path):
        return JSONResponse({"detail": "webapp.html not found"}, status_code=404)
    return FileResponse(path, media_type="text/html")


@APP.get("/cashapp")
def cashapp():
    path = os.path.join(os.path.dirname(__file__), "cashapp.html")
    if not os.path.exists(path):
        return JSONResponse({"detail": "cashapp.html not found"}, status_code=404)
    return FileResponse(path, media_type="text/html")



# ---------------------------
# Auth
# ---------------------------

@APP.post("/api/auth/telegram", response_model=AuthOut)
def auth_telegram(body: AuthTelegramIn, request: Request):
    user_obj = validate_telegram_init_data(body.initData, CFG.BOT_TOKEN)
    telegram_id = int(user_obj.get("id", 0))
    if not telegram_id:
        raise HTTPException(status_code=401, detail="Invalid Telegram user id")

    allow = refresh_allowlist_if_needed()
    allow_user = allow.get(telegram_id)
    if not allow_user or not allow_user.get("active"):
        raise HTTPException(status_code=403, detail="User not in allowlist or inactive")

    # sync single user from allowlist (in case file changed)
    sync_allowlist_to_db({telegram_id: allow_user})

    u = db_fetchone("SELECT * FROM users WHERE telegram_id=?;", (telegram_id,))
    if not u or int(u["active"]) != 1:
        raise HTTPException(status_code=403, detail="User inactive")

    # session token
    exp = int(time.time()) + 7 * 24 * 3600
    token = make_session_token(
        {
            "telegram_id": telegram_id,
            "role": str(u["role"]),
            "exp": exp,
        },
        CFG.SESSION_SECRET,
    )
    return {"token": token, "user": {"telegram_id": telegram_id, "name": u["name"], "role": u["role"]}}


@APP.get("/api/me")
def me(u: sqlite3.Row = Depends(get_current_user)):
    return {"id": u["id"], "telegram_id": u["telegram_id"], "name": u["name"], "role": u["role"]}


# ---------------------------
# Months
# ---------------------------

@APP.get("/api/months")
def list_months(
    year: int = Query(...),
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    rows = db_fetchall("SELECT * FROM months WHERE year=? ORDER BY month ASC;", (year,))
    return {"items": [dict(r) for r in rows]}

@APP.get("/api/months/latest")
def latest_month(
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    row = db_fetchone("SELECT * FROM months ORDER BY year DESC, month DESC LIMIT 1;")
    return {"item": dict(row) if row else None}

@APP.get("/api/months/latest")
def latest_month(
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    row = db_fetchone("SELECT * FROM months ORDER BY year DESC, month DESC LIMIT 1;")
    return {"item": dict(row) if row else None}



@APP.post("/api/months")
def create_month(
    body: MonthCreateIn,
    u: sqlite3.Row = Depends(require_role("admin")),
):
    tz = CFG.tzinfo()
    now = iso_now(tz)

    # start_balance default: previous month fact_balance
    start_balance = body.start_balance
    if start_balance is None:
        prev_y, prev_m = body.year, body.month - 1
        if prev_m == 0:
            prev_m = 12
            prev_y -= 1
        prev = db_fetchone("SELECT id FROM months WHERE year=? AND month=?;", (prev_y, prev_m))
        if prev:
            prev_summary = compute_month_summary(int(prev["id"]), ensure_tithe=True)
            start_balance = float(prev_summary["fact_balance"])
        else:
            start_balance = 0.0

    try:
        new_id = db_exec_returning_id(
            """
            INSERT INTO months (year, month, monthly_min_needed, start_balance, sundays_override, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?);
            """,
            (body.year, body.month, float(body.monthly_min_needed), float(start_balance), body.sundays_override, now, now),
        )
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Month already exists")

    after = db_fetchone("SELECT * FROM months WHERE id=?;", (new_id,))
    log_audit(int(u["id"]), "CREATE", "month", int(new_id), None, dict(after) if after else None)

    recalc_services_for_month(new_id)
    ensure_tithe_expense(new_id, user_id=int(u["id"]))
    return {"id": new_id}


@APP.get("/api/months/{month_id}/summary")
def month_summary(
    month_id: int,
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    return compute_month_summary(month_id, ensure_tithe=True)

@APP.get("/api/months/{month_id}/budget")
def list_month_budget(
    month_id: int,
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    get_month_by_id(month_id)
    return {"items": get_month_budget_rows(month_id)}


@APP.put("/api/months/{month_id}/budget")
def upsert_month_budget(
    month_id: int,
    body: List[MonthBudgetIn] = Body(default_factory=list),
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    get_month_by_id(month_id)
    tz = CFG.tzinfo()
    now = iso_now(tz)
    seen: set[int] = set()

    with db_connect() as conn:
        for item in body:
            category_id = int(item.category_id)
            if category_id in seen:
                raise HTTPException(status_code=400, detail="Duplicate category in payload")
            seen.add(category_id)
            cat = conn.execute("SELECT id FROM categories WHERE id=?;", (category_id,)).fetchone()
            if not cat:
                raise HTTPException(status_code=400, detail=f"Category {category_id} not found")
            limit_amount = float(item.limit_amount)
            warn_threshold = float(item.warn_threshold if item.warn_threshold is not None else 0.9)
            include_system = 1 if bool(item.include_system) else 0
            existing = conn.execute(
                "SELECT id FROM month_budgets WHERE month_id=? AND category_id=?;",
                (month_id, category_id),
            ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE month_budgets
                    SET limit_amount=?, warn_threshold=?, include_system=?, updated_at=?
                    WHERE id=?;
                    """,
                    (limit_amount, warn_threshold, include_system, now, int(existing["id"])),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO month_budgets (
                        month_id, category_id, limit_amount, warn_threshold,
                        include_system, created_by_user_id, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?);
                    """,
                    (month_id, category_id, limit_amount, warn_threshold, include_system, int(u["id"]), now, now),
                )
        if seen:
            placeholders = ",".join("?" for _ in seen)
            conn.execute(
                f"DELETE FROM month_budgets WHERE month_id=? AND category_id NOT IN ({placeholders});",
                (month_id, *seen),
            )
        else:
            conn.execute("DELETE FROM month_budgets WHERE month_id=?;", (month_id,))
        conn.commit()

    return {"items": get_month_budget_rows(month_id)}



@APP.get("/api/analytics/year")
def year_analytics(
    year: int = Query(...),
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    return compute_year_analytics(int(year))

@APP.get("/api/analytics/period")
def period_analytics(
    type: str = Query(..., alias="type"),
    year: int = Query(...),
    month: Optional[int] = Query(None),
    quarter: Optional[int] = Query(None),
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    include_top = str(u["role"]) in ("admin", "accountant")
    return compute_period_analytics(type, int(year), month, quarter, include_top_categories=include_top)



@APP.put("/api/months/{month_id}")
def update_month(
    month_id: int,
    body: MonthUpdateIn,
    u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Month not found")

    fields = []
    params: List[Any] = []
    if body.monthly_min_needed is not None:
        fields.append("monthly_min_needed=?")
        params.append(float(body.monthly_min_needed))
    if body.start_balance is not None:
        fields.append("start_balance=?")
        params.append(float(body.start_balance))
    if body.sundays_override is not None:
        fields.append("sundays_override=?")
        params.append(int(body.sundays_override))

    if not fields:
        return {"ok": True}

    params.append(iso_now(CFG.tzinfo()))
    params.append(month_id)

    db_exec(f"UPDATE months SET {', '.join(fields)}, updated_at=? WHERE id=?;", tuple(params))
    after = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    log_audit(int(u["id"]), "UPDATE", "month", month_id, dict(before), dict(after) if after else None)

    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    return {"ok": True}

@APP.post("/api/months/{month_id}/close")
def close_month(
    month_id: int,
    u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Month not found")

    before_closed = int(before["is_closed"] or 0) if table_has_column("months", "is_closed") else 0
    if before_closed == 1:
        return {"ok": True, "already_closed": True}

    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=int(u["id"]))

    now = iso_now(CFG.tzinfo())
    db_exec(
        """
        UPDATE months
        SET is_closed=1, closed_at=?, closed_by_user_id=?, updated_at=?
        WHERE id=?;
        """,
        (now, int(u["id"]), now, month_id),
    )
    after = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    log_audit(int(u["id"]), "CLOSE", "month", month_id, dict(before), dict(after) if after else None)
    return {"ok": True}


@APP.post("/api/months/{month_id}/reopen")
def reopen_month(
    month_id: int,
    u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Month not found")

    before_closed = int(before["is_closed"] or 0) if table_has_column("months", "is_closed") else 0
    if before_closed == 0:
        return {"ok": True, "already_open": True}

    now = iso_now(CFG.tzinfo())
    db_exec(
        """
        UPDATE months
        SET is_closed=0, closed_at=NULL, closed_by_user_id=NULL, updated_at=?
        WHERE id=?;
        """,
        (now, month_id),
    )
    after = db_fetchone("SELECT * FROM months WHERE id=?;", (month_id,))
    log_audit(int(u["id"]), "REOPEN", "month", month_id, dict(before), dict(after) if after else None)
    return {"ok": True}



# ---------------------------
# Services (donations)
# ---------------------------

@APP.get("/api/months/{month_id}/services")
def list_services(
    month_id: int,
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    recalc_services_for_month(month_id)
    rows = db_fetchall("SELECT * FROM services WHERE month_id=? ORDER BY service_date ASC;", (month_id,))
    return {"items": [dict(r) for r in rows]}


@APP.post("/api/months/{month_id}/services")
def create_service(
    month_id: int,
    body: ServiceIn,
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    ensure_month_open(month_id)
    m = get_month_by_id(month_id)
    tz = CFG.tzinfo()
    now = iso_now(tz)


    cashless = float(body.cashless)
    cash = float(body.cash)
    total = round(cashless + cash, 2)
    income_type = (body.income_type or "donation").strip().lower()
    if income_type not in ("donation", "other"):
        raise HTTPException(status_code=400, detail="Invalid income_type")
    service_date = body.service_date.isoformat()
    account = normalize_account(body.account)
    if cash > 0:
        request_id = create_cashflow_collect_request_if_needed(
            account=account,
            cash_amount=cash,
            cashless_amount=cashless,
            income_type=income_type,
            month_id=month_id,
            service_date=service_date,
            created_by_telegram_id=int(u["telegram_id"]),
        )
        if request_id:
            return JSONResponse(
                status_code=202,
                content={
                    "requires_signatures": True,
                    "request_id": request_id,
                    "detail": f"Сбор наличных требует подписей. Создан запрос №{request_id}.",
                },
            )
    before = db_fetchone(
        """
        SELECT * FROM services
        WHERE month_id=? AND service_date=? AND account=? AND income_type=?;
        """,
        (month_id, service_date, account, income_type),
    )

    if before:
        db_exec(
            """
            UPDATE services
            SET cashless=?, cash=?, total=?, income_type=?, account=?, updated_at=?
            WHERE id=?;
            """,
            (cashless, cash, total, income_type, account, now, before["id"]),
        )
        after = db_fetchone("SELECT * FROM services WHERE id=?;", (before["id"],))
        log_audit(int(u["id"]), "UPDATE", "service", int(before["id"]), dict(before), dict(after) if after else None)
        recalc_services_for_month(month_id)
        ensure_tithe_expense(month_id, user_id=int(u["id"]))
        return {"id": int(before["id"]), "updated": True}

    new_id = db_exec_returning_id(
        """
        INSERT INTO services (
            month_id, service_date, idx, cashless, cash, total,
            weekly_min_needed, mnsps_status, pvs_ratio, income_type, account,
            created_at, updated_at
        ) VALUES (?, ?, 0, ?, ?, ?, 0, 'Не собрана', 0, ?, ?, ?);
        """,
        (month_id, service_date, cashless, cash, total, income_type, account, now, now),
    )
    after = db_fetchone("SELECT * FROM services WHERE id=?;", (new_id,))
    log_audit(int(u["id"]), "CREATE", "service", int(new_id), None, dict(after) if after else None)

    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    return {"id": new_id, "updated": False}


@APP.put("/api/services/{service_id}")
def update_service(
    service_id: int,
    body: ServiceIn,
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    before = db_fetchone("SELECT * FROM services WHERE id=?;", (service_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Service not found")
    ensure_month_open(int(before["month_id"]))

    tz = CFG.tzinfo()
    now = iso_now(tz)

    cashless = float(body.cashless)
    cash = float(body.cash)
    total = round(cashless + cash, 2)
    income_type = (body.income_type or "donation").strip().lower()
    if income_type not in ("donation", "other"):
        raise HTTPException(status_code=400, detail="Invalid income_type")
    new_date = body.service_date.isoformat()
    account = normalize_account(body.account)
    existing = db_fetchone(
        """
        SELECT id FROM services
        WHERE month_id=? AND service_date=? AND account=? AND income_type=?;
        """,
        (int(before["month_id"]), new_date, account, income_type),
    )
    if existing and int(existing["id"]) != int(service_id):
        raise HTTPException(status_code=409, detail="Service already exists for this account and date")

    db_exec(
        """
        UPDATE services
        SET service_date=?, cashless=?, cash=?, total=?, income_type=?, account=?, updated_at=?
        WHERE id=?;
        """,
        (new_date, cashless, cash, total, income_type, account, now, service_id),
    )
    after = db_fetchone("SELECT * FROM services WHERE id=?;", (service_id,))
    log_audit(int(u["id"]), "UPDATE", "service", service_id, dict(before), dict(after) if after else None)

    month_id = int(before["month_id"])
    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    return {"ok": True}


@APP.delete("/api/services/{service_id}")
def delete_service(
    service_id: int,
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    before = db_fetchone("SELECT * FROM services WHERE id=?;", (service_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Service not found")
    ensure_month_open(int(before["month_id"]))

    month_id = int(before["month_id"])
    db_exec("DELETE FROM services WHERE id=?;", (service_id,))
    log_audit(int(u["id"]), "DELETE", "service", service_id, dict(before), None)

    recalc_services_for_month(month_id)
    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    return {"ok": True}


# ---------------------------
# Categories + aliases
# ---------------------------

@APP.get("/api/categories")
def list_categories(
        u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    return {"items": get_categories_payload()}


@APP.post("/api/categories")
def create_category(
        body: CategoryCreateIn,
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    name = str(body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    existing = db_fetchone("SELECT * FROM categories WHERE name=?;", (name,))
    if existing:
        raise HTTPException(status_code=409, detail="Category already exists")
    now = iso_now(CFG.tzinfo())
    new_id = db_exec_returning_id(
        """
        INSERT INTO categories (name, is_active, sort_order, created_at, updated_at)
        VALUES (?, 1, 0, ?, ?);
        """,
        (name, now, now),
    )
    after = db_fetchone("SELECT * FROM categories WHERE id=?;", (new_id,))
    log_audit(int(u["id"]), "CREATE", "category", new_id, None, dict(after) if after else None)
    return {"id": new_id}


@APP.put("/api/categories/{category_id}")
def update_category(
        category_id: int,
        body: CategoryUpdateIn,
        u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM categories WHERE id=?;", (category_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Category not found")

    updates = []
    params: List[Any] = []
    if body.name is not None:
        name = str(body.name or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Name is required")
        existing = db_fetchone("SELECT id FROM categories WHERE name=?;", (name,))
        if existing and int(existing["id"]) != int(category_id):
            raise HTTPException(status_code=409, detail="Category name already exists")
        updates.append("name=?")
        params.append(name)
    if body.is_active is not None:
        updates.append("is_active=?")
        params.append(1 if body.is_active else 0)
    if body.sort_order is not None:
        updates.append("sort_order=?")
        params.append(int(body.sort_order))
    if not updates:
        return {"ok": True}
    updates.append("updated_at=?")
    params.append(iso_now(CFG.tzinfo()))
    params.append(category_id)
    db_exec(
        f"UPDATE categories SET {', '.join(updates)} WHERE id=?;",
        tuple(params),
    )
    after = db_fetchone("SELECT * FROM categories WHERE id=?;", (category_id,))
    log_audit(int(u["id"]), "UPDATE", "category", category_id, dict(before), dict(after) if after else None)
    return {"ok": True}


@APP.post("/api/categories/{category_id}/aliases")
def create_category_alias(
        category_id: int,
        body: CategoryAliasCreateIn,
        u: sqlite3.Row = Depends(require_role("admin")),
):
    category = db_fetchone("SELECT * FROM categories WHERE id=?;", (category_id,))
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    alias = str(body.alias or "").strip()
    if not alias:
        raise HTTPException(status_code=400, detail="Alias is required")
    alias_norm = normalize_alias(alias)
    if not alias_norm:
        raise HTTPException(status_code=400, detail="Alias is invalid")
    existing = db_fetchone("SELECT id FROM category_aliases WHERE alias_norm=?;", (alias_norm,))
    if existing:
        raise HTTPException(status_code=409, detail="Alias already exists")
    now = iso_now(CFG.tzinfo())
    new_id = db_exec_returning_id(
        """
        INSERT INTO category_aliases (category_id, alias, alias_norm, created_at)
        VALUES (?, ?, ?, ?);
        """,
        (category_id, alias, alias_norm, now),
    )
    after = db_fetchone("SELECT * FROM category_aliases WHERE id=?;", (new_id,))
    log_audit(int(u["id"]), "CREATE", "category_alias", new_id, None, dict(after) if after else None)
    return {"id": new_id}


@APP.delete("/api/category-aliases/{alias_id}")
def delete_category_alias(
        alias_id: int,
        u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM category_aliases WHERE id=?;", (alias_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Alias not found")
    db_exec("DELETE FROM category_aliases WHERE id=?;", (alias_id,))
    log_audit(int(u["id"]), "DELETE", "category_alias", alias_id, dict(before), None)
    return {"ok": True}


@APP.post("/api/categories/merge")
def merge_categories(
        body: CategoryMergeIn,
        u: sqlite3.Row = Depends(require_role("admin")),
):
    target = db_fetchone("SELECT * FROM categories WHERE id=?;", (body.target_id,))
    if not target:
        raise HTTPException(status_code=404, detail="Target category not found")
    source_ids = [int(x) for x in body.source_ids if int(x) != int(body.target_id)]
    if not source_ids:
        raise HTTPException(status_code=400, detail="Source categories are required")

    sources = []
    for source_id in source_ids:
        row = db_fetchone("SELECT * FROM categories WHERE id=?;", (source_id,))
        if not row:
            raise HTTPException(status_code=404, detail=f"Source category {source_id} not found")
        sources.append(row)

    now = iso_now(CFG.tzinfo())
    before = {
        "target": dict(target),
        "sources": [dict(s) for s in sources],
    }
    with db_connect() as conn:
        for source in sources:
            conn.execute(
                "UPDATE expenses SET category=? WHERE category=?;",
                (target["name"], source["name"]),
            )
            conn.execute(
                "UPDATE category_aliases SET category_id=? WHERE category_id=?;",
                (int(target["id"]), int(source["id"])),
            )
            conn.execute(
                "UPDATE categories SET is_active=0, updated_at=? WHERE id=?;",
                (now, int(source["id"])),
            )
        conn.commit()
    after = {
        "target": dict(db_fetchone("SELECT * FROM categories WHERE id=?;", (target["id"],)) or {}),
        "sources": [dict(db_fetchone("SELECT * FROM categories WHERE id=?;", (int(s["id"]),)) or {}) for s in sources],
    }
    log_audit(int(u["id"]), "MERGE", "category", int(target["id"]), before, after)
    return {"ok": True}


@APP.post("/api/categories/rename-mass")
def rename_categories_mass(
        body: CategoryRenameMassIn,
        u: sqlite3.Row = Depends(require_role("admin")),
):
    from_name = str(body.from_ or "").strip()
    to_name = str(body.to or "").strip()
    if not from_name or not to_name:
        raise HTTPException(status_code=400, detail="Invalid rename payload")
    before_count = db_fetchone(
        "SELECT COUNT(*) AS cnt FROM expenses WHERE category=?;",
        (from_name,),
    )
    db_exec("UPDATE expenses SET category=? WHERE category=?;", (to_name, from_name))
    after_count = db_fetchone(
        "SELECT COUNT(*) AS cnt FROM expenses WHERE category=?;",
        (to_name,),
    )
    log_audit(
        int(u["id"]),
        "RENAME_MASS",
        "category",
        None,
        {"from": from_name, "count": int(before_count["cnt"] if before_count else 0)},
        {"to": to_name, "count": int(after_count["cnt"] if after_count else 0)},
    )
    return {"ok": True}


# ---------------------------
# Tags
# ---------------------------

@APP.get("/api/tags")
def list_tags(
        u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    return {"items": get_tags_payload()}


@APP.post("/api/tags")
def create_tag(
        body: Dict[str, Any] = Body(...),
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    name = str(body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    name_norm = normalize_tag_name(name)
    if not name_norm:
        raise HTTPException(status_code=400, detail="Invalid tag name")
    existing = db_fetchone("SELECT * FROM tags WHERE name_norm=?;", (name_norm,))
    if existing:
        raise HTTPException(status_code=409, detail="Tag already exists")
    now = iso_now(CFG.tzinfo())
    new_id = db_exec_returning_id(
        """
        INSERT INTO tags (name, name_norm, created_at, updated_at)
        VALUES (?, ?, ?, ?);
        """,
        (name, name_norm, now, now),
    )
    after = db_fetchone("SELECT * FROM tags WHERE id=?;", (new_id,))
    log_audit(int(u["id"]), "CREATE", "tag", new_id, None, dict(after) if after else None)
    return {"id": new_id}


@APP.put("/api/tags/{tag_id}")
def update_tag(
        tag_id: int,
        body: Dict[str, Any] = Body(...),
        u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM tags WHERE id=?;", (tag_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Tag not found")
    name = str(body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    name_norm = normalize_tag_name(name)
    if not name_norm:
        raise HTTPException(status_code=400, detail="Invalid tag name")
    existing = db_fetchone("SELECT id FROM tags WHERE name_norm=?;", (name_norm,))
    if existing and int(existing["id"]) != int(tag_id):
        raise HTTPException(status_code=409, detail="Tag already exists")
    now = iso_now(CFG.tzinfo())
    db_exec(
        "UPDATE tags SET name=?, name_norm=?, updated_at=? WHERE id=?;",
        (name, name_norm, now, tag_id),
    )
    after = db_fetchone("SELECT * FROM tags WHERE id=?;", (tag_id,))
    log_audit(int(u["id"]), "UPDATE", "tag", tag_id, dict(before), dict(after) if after else None)
    return {"ok": True}


@APP.delete("/api/tags/{tag_id}")
def delete_tag(
        tag_id: int,
        u: sqlite3.Row = Depends(require_role("admin")),
):
    before = db_fetchone("SELECT * FROM tags WHERE id=?;", (tag_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Tag not found")
    db_exec("DELETE FROM tags WHERE id=?;", (tag_id,))
    log_audit(int(u["id"]), "DELETE", "tag", tag_id, dict(before), None)
    return {"ok": True}




# ---------------------------
# Expenses
# ---------------------------

@APP.get("/api/months/{month_id}/expenses")
def list_expenses(
        month_id: int,
        tags: Optional[str] = Query(default=None),
        mode: str = Query(default="any"),
        u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    ensure_tithe_expense(month_id, user_id=None)
    tag_filters: List[str] = []
    if tags:
        tag_filters = normalize_tag_list(tags.split(","))
    tag_norms: List[str] = []
    seen_norms: set[str] = set()
    for tag in tag_filters:
        norm = normalize_tag_name(tag)
        if not norm or norm in seen_norms:
            continue
        seen_norms.add(norm)
        tag_norms.append(norm)
    if mode not in ("any", "all"):
        raise HTTPException(status_code=400, detail="Invalid tag filter mode")

    params: List[Any] = [month_id]
    tag_clause = ""
    if tag_norms:
        placeholders = ",".join("?" for _ in tag_norms)
        if mode == "any":
            tag_clause = f"""
                AND EXISTS (
                    SELECT 1
                    FROM expense_tags et
                    JOIN tags t ON t.id=et.tag_id
                    WHERE et.expense_id=e.id AND t.name_norm IN ({placeholders})
                )
            """
            params.extend(tag_norms)
        else:
            tag_clause = f"""
                AND (
                    SELECT COUNT(DISTINCT t.name_norm)
                    FROM expense_tags et
                    JOIN tags t ON t.id=et.tag_id
                    WHERE et.expense_id=e.id AND t.name_norm IN ({placeholders})
                ) = ?
            """
            params.extend(tag_norms)
            params.append(len(tag_norms))

    rows = db_fetchall(
        f"""
        SELECT
            e.*,
            (
                SELECT COUNT(*)
                FROM attachments a
                WHERE a.entity_type='expense' AND a.entity_id=e.id
            ) AS attachments_count
        FROM expenses e
        WHERE e.month_id=? {tag_clause}
        ORDER BY e.expense_date DESC, e.id DESC;
        """,
        tuple(params),
    )
    items = [dict(r) for r in rows]
    tag_map = fetch_expense_tags([int(r["id"]) for r in rows])
    for item in items:
        item["tags"] = tag_map.get(int(item["id"]), [])
    return {"items": items}



@APP.post("/api/months/{month_id}/expenses")
def create_expense(
    month_id: int,
    body: ExpenseIn,
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    ensure_month_open(month_id)
    tz = CFG.tzinfo()
    now = iso_now(tz)

    account = normalize_account(body.account)
    category = resolve_category(body.category, str(u["role"]), int(u["id"]))
    total = round(float(body.qty) * float(body.unit_amount), 2)
    new_id = db_exec_returning_id(
        """
        INSERT INTO expenses (
            month_id, expense_date, category, title, qty, unit_amount, total, comment,
            is_system, account, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?);
        """,
        (
            month_id,
            body.expense_date.isoformat(),
            category,
            body.title,
            float(body.qty),
            float(body.unit_amount),
            total,
            body.comment,
            account,
            now,
            now,
        ),
    )
    tag_names = set_expense_tags(new_id, body.tags, str(u["role"]), int(u["id"]))
    after = db_fetchone("SELECT * FROM expenses WHERE id=?;", (new_id,))
    after_payload = dict(after) if after else None
    if after_payload is not None:
        after_payload["tags"] = tag_names
    log_audit(int(u["id"]), "CREATE", "expense", int(new_id), None, after_payload)

    # ensure tithe exists (depends on income; no harm to upsert)
    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    warnings: List[Dict[str, Any]] = []
    if account == "main":
        budget_warning = get_budget_warning_for_category(month_id, category)
        if budget_warning:
            warnings.append(budget_warning)
    return {"id": new_id, "warnings": warnings}


@APP.put("/api/expenses/{expense_id}")
def update_expense(
    expense_id: int,
    body: ExpenseIn,
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    before = db_fetchone("SELECT * FROM expenses WHERE id=?;", (expense_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Expense not found")
    ensure_month_open(int(before["month_id"]))
    if int(before["is_system"]) == 1:
        # system row edit blocked by default
        raise HTTPException(status_code=403, detail="System expense cannot be edited directly")

    tz = CFG.tzinfo()
    now = iso_now(tz)

    account = normalize_account(body.account)
    category = resolve_category(body.category, str(u["role"]), int(u["id"]))
    total = round(float(body.qty) * float(body.unit_amount), 2)
    before_payload = dict(before)
    before_payload["tags"] = get_expense_tag_names(expense_id)

    db_exec(
        """
        UPDATE expenses
        SET expense_date=?, category=?, title=?, qty=?, unit_amount=?, total=?, comment=?, account=?, updated_at=?
        WHERE id=?;
        """,
        (
            body.expense_date.isoformat(),
            category,
            body.title,
            float(body.qty),
            float(body.unit_amount),
            total,
            body.comment,
            account,
            now,
            expense_id,
        ),
    )
    tag_names = set_expense_tags(expense_id, body.tags, str(u["role"]), int(u["id"]))
    after = db_fetchone("SELECT * FROM expenses WHERE id=?;", (expense_id,))
    after_payload = dict(after) if after else None
    if after_payload is not None:
        after_payload["tags"] = tag_names
    log_audit(int(u["id"]), "UPDATE", "expense", expense_id, before_payload, after_payload)

    ensure_tithe_expense(int(before["month_id"]), user_id=int(u["id"]))
    return {"ok": True}


@APP.delete("/api/expenses/{expense_id}")
def delete_expense(
    expense_id: int,
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    before = db_fetchone("SELECT * FROM expenses WHERE id=?;", (expense_id,))
    if not before:
        raise HTTPException(status_code=404, detail="Expense not found")
    ensure_month_open(int(before["month_id"]))
    if int(before["is_system"]) == 1:
        raise HTTPException(status_code=403, detail="System expense cannot be deleted")

    before_payload = dict(before)
    before_payload["tags"] = get_expense_tag_names(expense_id)

    month_id = int(before["month_id"])
    db_exec("DELETE FROM expenses WHERE id=?;", (expense_id,))
    log_audit(int(u["id"]), "DELETE", "expense", expense_id, before_payload, None)

    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    return {"ok": True}


# ---------------------------
# Drafts
# ---------------------------

@APP.post("/api/months/{month_id}/drafts/expenses")
def create_expense_draft(
        month_id: int,
        body: Dict[str, Any] = Body(...),
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    ensure_month_open(month_id)
    tz = CFG.tzinfo()
    now = iso_now(tz)
    payload = normalize_expense_draft_payload(body, tz)
    payload["category"] = resolve_category(payload.get("category"), str(u["role"]), int(u["id"]))
    new_id = db_exec_returning_id(
        """
        INSERT INTO drafts (
            kind, month_id, created_by_user_id, payload_json, status, created_at, updated_at
        ) VALUES ('expense', ?, ?, ?, 'draft', ?, ?);
        """,
        (
            month_id,
            int(u["id"]),
            json.dumps(payload, ensure_ascii=False),
            now,
            now,
        ),
    )
    after = db_fetchone("SELECT * FROM drafts WHERE id=?;", (new_id,))
    log_audit(int(u["id"]), "CREATE", "draft", new_id, None, dict(after) if after else None)
    return {"id": new_id}


@APP.put("/api/drafts/{draft_id}")
def update_draft(
        draft_id: int,
        body: Dict[str, Any] = Body(...),
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    existing = get_draft_or_404(draft_id)
    if existing["status"] != "draft":
        raise HTTPException(status_code=409, detail="Draft is not editable")
    if u["role"] != "admin" and int(existing["created_by_user_id"]) != int(u["id"]):
        raise HTTPException(status_code=403, detail="Forbidden")
    ensure_month_open(int(existing["month_id"]))

    tz = CFG.tzinfo()
    now = iso_now(tz)
    payload = normalize_expense_draft_payload(body, tz)
    payload["category"] = resolve_category(payload.get("category"), str(u["role"]), int(u["id"]))
    db_exec(
        "UPDATE drafts SET payload_json=?, updated_at=? WHERE id=?;",
        (json.dumps(payload, ensure_ascii=False), now, draft_id),
    )
    after = db_fetchone("SELECT * FROM drafts WHERE id=?;", (draft_id,))
    log_audit(int(u["id"]), "UPDATE", "draft", draft_id, dict(existing), dict(after) if after else None)
    return {"ok": True}


@APP.get("/api/months/{month_id}/drafts")
def list_drafts(
        month_id: int,
        kind: str = Query("expense"),
        scope: str = Query("mine"),
        u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    if kind != "expense":
        raise HTTPException(status_code=400, detail="Unsupported draft kind")
    if scope not in ("mine", "all"):
        raise HTTPException(status_code=400, detail="Invalid scope")
    if scope == "all" and u["role"] != "admin":
        raise HTTPException(status_code=403, detail="Insufficient role for scope=all")

    params: List[Any] = [month_id, kind, "draft"]
    sql = """
            SELECT *
            FROM drafts
            WHERE month_id=? AND kind=? AND status=?
        """
    if scope == "mine":
        sql += " AND created_by_user_id=?"
        params.append(int(u["id"]))
    sql += " ORDER BY updated_at DESC, id DESC;"

    rows = db_fetchall(sql, tuple(params))
    items: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        payload = json.loads(r["payload_json"])
        item["payload"] = payload
        item["summary"] = draft_payload_summary(payload)
        items.append(item)
    return {"items": items}


@APP.post("/api/drafts/{draft_id}/submit")
def submit_draft(
        draft_id: int,
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    draft = get_draft_or_404(draft_id)
    if draft["status"] != "draft":
        raise HTTPException(status_code=409, detail="Draft is not submittable")
    if u["role"] != "admin" and int(draft["created_by_user_id"]) != int(u["id"]):
        raise HTTPException(status_code=403, detail="Forbidden")

    month_id = int(draft["month_id"])
    ensure_month_open(month_id)

    tz = CFG.tzinfo()
    now = iso_now(tz)
    payload = normalize_expense_draft_payload(json.loads(draft["payload_json"]), tz)
    payload["category"] = resolve_category(payload.get("category"), str(u["role"]), int(u["id"]))
    total = round(float(payload["qty"]) * float(payload["unit_amount"]), 2)
    account = normalize_account(payload.get("account"))

    new_id = db_exec_returning_id(
        """
        INSERT INTO expenses (
            month_id, expense_date, category, title, qty, unit_amount, total, comment,
            is_system, account, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?);
        """,
        (
            month_id,
            payload["expense_date"],
            payload["category"],
            payload["title"],
            float(payload["qty"]),
            float(payload["unit_amount"]),
            total,
            payload.get("comment"),
            account,
            now,
            now,
        ),
    )
    tag_names = set_expense_tags(new_id, payload.get("tags"), str(u["role"]), int(u["id"]))
    after = db_fetchone("SELECT * FROM expenses WHERE id=?;", (new_id,))
    after_payload = dict(after) if after else None
    if after_payload is not None:
        after_payload["tags"] = tag_names
    log_audit(int(u["id"]), "CREATE", "expense", new_id, None, after_payload)

    db_exec(
        "UPDATE drafts SET status='submitted', updated_at=? WHERE id=?;",
        (now, draft_id),
    )
    log_audit(int(u["id"]), "SUBMIT", "draft", draft_id, dict(draft), None)

    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    return {"expense_id": new_id}


@APP.delete("/api/drafts/{draft_id}")
def delete_draft(
        draft_id: int,
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    draft = get_draft_or_404(draft_id)
    if draft["status"] != "draft":
        raise HTTPException(status_code=409, detail="Draft cannot be deleted")
    if u["role"] != "admin" and int(draft["created_by_user_id"]) != int(u["id"]):
        raise HTTPException(status_code=403, detail="Forbidden")
    ensure_month_open(int(draft["month_id"]))

    now = iso_now(CFG.tzinfo())
    db_exec("UPDATE drafts SET status='deleted', updated_at=? WHERE id=?;", (now, draft_id))
    log_audit(int(u["id"]), "DELETE", "draft", draft_id, dict(draft), None)
    return {"ok": True}


# ---------------------------
# Attachments
# ---------------------------

@APP.post("/api/expenses/{expense_id}/attachments")
def upload_expense_attachment(
        expense_id: int,
        file: UploadFile = File(...),
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    expense = get_expense_or_404(expense_id)
    ensure_month_open(int(expense["month_id"]))

    content_type = (file.content_type or "").lower()
    if content_type not in ALLOWED_ATTACHMENT_MIME_TYPES:
        raise HTTPException(status_code=415, detail="Unsupported file type")

    existing = db_fetchone(
        """
        SELECT COUNT(*) AS cnt
        FROM attachments
        WHERE entity_type='expense' AND entity_id=?;
        """,
        (expense_id,),
    )
    if existing and int(existing["cnt"]) >= MAX_ATTACHMENTS_PER_EXPENSE:
        raise HTTPException(status_code=409, detail="Attachments limit reached")

    now = iso_now(CFG.tzinfo())
    orig_filename = (file.filename or "attachment").strip()

    placeholder_id = db_exec_returning_id(
        """
        INSERT INTO attachments (
            entity_type, entity_id, orig_filename, stored_filename,
            mime, size_bytes, sha256, created_by_user_id, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        (
            "expense",
            expense_id,
            orig_filename,
            "pending",
            content_type,
            0,
            "pending",
            int(u["id"]),
            now,
        ),
    )

    ext = attachment_extension(orig_filename, content_type)
    stored_filename = f"{placeholder_id}_{int(time.time())}_{secrets.token_hex(4)}{ext}"
    target_dir = attachment_storage_dir(expense_id, now)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / stored_filename

    size_bytes = 0
    hasher = hashlib.sha256()
    try:
        with target_path.open("wb") as out:
            while True:
                chunk = file.file.read(1024 * 1024)
                if not chunk:
                    break
                size_bytes += len(chunk)
                if size_bytes > MAX_ATTACHMENT_BYTES:
                    raise HTTPException(status_code=413, detail="File too large")
                hasher.update(chunk)
                out.write(chunk)
    except HTTPException:
        if target_path.exists():
            target_path.unlink()
        db_exec("DELETE FROM attachments WHERE id=?;", (placeholder_id,))
        raise
    except Exception:
        if target_path.exists():
            target_path.unlink()
        db_exec("DELETE FROM attachments WHERE id=?;", (placeholder_id,))
        raise HTTPException(status_code=500, detail="Failed to store attachment")

    db_exec(
        """
        UPDATE attachments
        SET stored_filename=?, size_bytes=?, sha256=?
        WHERE id=?;
        """,
        (stored_filename, size_bytes, hasher.hexdigest(), placeholder_id),
    )

    after = db_fetchone("SELECT * FROM attachments WHERE id=?;", (placeholder_id,))
    log_audit(int(u["id"]), "CREATE_ATTACHMENT", "attachment", placeholder_id, None, dict(after) if after else None)

    return {
        "attachment": {
            "id": placeholder_id,
            "entity_type": "expense",
            "entity_id": expense_id,
            "orig_filename": orig_filename,
            "mime": content_type,
            "size_bytes": size_bytes,
            "created_at": now,
        }
    }


@APP.get("/api/expenses/{expense_id}/attachments")
def list_expense_attachments(
        expense_id: int,
        u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    get_expense_or_404(expense_id)
    rows = db_fetchall(
        """
        SELECT id, entity_type, entity_id, orig_filename, mime, size_bytes, created_at
        FROM attachments
        WHERE entity_type='expense' AND entity_id=?
        ORDER BY id DESC;
        """,
        (expense_id,),
    )
    return {"items": [dict(r) for r in rows]}


@APP.get("/api/attachments/{attachment_id}")
def get_attachment(
        attachment_id: int,
        inline: int = Query(1),
        u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    row = get_attachment_or_404(attachment_id)
    if row["entity_type"] != "expense":
        raise HTTPException(status_code=404, detail="Attachment not found")
    expense = db_fetchone("SELECT id FROM expenses WHERE id=?;", (int(row["entity_id"]),))
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")

    target_path = attachment_storage_dir(int(row["entity_id"]), row["created_at"]) / row["stored_filename"]
    if not target_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    disposition = "inline" if int(inline or 1) == 1 else "attachment"
    headers = {"Content-Disposition": f'{disposition}; filename="{row["orig_filename"]}"'}
    return FileResponse(target_path, media_type=row["mime"], filename=row["orig_filename"], headers=headers)


@APP.delete("/api/attachments/{attachment_id}")
def delete_attachment(
        attachment_id: int,
        u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    row = get_attachment_or_404(attachment_id)
    if row["entity_type"] != "expense":
        raise HTTPException(status_code=404, detail="Attachment not found")
    expense = db_fetchone("SELECT id, month_id FROM expenses WHERE id=?;", (int(row["entity_id"]),))
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    ensure_month_open(int(expense["month_id"]))

    target_path = attachment_storage_dir(int(row["entity_id"]), row["created_at"]) / row["stored_filename"]
    if target_path.exists():
        target_path.unlink()

    db_exec("DELETE FROM attachments WHERE id=?;", (attachment_id,))
    log_audit(int(u["id"]), "DELETE_ATTACHMENT", "attachment", attachment_id, dict(row), None)
    return {"ok": True}


# ---------------------------
# Settings
# ---------------------------

@APP.get("/api/settings")
def api_get_settings(u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer"))):
    s = get_settings()
    out = dict(s)
    out["daily_expenses_enabled"] = bool(int(out.get("daily_expenses_enabled") or 0))
    return out


@APP.put("/api/settings")
async def api_update_settings(
    body: SettingsUpdateIn,
    u: sqlite3.Row = Depends(require_role("admin", "viewer")),
):
    role = str(u["role"])
    if role != "admin":
        non_theme_fields = [
            body.report_chat_id,
            body.sunday_report_time,
            body.month_report_time,
            body.timezone,
            body.daily_expenses_enabled,
        ]
        if body.ui_theme is None or any(field is not None for field in non_theme_fields):
            raise HTTPException(status_code=403, detail="Only theme updates allowed")

    before = get_settings()
    fields = []
    params: List[Any] = []

    if body.report_chat_id is not None:
        report_chat_id = int(body.report_chat_id) if body.report_chat_id is not None else None
        if report_chat_id is not None:
            await ensure_report_chat_reachable(report_chat_id)
        fields.append("report_chat_id=?")
        params.append(report_chat_id)

    if body.sunday_report_time is not None:
        fields.append("sunday_report_time=?")
        params.append(body.sunday_report_time.strip())

    if body.month_report_time is not None:
        fields.append("month_report_time=?")
        params.append(body.month_report_time.strip())

    if body.timezone is not None:
        fields.append("timezone=?")
        params.append(body.timezone.strip())

    if body.ui_theme is not None:
        fields.append("ui_theme=?")
        params.append(body.ui_theme.strip())

    if body.daily_expenses_enabled is not None:
        fields.append("daily_expenses_enabled=?")
        params.append(1 if body.daily_expenses_enabled else 0)

    if not fields:
        return {"ok": True}

    params.append(iso_now(CFG.tzinfo()))
    db_exec(f"UPDATE settings SET {', '.join(fields)}, updated_at=? WHERE id=?;", tuple(params + [int(before["id"])]))

    after = get_settings()
    log_audit(int(u["id"]), "UPDATE", "settings", int(after["id"]), dict(before), dict(after))

    # reschedule planner
    reschedule_jobs()
    return {"ok": True}

# ---------------------------
# Monitoring API
# ---------------------------

@APP.get("/api/admin/monitor/overview")
def api_monitor_overview(
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    tzinfo = CFG.tzinfo()
    since = (dt.datetime.now(tzinfo) - dt.timedelta(hours=24)).replace(microsecond=0).isoformat()
    error_row = db_fetchone(
        "SELECT COUNT(*) AS c FROM system_logs WHERE level='ERROR' AND created_at >= ?;",
        (since,),
    )
    fail_row = db_fetchone(
        "SELECT COUNT(*) AS c FROM message_deliveries WHERE status='fail' AND created_at >= ?;",
        (since,),
    )
    job_rows = db_fetchall(
        """
        SELECT *
        FROM job_runs
        WHERE id IN (SELECT MAX(id) FROM job_runs GROUP BY job_id)
        ORDER BY started_at DESC;
        """
    )
    return {
        "since": since,
        "errors_24h": int(error_row["c"]) if error_row else 0,
        "failed_deliveries_24h": int(fail_row["c"]) if fail_row else 0,
        "jobs": [dict(r) for r in job_rows],
    }


@APP.get("/api/admin/monitor/logs")
def api_monitor_logs(
    level: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    params: List[Any] = []
    where = ""
    if level:
        where = "WHERE level=?"
        params.append(level.upper())
    params.append(int(limit))
    rows = db_fetchall(
        f"SELECT * FROM system_logs {where} ORDER BY created_at DESC LIMIT ?;",
        tuple(params),
    )
    return {"items": [dict(r) for r in rows]}


@APP.get("/api/admin/monitor/jobs")
def api_monitor_jobs(
    limit: int = Query(50, ge=1, le=500),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    rows = db_fetchall(
        "SELECT * FROM job_runs ORDER BY started_at DESC LIMIT ?;",
        (int(limit),),
    )
    return {"items": [dict(r) for r in rows]}


@APP.get("/api/admin/monitor/deliveries")
def api_monitor_deliveries(
    limit: int = Query(50, ge=1, le=500),
    kind: Optional[str] = Query(None),
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    params: List[Any] = []
    where = ""
    if kind:
        where = "WHERE kind=?"
        params.append(kind)
    params.append(int(limit))
    rows = db_fetchall(
        f"SELECT * FROM message_deliveries {where} ORDER BY created_at DESC LIMIT ?;",
        tuple(params),
    )
    return {"items": [dict(r) for r in rows]}


# ---------------------------
# Diagnostics API
# ---------------------------

DIAG_RUNNING_TASKS: Dict[int, asyncio.Task] = {}
DIAG_CANCEL_EVENTS: Dict[int, asyncio.Event] = {}


def _mask_secret(value: str) -> str:
    v = str(value or "")
    if len(v) <= 8:
        return "***"
    return f"{v[:4]}***{v[-4:]}"


def _json_loads(raw: Optional[str], fallback: Any) -> Any:
    if not raw:
        return fallback
    try:
        return json.loads(raw)
    except Exception:
        return fallback


class DiagnosticsRunOptions(BaseModel):
    telegram_send_test: bool = False
    telegram_target_id: Optional[int] = None
    backup_full: bool = False
    timeout_sec: int = Field(default=120, ge=10, le=1800)


class DiagnosticsRunIn(BaseModel):
    suite: str = Field(default="quick")
    mode: str = Field(default="safe")
    options: DiagnosticsRunOptions = Field(default_factory=DiagnosticsRunOptions)


class DiagnosticsContext:
    def __init__(self, run_id: int, suite: str, mode: str, options: DiagnosticsRunOptions):
        self.run_id = int(run_id)
        self.suite = suite
        self.mode = mode
        self.options = options
        self.cancel_event: asyncio.Event = DIAG_CANCEL_EVENTS.get(run_id, asyncio.Event())
        self.sandbox_dir: Optional[Path] = None
        self.sandbox_db: Optional[Path] = None

    def selected_db_path(self) -> Path:
        return self.sandbox_db or Path(CFG.DB_PATH)


class DiagnosticsCheck:
    def __init__(
        self,
        key: str,
        title: str,
        severity: str,
        allowed_modes: Set[str],
        fn: Callable[[DiagnosticsContext], Awaitable[Dict[str, Any]]],
    ):
        self.key = key
        self.title = title
        self.severity = severity
        self.allowed_modes = allowed_modes
        self.fn = fn


async def _diag_check_preflight(ctx: DiagnosticsContext) -> Dict[str, Any]:
    required = {
        "BOT_TOKEN": CFG.BOT_TOKEN,
        "DB_PATH": CFG.DB_PATH,
        "USERS_JSON": CFG.USERS_JSON_PATH,
        "APP_URL": CFG.APP_URL,
        "WEBAPP_URL": CFG.WEBAPP_URL,
        "TIMEZONE": CFG.TZ,
    }
    missing = [k for k, v in required.items() if not str(v or "").strip()]
    users_raw = Path(CFG.USERS_JSON_PATH).read_text(encoding="utf-8")
    users = json.loads(users_raw)
    users = users.get("users") if isinstance(users, dict) else users
    has_admin = any(int(u.get("active", 1)) == 1 and str(u.get("role", "")) == "admin" for u in (users or []))
    if not has_admin:
        raise RuntimeError("No active admin in users.json")
    for name in ("webapp.html", "cashapp.html"):
        if not (BASE_DIR / name).exists():
            raise RuntimeError(f"Missing template: {name}")
    for d in (ATTACHMENTS_DIR, BACKUPS_DIR):
        d.mkdir(parents=True, exist_ok=True)
        test = d / f".diag_write_{ctx.run_id}"
        test.write_text("ok", encoding="utf-8")
        test.unlink(missing_ok=True)
    details = {
        "missing": missing,
        "token_masked": _mask_secret(CFG.BOT_TOKEN),
    }
    if missing:
        return {"status": "warn", "message": "Missing required env vars", "details": details}
    return {"status": "success", "message": "Preflight checks passed", "details": details}


async def _diag_check_db_integrity(ctx: DiagnosticsContext) -> Dict[str, Any]:
    required_tables = [
        "users", "months", "expenses", "categories", "tags", "drafts", "attachments",
        "settings", "system_logs", "job_runs", "message_deliveries",
    ]
    path = ctx.selected_db_path()
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON;")
        tables = {r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table';")}
        missing = [t for t in required_tables if t not in tables]
        integ = conn.execute("PRAGMA integrity_check;").fetchone()[0]
    details = {"db_path": str(path), "missing_tables": missing, "integrity": integ}
    if missing or integ.lower() != "ok":
        return {"status": "fail", "message": "DB integrity check failed", "details": details}
    return {"status": "success", "message": "DB integrity check OK", "details": details}


async def _diag_check_scheduler(ctx: DiagnosticsContext) -> Dict[str, Any]:
    reschedule_jobs()
    jobs = [j.id for j in scheduler.get_jobs()]
    expected = {"sunday_report", "month_report", "daily_expenses"}
    missing = sorted(expected - set(jobs))
    if missing:
        return {"status": "warn", "message": "Some scheduler jobs are missing", "details": {"missing": missing, "jobs": jobs}}
    return {"status": "success", "message": "Scheduler jobs present", "details": {"jobs": jobs}}


async def _diag_check_sandbox_crud(ctx: DiagnosticsContext) -> Dict[str, Any]:
    if ctx.mode == "safe":
        return {"status": "skipped", "message": "CRUD skipped in SAFE mode", "details": {}}
    with sqlite3.connect(ctx.selected_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        now = iso_now(CFG.tzinfo())
        conn.execute("INSERT INTO users (telegram_id, name, role, active, created_at) VALUES (?, ?, ?, 1, ?);", (9000000 + ctx.run_id, "Diag", "admin", now))
        user_id = int(conn.execute("SELECT id FROM users WHERE telegram_id=?;", (9000000 + ctx.run_id,)).fetchone()[0])
        conn.execute("INSERT INTO months (year, month, monthly_min_needed, start_balance, created_at, updated_at) VALUES (?, ?, 0, 0, ?, ?);", (2090, (ctx.run_id % 12) + 1, now, now))
        month_id = int(conn.execute("SELECT id FROM months WHERE year=2090 ORDER BY id DESC LIMIT 1;").fetchone()[0])
        conn.execute("INSERT INTO categories (name, is_active, sort_order, created_at, updated_at) VALUES (?, 1, 0, ?, ?);", (f"diag-{ctx.run_id}", now, now))
        category_id = int(conn.execute("SELECT id FROM categories WHERE name=?;", (f"diag-{ctx.run_id}",)).fetchone()[0])
        conn.execute("INSERT INTO expenses (month_id, expense_date, category_id, amount, note, created_by_user_id, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?);", (month_id, now[:10], category_id, 123.45, "diag", user_id, now, now))
        expense_id = int(conn.execute("SELECT id FROM expenses WHERE month_id=? ORDER BY id DESC LIMIT 1;", (month_id,)).fetchone()[0])
        conn.execute("DELETE FROM expenses WHERE id=?;", (expense_id,))
        conn.commit()
    return {"status": "success", "message": "Sandbox CRUD checks passed", "details": {"db_path": str(ctx.selected_db_path())}}


async def _diag_check_exports(ctx: DiagnosticsContext) -> Dict[str, Any]:
    with sqlite3.connect(ctx.selected_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        month = conn.execute("SELECT id FROM months ORDER BY year DESC, month DESC LIMIT 1;").fetchone()
    if not month:
        return {"status": "warn", "message": "No month in DB; export checks skipped", "details": {}}
    if ctx.mode == "sandbox":
        return {"status": "success", "message": "Sandbox export precheck passed", "details": {"month_id": int(month["id"])}}
    png, _, _ = build_month_report_png(int(month["id"]), preset="square", pixel_ratio=1, dpi=96)
    is_png = png.startswith(b"\x89PNG\r\n\x1a\n") and len(png) > 100
    if not is_png:
        return {"status": "fail", "message": "PNG export signature check failed", "details": {"size": len(png)}}
    return {"status": "success", "message": "Export checks passed", "details": {"png_size": len(png)}}


async def _diag_check_backups(ctx: DiagnosticsContext) -> Dict[str, Any]:
    source_db = ctx.selected_db_path()
    backup_dir = ctx.sandbox_dir / "backups" if ctx.sandbox_dir else BACKUPS_DIR
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now(CFG.tzinfo()).strftime("%Y%m%d_%H%M%S")
    db_backup = backup_dir / f"diag_{ctx.run_id}_{stamp}.sqlite3"
    shutil.copy2(source_db, db_backup)
    details: Dict[str, Any] = {"db_backup": str(db_backup), "db_size": db_backup.stat().st_size}
    if ctx.options.backup_full:
        full_path = backup_dir / f"diag_{ctx.run_id}_{stamp}.zip"
        with zipfile.ZipFile(full_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.write(source_db, arcname="db.sqlite3")
        with zipfile.ZipFile(full_path, "r") as zf:
            details["full_zip_entries"] = zf.namelist()
        details["full_zip"] = str(full_path)
    return {"status": "success", "message": "Backup checks passed", "details": details}


async def _diag_check_telegram(ctx: DiagnosticsContext) -> Dict[str, Any]:
    if ctx.mode != "live":
        return {"status": "skipped", "message": "Telegram checks require LIVE mode", "details": {}}
    me = await BOT.get_me()
    details = {"bot_id": me.id, "username": me.username}
    if ctx.options.telegram_send_test:
        if not ctx.options.telegram_target_id:
            return {"status": "warn", "message": "telegram_target_id is required for send test", "details": details}
        await BOT.send_message(int(ctx.options.telegram_target_id), f"🧪 Diagnostics run #{ctx.run_id} completed")
        details["sent_to"] = int(ctx.options.telegram_target_id)
    return {"status": "success", "message": "Telegram checks passed", "details": details}


def _build_checks(suite: str) -> List[DiagnosticsCheck]:
    base = [
        DiagnosticsCheck("preflight", "Preflight checks", "critical", {"safe", "sandbox", "live"}, _diag_check_preflight),
        DiagnosticsCheck("db.integrity", "DB integrity", "critical", {"safe", "sandbox", "live"}, _diag_check_db_integrity),
        DiagnosticsCheck("scheduler.jobs", "Scheduler jobs", "major", {"safe", "sandbox", "live"}, _diag_check_scheduler),
    ]
    if suite == "quick":
        return base
    if suite == "integrations":
        return base + [
            DiagnosticsCheck("telegram.integration", "Telegram integration", "major", {"live"}, _diag_check_telegram),
            DiagnosticsCheck("backups", "Backup diagnostics", "major", {"safe", "sandbox", "live"}, _diag_check_backups),
        ]
    return base + [
        DiagnosticsCheck("sandbox.crud", "Sandbox CRUD", "critical", {"sandbox", "live"}, _diag_check_sandbox_crud),
        DiagnosticsCheck("exports", "Export checks", "major", {"safe", "sandbox", "live"}, _diag_check_exports),
        DiagnosticsCheck("backups", "Backup diagnostics", "major", {"safe", "sandbox", "live"}, _diag_check_backups),
        DiagnosticsCheck("telegram.integration", "Telegram integration", "major", {"live"}, _diag_check_telegram),
    ]


def _diag_set_run_status(run_id: int, status: str, summary: Dict[str, int], started_at: str) -> None:
    finished_at = iso_now(CFG.tzinfo())
    started = dt.datetime.fromisoformat(started_at)
    finished = dt.datetime.fromisoformat(finished_at)
    duration_ms = int((finished - started).total_seconds() * 1000)
    db_exec(
        """
        UPDATE diagnostic_runs
        SET status=?, finished_at=?, duration_ms=?, summary_json=?
        WHERE id=?;
        """,
        (status, finished_at, duration_ms, json.dumps(summary, ensure_ascii=False), int(run_id)),
    )


async def _run_diagnostics(run_id: int, suite: str, mode: str, options: DiagnosticsRunOptions) -> None:
    started_at = iso_now(CFG.tzinfo())
    ctx = DiagnosticsContext(run_id, suite, mode, options)
    checks = _build_checks(suite)
    final_status = "success"
    counters = {"success": 0, "warn": 0, "fail": 0, "skipped": 0}

    db_exec("UPDATE diagnostic_runs SET status='running', started_at=? WHERE id=?;", (started_at, int(run_id)))
    if mode in ("sandbox", "live"):
        sandbox_dir = Path(tempfile.mkdtemp(prefix=f"diag_{run_id}_"))
        ctx.sandbox_dir = sandbox_dir
        ctx.sandbox_db = sandbox_dir / "db.sqlite3"
        shutil.copy2(CFG.DB_PATH, ctx.sandbox_db)

    try:
        for chk in checks:
            if ctx.cancel_event.is_set():
                final_status = "canceled"
                break
            step_started = iso_now(CFG.tzinfo())
            step_id = db_exec_returning_id(
                """
                INSERT INTO diagnostic_steps (run_id, key, title, severity, status, started_at)
                VALUES (?, ?, ?, ?, 'running', ?);
                """,
                (int(run_id), chk.key, chk.title, chk.severity, step_started),
            )
            status = "skipped"
            message = "Skipped"
            details: Dict[str, Any] = {}
            trace = None
            try:
                if mode not in chk.allowed_modes:
                    status = "skipped"
                    message = f"Step is not allowed in {mode} mode"
                else:
                    result = await asyncio.wait_for(chk.fn(ctx), timeout=float(options.timeout_sec))
                    status = str(result.get("status", "success"))
                    message = str(result.get("message", ""))
                    details = result.get("details", {}) if isinstance(result.get("details", {}), dict) else {"value": result.get("details")}
            except Exception as exc:
                status = "fail"
                message = str(exc)
                trace = traceback.format_exc(limit=6)
                details = {"error": str(exc)}

            step_finished = iso_now(CFG.tzinfo())
            step_duration = int((dt.datetime.fromisoformat(step_finished) - dt.datetime.fromisoformat(step_started)).total_seconds() * 1000)
            if trace:
                details["trace"] = trace
            db_exec(
                """
                UPDATE diagnostic_steps
                SET status=?, finished_at=?, duration_ms=?, message=?, details_json=?
                WHERE id=?;
                """,
                (status, step_finished, step_duration, message[:1000], json.dumps(details, ensure_ascii=False), int(step_id)),
            )
            counters[status] = counters.get(status, 0) + 1
            if status == "fail":
                final_status = "fail"
            elif status == "warn" and final_status != "fail":
                final_status = "warn"
    finally:
        if final_status == "success" and counters.get("warn", 0) > 0:
            final_status = "warn"
        _diag_set_run_status(run_id, final_status, counters, started_at)
        if ctx.sandbox_dir:
            shutil.rmtree(ctx.sandbox_dir, ignore_errors=True)
        DIAG_RUNNING_TASKS.pop(run_id, None)
        DIAG_CANCEL_EVENTS.pop(run_id, None)


def _diag_fetch_run(run_id: int) -> Dict[str, Any]:
    row = db_fetchone("SELECT * FROM diagnostic_runs WHERE id=?;", (int(run_id),))
    if not row:
        raise HTTPException(status_code=404, detail="Diagnostics run not found")
    steps = db_fetchall("SELECT * FROM diagnostic_steps WHERE run_id=? ORDER BY id;", (int(run_id),))
    out = dict(row)
    out["options"] = _json_loads(out.pop("options_json", None), {})
    out["summary"] = _json_loads(out.pop("summary_json", None), {})
    out["steps"] = []
    for st in steps:
        item = dict(st)
        item["details"] = _json_loads(item.pop("details_json", None), {})
        out["steps"].append(item)
    return out


@APP.post("/api/admin/diagnostics/run")
async def api_diagnostics_run(
    body: DiagnosticsRunIn = Body(default_factory=DiagnosticsRunIn),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    suite = str(body.suite or "quick").lower()
    mode = str(body.mode or "safe").lower()
    if suite not in {"quick", "full", "integrations"}:
        raise HTTPException(status_code=400, detail="Invalid suite")
    if mode not in {"safe", "sandbox", "live"}:
        raise HTTPException(status_code=400, detail="Invalid mode")
    active = db_fetchone("SELECT id FROM diagnostic_runs WHERE status='running' ORDER BY id DESC LIMIT 1;")
    if active:
        raise HTTPException(status_code=409, detail="Diagnostics run already in progress")

    now = iso_now(CFG.tzinfo())
    user_id = int(u["id"])
    user_exists = db_fetchone("SELECT id FROM users WHERE id=?;", (user_id,))
    run_id = db_exec_returning_id(
        """
        INSERT INTO diagnostic_runs (created_at, created_by_user_id, suite, mode, status, options_json)
        VALUES (?, ?, ?, ?, 'queued', ?);
        """,
        (now, user_id if user_exists else None, suite, mode, json.dumps(body.options.model_dump(), ensure_ascii=False)),
    )
    DIAG_CANCEL_EVENTS[run_id] = asyncio.Event()
    task = asyncio.create_task(_run_diagnostics(run_id, suite, mode, body.options))
    DIAG_RUNNING_TASKS[run_id] = task
    return {"run_id": run_id, "status": "running"}


@APP.get("/api/admin/diagnostics/runs/{run_id}")
def api_diagnostics_run_get(run_id: int, u: sqlite3.Row = Depends(require_role("admin"))):
    return _diag_fetch_run(run_id)


@APP.get("/api/admin/diagnostics/runs")
def api_diagnostics_runs_list(
    limit: int = Query(50, ge=1, le=500),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    rows = db_fetchall("SELECT * FROM diagnostic_runs ORDER BY id DESC LIMIT ?;", (int(limit),))
    items = []
    for row in rows:
        r = dict(row)
        r["options"] = _json_loads(r.pop("options_json", None), {})
        r["summary"] = _json_loads(r.pop("summary_json", None), {})
        items.append(r)
    return {"items": items}


@APP.post("/api/admin/diagnostics/runs/{run_id}/cancel")
def api_diagnostics_cancel(run_id: int, u: sqlite3.Row = Depends(require_role("admin"))):
    row = db_fetchone("SELECT * FROM diagnostic_runs WHERE id=?;", (int(run_id),))
    if not row:
        raise HTTPException(status_code=404, detail="Diagnostics run not found")
    ev = DIAG_CANCEL_EVENTS.get(int(run_id))
    if ev:
        ev.set()
    return {"ok": True}


@APP.get("/api/admin/diagnostics/runs/{run_id}/download")
def api_diagnostics_download(
    run_id: int,
    format: str = Query("json"),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    payload = _diag_fetch_run(run_id)
    fmt = str(format or "json").lower()
    if fmt == "json":
        return JSONResponse(payload)
    if fmt != "html":
        raise HTTPException(status_code=400, detail="Unsupported format")
    html = [
        "<html><head><meta charset='utf-8'><title>Diagnostics report</title></head><body>",
        f"<h1>Diagnostics run #{payload['id']}</h1>",
        f"<p>Status: <b>{payload['status']}</b></p>",
        "<ul>",
    ]
    for st in payload.get("steps", []):
        html.append(f"<li><b>{st.get('status')}</b> {st.get('title')} — {st.get('message') or ''}</li>")
    html.append("</ul></body></html>")
    return Response("".join(html), media_type="text/html; charset=utf-8")


@APP.post("/api/admin/system/full-test")
async def api_admin_full_test(u: sqlite3.Row = Depends(require_role("admin"))):
    options = DiagnosticsRunOptions(timeout_sec=60)
    ctx = DiagnosticsContext(0, "quick", "safe", options)
    checks = [
        ("database", _diag_check_db_integrity),
        ("settings", _diag_check_preflight),
        ("scheduler", _diag_check_scheduler),
        ("core_data", _diag_check_exports),
        ("storage", _diag_check_backups),
    ]
    out = []
    worst = "ok"
    for name, fn in checks:
        res = await fn(ctx)
        status = str(res.get("status", "success"))
        if status == "fail":
            worst = "fail"
        elif status == "warn" and worst != "fail":
            worst = "warn"
        out.append({"name": name, "status": status, "message": res.get("message", "")})
    return {"status": worst, "checks": out}


# ---------------------------
# Reports (manual trigger via API)
# ---------------------------

@APP.post("/api/reports/sunday")
async def api_report_sunday(u: sqlite3.Row = Depends(require_role("admin", "accountant"))):
    s = get_settings()
    recipients = list_report_recipients(s)
    if not recipients:
        raise HTTPException(status_code=400, detail="No report recipients configured")

    tzinfo = ZoneInfo(str(s["timezone"] or CFG.TZ))
    today = dt.datetime.now(tzinfo).date()
    await send_sunday_reports_bundle(today, recipients, raise_on_error=True)
    return {"ok": True}



@APP.post("/api/reports/month_expenses")
async def api_report_month_expenses(u: sqlite3.Row = Depends(require_role("admin", "accountant"))):
    s = get_settings()
    recipients = list_report_recipients(s)
    if not recipients:
        raise HTTPException(status_code=400, detail="No report recipients configured")

    tzinfo = ZoneInfo(str(s["timezone"] or CFG.TZ))
    today = dt.datetime.now(tzinfo).date()
    text, kb = build_month_expenses_report_text(today)
    await send_report_to_recipients(text, kb, recipients, raise_on_error=False, kind="report")
    return {"ok": True}

@APP.post("/api/reports/test")
async def api_report_test(u: sqlite3.Row = Depends(require_role("admin", "accountant"))):
    s = get_settings()
    recipients = list_report_recipients(s)
    if not recipients:
        raise HTTPException(status_code=400, detail="No report recipients configured")

    tzinfo = ZoneInfo(str(s["timezone"] or CFG.TZ))
    now = dt.datetime.now(tzinfo)
    text = (
        "✅ Тестовый отчёт\n"
        f"Если вы это видите, доставка работает.\n"
        f"Время: {now:%Y-%m-%d %H:%M:%S %Z}"
    )
    await send_report_to_recipients(text, None, recipients, raise_on_error=False, kind="report")
    return {"ok": True}

# ---------------------------
# Backups
# ---------------------------

@APP.get("/api/backups")
def api_list_backups(u: sqlite3.Row = Depends(require_role("admin"))):
    tzinfo = backups_tzinfo()
    return list_backups(tzinfo)


@APP.post("/api/backups/run")
def api_run_backup(
    backup_type: str = Query("db"),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    if backup_type not in ("db", "full"):
        raise HTTPException(status_code=400, detail="Invalid backup type")
    tzinfo = backups_tzinfo()
    if backup_type == "db":
        path = create_db_backup(tzinfo)
    else:
        path = create_full_backup(tzinfo)
    stat = path.stat()
    return {
        "name": path.name,
        "type": backup_type,
        "size_bytes": stat.st_size,
        "created_at": dt.datetime.now(tzinfo).isoformat(),
    }


@APP.get("/api/backups/{name}/download")
def api_download_backup(
    name: str,
    u: sqlite3.Row = Depends(require_role("admin")),
):
    ensure_backups_dir()
    if "/" in name or "\\" in name:
        raise HTTPException(status_code=404, detail="Backup not found")
    path = (BACKUPS_DIR / name).resolve()
    if path.parent != BACKUPS_DIR.resolve() or not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Backup not found")
    return FileResponse(path, media_type="application/octet-stream", filename=name)


@APP.post("/api/backups/restore")
async def api_restore_backup(
    file: UploadFile = File(...),
    u: sqlite3.Row = Depends(require_role("admin")),
):
    if not file:
        raise HTTPException(status_code=400, detail="Missing file")
    tzinfo = backups_tzinfo()
    temp_dir = Path(tempfile.mkdtemp(prefix="restore_"))
    temp_file = temp_dir / (file.filename or "backup_upload")
    backup_db_path = None
    try:
        with temp_file.open("wb") as f:
            total_written = 0
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                total_written += len(chunk)
                if total_written > MAX_BACKUP_UPLOAD_BYTES:
                    raise HTTPException(status_code=413, detail="Backup file is too large")
                f.write(chunk)
        if temp_file.stat().st_size == 0:
            raise HTTPException(status_code=400, detail="Empty file")

        candidate_db = temp_file
        uploads_dir = None
        if zipfile.is_zipfile(temp_file):
            extract_dir = temp_dir / "extract"
            extract_dir.mkdir()
            with zipfile.ZipFile(temp_file, "r") as zf:
                if "db.sqlite3" not in zf.namelist():
                    raise HTTPException(status_code=400, detail="Archive missing db.sqlite3")
                safe_extract_zip(zf, extract_dir)
            candidate_db = extract_dir / "db.sqlite3"
            uploads_dir = extract_dir / "uploads"

        if not candidate_db.exists() or not validate_sqlite_file(candidate_db):
            raise HTTPException(status_code=400, detail="Invalid sqlite3 backup")

        current_db = Path(CFG.DB_PATH)
        stamp = dt.datetime.now(tzinfo).strftime("%Y%m%d_%H%M%S")
        backup_db_path = current_db.with_name(f"{current_db.name}.before_restore.{stamp}")

        if current_db.exists():
            os.replace(current_db, backup_db_path)

        temp_restore = current_db.with_name(f".restore_tmp_{stamp}.sqlite3")
        shutil.copy2(candidate_db, temp_restore)
        os.replace(temp_restore, current_db)
        init_db()

        if uploads_dir is not None and uploads_dir.exists() and uploads_dir.is_dir():
            uploads_backup = UPLOADS_DIR.with_name(f"{UPLOADS_DIR.name}.before_restore.{stamp}")
            if UPLOADS_DIR.exists():
                if uploads_backup.exists():
                    shutil.rmtree(uploads_backup)
                shutil.move(str(UPLOADS_DIR), str(uploads_backup))
            shutil.copytree(uploads_dir, UPLOADS_DIR)

        return {"ok": True}
    except HTTPException:
        raise
    except Exception:
        if backup_db_path and backup_db_path.exists() and not Path(CFG.DB_PATH).exists():
            try:
                os.replace(backup_db_path, CFG.DB_PATH)
            except Exception:
                pass
        raise HTTPException(status_code=400, detail="Backup restore failed")
    finally:
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass



# ---------------------------
# Export
# ---------------------------

PNG_PRESETS = {
    "landscape": (1600, 900),
    "square": (1080, 1080),
    "story": (1080, 1920),
}

RU_MONTHS = [
    "января",
    "февраля",
    "марта",
    "апреля",
    "мая",
    "июня",
    "июля",
    "августа",
    "сентября",
    "октября",
    "ноября",
    "декабря",
]


def require_pillow() -> Tuple[Any, Any, Any]:
    if importlib.util.find_spec("PIL") is None:
        raise HTTPException(status_code=501, detail="Pillow (PIL) is required for PNG export")
    from PIL import Image, ImageDraw, ImageFont

    return Image, ImageDraw, ImageFont


def load_ttf_font(image_font: Any, size: int, bold: bool = False) -> Any:
    """
    Пытаемся грузить красивый UI-шрифт (Menlo/Inter, если положишь рядом),
    иначе системные DejaVu/Liberation.
    """
    base_dir = Path(__file__).resolve().parent

    local_candidates = [
        base_dir / ("Menlo-Bold.ttf" if bold else "Menlo-Regular.ttf"),
        base_dir / "Menlo.ttf",
        base_dir / ("Inter-Bold.ttf" if bold else "Inter-Regular.ttf"),
        base_dir / ("Inter-SemiBold.ttf" if bold else "Inter-Regular.ttf"),
        base_dir / ("SF-Pro-Display-Bold.ttf" if bold else "SF-Pro-Display-Regular.ttf"),
    ]

    sys_candidates = [
        "/Library/Fonts/Menlo.ttc",
        "/Library/Fonts/Menlo.ttf",
        "/System/Library/Fonts/Menlo.ttc",
        "/System/Library/Fonts/Menlo.ttf",
        "/usr/share/fonts/truetype/menlo/Menlo-Bold.ttf" if bold else "/usr/share/fonts/truetype/menlo/Menlo-Regular.ttf",
        "/usr/share/fonts/truetype/menlo/Menlo.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf" if bold else "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
    ]

    for p in [*local_candidates, *sys_candidates]:
        try:
            p_str = str(p)
            if p_str.startswith("/") and not Path(p_str).exists():
                continue
            return image_font.truetype(p_str, size=size)
        except Exception:
            continue

    return image_font.load_default()



def text_bbox(draw: Any, text: str, font: Any) -> Tuple[int, int]:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def fit_text_ellipsis(draw: Any, text: str, font: Any, max_width: int) -> str:
    if max_width <= 0:
        return ""
    if text_bbox(draw, text, font)[0] <= max_width:
        return text
    ellipsis = "…"
    if text_bbox(draw, ellipsis, font)[0] > max_width:
        return ""

    lo, hi = 0, len(text)
    while lo < hi:
        mid = (lo + hi + 1) // 2
        candidate = text[:mid].rstrip() + ellipsis
        if text_bbox(draw, candidate, font)[0] <= max_width:
            lo = mid
        else:
            hi = mid - 1
    return text[:lo].rstrip() + ellipsis


def find_max_font_size_for_rows(
    image_font: Any,
    draw: Any,
    rows: List[Tuple[str, str]],
    max_label_w: int,
    max_value_w: int,
    max_row_h: int,
    min_size: int,
    max_size: int,
) -> int:
    for size in range(max_size, min_size - 1, -1):
        font = load_ttf_font(image_font, size=size, bold=False)
        fits_height = True
        fits_width = True
        for label, value in rows:
            _, label_h = text_bbox(draw, label, font)
            _, value_h = text_bbox(draw, value, font)
            if max(label_h, value_h) > max_row_h:
                fits_height = False
                break
            if text_bbox(draw, label, font)[0] > max_label_w or text_bbox(draw, value, font)[0] > max_value_w:
                fits_width = False
        if fits_height and fits_width:
            return size
    return min_size


def draw_card(
    img: Any,
    xy: Tuple[int, int, int, int],
    radius: int,
    fill: Tuple[int, int, int],
    shadow_alpha: int = 40,
    shadow_blur: int = 18,
    shadow_offset: Tuple[int, int] = (0, 10),
    outline: Optional[Tuple[int, int, int]] = None,
    outline_width: int = 0,
) -> None:
    """
    Рисует карточку с мягкой тенью (GaussianBlur) на RGBA-канвасе.
    img должен быть RGBA.
    """
    try:
        from PIL import Image, ImageDraw, ImageFilter
    except Exception:
        # fallback на старое поведение без blur
        draw = img  # если кто-то случайно передал draw
        x0, y0, x1, y1 = xy
        draw.rounded_rectangle((x0, y0, x1, y1), radius=radius, fill=fill)
        return

    x0, y0, x1, y1 = map(int, xy)
    w = max(1, x1 - x0)
    h = max(1, y1 - y0)

    pad = shadow_blur * 2
    layer_w = w + pad * 2
    layer_h = h + pad * 2

    # shadow layer (локальный — быстрее, чем на весь холст)
    shadow = Image.new("RGBA", (layer_w, layer_h), (0, 0, 0, 0))
    sd = ImageDraw.Draw(shadow)
    sd.rounded_rectangle(
        (pad, pad, pad + w, pad + h),
        radius=radius,
        fill=(0, 0, 0, max(0, min(255, shadow_alpha))),
    )
    shadow = shadow.filter(ImageFilter.GaussianBlur(shadow_blur))

    ox, oy = shadow_offset
    img.alpha_composite(shadow, dest=(x0 - pad + ox, y0 - pad + oy))

    # card body
    card = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    cd = ImageDraw.Draw(card)
    cd.rounded_rectangle(
        (0, 0, w, h),
        radius=radius,
        fill=(*fill, 255),
        outline=(*outline, 255) if outline else None,
        width=int(outline_width) if outline and outline_width else 0,
    )
    img.alpha_composite(card, dest=(x0, y0))



def draw_pill(
    draw: Any,
    xy: Tuple[int, int],
    text: str,
    font: Any,
    fill: Tuple[int, int, int],
    text_color: Tuple[int, int, int],
    padding: Tuple[int, int] = (12, 6),
    radius: int = 12,
    outline: Optional[Tuple[int, int, int]] = None,
    outline_width: int = 0,
) -> Tuple[int, int]:
    x, y = xy
    text_w, text_h = text_bbox(draw, text, font)
    pad_x, pad_y = padding
    pill_w = text_w + pad_x * 2
    pill_h = text_h + pad_y * 2

    draw.rounded_rectangle((x, y, x + pill_w, y + pill_h), radius=radius, fill=fill)
    if outline and outline_width > 0:
        draw.rounded_rectangle((x, y, x + pill_w, y + pill_h), radius=radius, outline=outline, width=outline_width)

    draw.text((x + pad_x, y + pad_y), text, font=font, fill=text_color)
    return pill_w, pill_h



def render_month_report_png(
    month_row: sqlite3.Row,
    summary: Dict[str, Any],
    services: List[sqlite3.Row],
    top_categories: List[Dict[str, Any]],
    expenses: List[sqlite3.Row],
    subaccount_services: List[sqlite3.Row],
    subaccount_expenses: List[sqlite3.Row],
    preset: str,
    pixel_ratio: int = 2,   # <-- плотность пикселей (1..4)
    dpi: int = 192,         # <-- DPI метаданные (72..600)
) -> bytes:
    Image, ImageDraw, ImageFont = require_pillow()

    # размеры делаем "ретина"
    base_w, base_h = PNG_PRESETS[preset]
    pixel_ratio = int(max(1, min(4, pixel_ratio)))
    w, h = int(base_w * pixel_ratio), int(base_h * pixel_ratio)
    scale = w / 1600.0

    # палитра (чуть спокойнее/дороже)
    color_bg = (243, 246, 251)
    color_card = (255, 255, 255)
    color_card2 = (241, 245, 249)
    color_text = (15, 23, 42)
    color_muted = (71, 85, 105)
    color_muted2 = (100, 116, 139)
    color_stroke = (226, 232, 240)

    color_accent = (16, 185, 129)        # green
    color_accent_soft = (209, 250, 229)

    color_danger = (239, 68, 68)         # red
    color_danger_soft = (254, 226, 226)

    color_warn = (245, 158, 11)          # amber
    color_warn_soft = (254, 243, 199)

    # типографика
    font_title = load_ttf_font(ImageFont, size=int(44 * scale), bold=True)
    font_kpi_value = load_ttf_font(ImageFont, size=int(34 * scale), bold=True)
    font_kpi_label = load_ttf_font(ImageFont, size=int(16 * scale), bold=False)
    font_section = load_ttf_font(ImageFont, size=int(20 * scale), bold=True)
    font_body = load_ttf_font(ImageFont, size=int(16 * scale), bold=False)
    font_small = load_ttf_font(ImageFont, size=int(13 * scale), bold=False)
    font_bar_value = load_ttf_font(ImageFont, size=int(15 * scale), bold=True)

    margin = int(44 * scale)
    gap = int(20 * scale)
    radius = int(18 * scale)

    # dynamic height for text-heavy blocks (plan + expense structure)
    plan_items = [
        ("МНСП", fmt_money(float(summary.get("monthly_min_needed") or 0.0))),
        ("Выполнение МНСП", fmt_percent_1(float(summary.get("monthly_completion") or 0.0))),
        ("СДДР", fmt_money(float(summary.get("sddr") or 0.0))),
        ("К прошлому месяцу", f"{float(summary.get('psdpm') or 0.0) * 100:.1f}%"),
        ("Среднее пожертвование", fmt_money(float(summary.get("avg_sunday") or 0.0))),
    ]

    kpi_y = margin + int(64 * scale)
    kpi_h = int(148 * scale)
    section_y = kpi_y + kpi_h + int(30 * scale)
    footer_h = int(36 * scale)
    left_w = int(w * 0.40)

    tmp_img = Image.new("RGBA", (32, 32), (0, 0, 0, 0))
    tmp_draw = ImageDraw.Draw(tmp_img)
    progress_font = load_ttf_font(ImageFont, size=max(11, int(13 * scale)), bold=False)
    _, progress_label_h = text_bbox(tmp_draw, "Прогресс", progress_font)
    body_line_h = text_bbox(tmp_draw, "Ag", font_body)[1]

    plan_rows_h = len(plan_items) * max(body_line_h + int(6 * scale), int(20 * scale))
    plan_min_h = int(56 * scale) + plan_rows_h + int(12 * scale) + progress_label_h + max(int(4 * scale), int(6 * scale)) + max(8, int(10 * scale)) + int(18 * scale)
    plan_target_h = max(int(252 * scale), plan_min_h)

    legend_rows = min(6, max(1, len([c for c in top_categories if float(c.get("sum") or 0.0) > 0.0])))
    legend_h = int(8 * scale) + legend_rows * int(24 * scale)
    expenses_target_h = max(int(220 * scale), int(52 * scale) + legend_h + int(16 * scale))

    top_h_min = max(int(280 * scale), plan_target_h + gap + expenses_target_h)
    list_card_min_h = max(int(180 * scale), int(220 * scale))
    required_content_h = top_h_min + gap + list_card_min_h
    required_h = section_y + required_content_h + footer_h + margin
    if required_h > h:
        h = required_h

    # RGBA холст для мягких теней
    bg = Image.new("RGBA", (w, h), (*color_bg, 255))
    draw = ImageDraw.Draw(bg)

    # title
    month_name = RU_MONTHS[int(month_row["month"]) - 1]
    title = f"Отчёт за {month_name} {int(month_row['year'])}"
    draw.text((margin, margin), title, font=font_title, fill=color_text)

    # KPI
    kpi_w = int((w - margin * 2 - gap * 3) / 4)

    income = float(summary.get("month_income_sum") or 0.0)
    spend = float(summary.get("month_expenses_sum") or 0.0)
    bal = float(summary.get("month_balance") or 0.0)
    fbal = float(summary.get("fact_balance") or 0.0)

    kpis = [
        ("Доход", fmt_money(income), color_accent, color_accent_soft),
        ("Расход", fmt_money(spend), color_danger, color_danger_soft),
        ("Баланс месяца", fmt_money(bal), (color_accent if bal >= 0 else color_danger), (color_accent_soft if bal >= 0 else color_danger_soft)),
        ("Факт. баланс", fmt_money(fbal), (color_accent if fbal >= 0 else color_danger), (color_accent_soft if fbal >= 0 else color_danger_soft)),
    ]

    for idx, (label, value, kcol, ksoft) in enumerate(kpis):
        x0 = margin + idx * (kpi_w + gap)
        card = (x0, kpi_y, x0 + kpi_w, kpi_y + kpi_h)
        is_fact_balance = (label == "Факт. баланс")
        card_fill = (245, 255, 250) if is_fact_balance and fbal >= 0 else ((255, 246, 246) if is_fact_balance else color_card)
        card_outline = kcol if is_fact_balance else None
        card_outline_width = max(2, int(2 * scale)) if is_fact_balance else 0
        draw_card(
            bg, card, radius, card_fill,
            shadow_alpha=38, shadow_blur=int(18 * scale), shadow_offset=(0, int(10 * scale)),
            outline=card_outline, outline_width=card_outline_width,
        )

        # маленький акцентный индикатор слева
        stripe_w = max(3, int(4 * scale))
        draw.rounded_rectangle(
            (x0 + int(14 * scale), kpi_y + int(18 * scale), x0 + int(14 * scale) + stripe_w, kpi_y + kpi_h - int(18 * scale)),
            radius=int(6 * scale),
            fill=kcol,
        )

        draw.text((x0 + int(26 * scale), kpi_y + int(18 * scale)), label, font=font_kpi_label, fill=color_muted2)
        draw.text((x0 + int(26 * scale), kpi_y + int(58 * scale)), value, font=font_kpi_value, fill=color_text)

        # pill справа сверху (мягкая)
        pill_text = "₽"
        pill_w, pill_h = draw_pill(
            draw,
            (x0 + kpi_w - int(56 * scale), kpi_y + int(16 * scale)),
            pill_text,
            font_small,
            fill=ksoft,
            text_color=kcol,
            radius=int(12 * scale),
        )

    # Layout columns
    content_h = h - section_y - footer_h - margin

    left_w = int(w * 0.40)
    right_w = w - margin * 2 - gap - left_w
    left_x = margin
    right_x = left_x + left_w + gap

    list_card_h = min(int(260 * scale), max(int(180 * scale), int(content_h * 0.30)))
    top_h = max(content_h - list_card_h - gap, int(280 * scale))

    # --- PLAN CARD (с прогресс-баром) ---
    plan_h = max(plan_target_h, int(top_h * 0.52))
    plan_card = (left_x, section_y, left_x + left_w, section_y + plan_h)
    draw_card(bg, plan_card, radius, color_card, shadow_alpha=38, shadow_blur=int(18 * scale), shadow_offset=(0, int(10 * scale)))
    draw.text((left_x + int(18 * scale), section_y + int(16 * scale)), "План/цели", font=font_section, fill=color_text)

    psdpm = summary.get("psdpm")
    psdpm_text = f"{float(psdpm)*100:.1f}%" if isinstance(psdpm, (int, float)) else "—"
    completion_val = float(summary.get("monthly_completion") or 0.0)  # 0..1
    completion_txt = fmt_percent_1(completion_val)

    plan_items = [
        ("МНСП", fmt_money(float(summary.get("monthly_min_needed") or 0.0))),
        ("Выполнение МНСП", completion_txt),
        ("СДДР", fmt_money(float(summary.get("sddr") or 0.0))),
        ("К прошлому месяцу", psdpm_text),
        ("Среднее пожертвование", fmt_money(float(summary.get("avg_sunday") or 0.0))),
    ]

    # progress bar
    bar_x0 = left_x + int(18 * scale)
    bar_w = left_w - int(36 * scale)
    bar_h = max(8, int(10 * scale))
    r = bar_h // 2

    progress_font = load_ttf_font(ImageFont, size=max(11, int(13 * scale)), bold=False)
    progress_label_w, progress_label_h = text_bbox(draw, "Прогресс", progress_font)

    plan_top_y = section_y + int(56 * scale)
    plan_inner_right = left_x + left_w - int(18 * scale)
    plan_bottom_padding = int(18 * scale)
    bar_gap_top = max(int(4 * scale), int(6 * scale))
    bar_gap_bottom = int(12 * scale)
    bar_y0 = plan_card[3] - plan_bottom_padding - bar_h
    progress_y = max(plan_top_y, bar_y0 - bar_gap_top - progress_label_h)

    rows_area_bottom = progress_y - bar_gap_bottom
    plan_available_h = max(int(20 * scale), rows_area_bottom - plan_top_y)
    plan_row_h = max(int(18 * scale), int(plan_available_h / max(1, len(plan_items))))

    max_label_w = max(int(80 * scale), int(bar_w * 0.58))
    max_value_w = max(int(72 * scale), bar_w - max_label_w - int(12 * scale))

    plan_font_size = find_max_font_size_for_rows(
        ImageFont,
        draw,
        plan_items,
        max_label_w=max_label_w,
        max_value_w=max_value_w,
        max_row_h=plan_row_h,
        min_size=max(10, int(11 * scale)),
        max_size=max(12, int(16 * scale)),
    )
    plan_font = load_ttf_font(ImageFont, size=plan_font_size, bold=False)

    for idx, (label, value) in enumerate(plan_items):
        line_top = plan_top_y + idx * plan_row_h
        label_text = fit_text_ellipsis(draw, label, plan_font, max_label_w)
        value_text = fit_text_ellipsis(draw, value, plan_font, max_value_w)

        _, label_h = text_bbox(draw, label_text or " ", plan_font)
        _, value_h = text_bbox(draw, value_text or " ", plan_font)
        line_y = line_top + max(0, int((plan_row_h - max(label_h, value_h)) / 2))

        draw.text((bar_x0, line_y), label_text, font=plan_font, fill=color_muted)
        value_w, _ = text_bbox(draw, value_text, plan_font)
        draw.text((plan_inner_right - value_w, line_y), value_text, font=plan_font, fill=color_text)

    draw.text((bar_x0, progress_y), "Прогресс", font=progress_font, fill=color_muted2)
    draw.rounded_rectangle((bar_x0, bar_y0, bar_x0 + bar_w, bar_y0 + bar_h), radius=r, fill=color_card2)
    fill_w = int(bar_w * max(0.0, min(1.0, completion_val)))
    if fill_w > 0:
        draw.rounded_rectangle((bar_x0, bar_y0, bar_x0 + fill_w, bar_y0 + bar_h), radius=r, fill=color_accent)

    # --- EXPENSE STRUCTURE (DONUT) ---
    expenses_card_y = section_y + plan_h + gap
    expenses_card = (left_x, expenses_card_y, left_x + left_w, section_y + top_h)
    draw_card(
        bg, expenses_card, radius, color_card,
        shadow_alpha=38, shadow_blur=int(18 * scale), shadow_offset=(0, int(10 * scale)),
        outline=color_stroke, outline_width=max(1, int(1 * scale)),
    )
    draw.text((left_x + int(18 * scale), expenses_card_y + int(16 * scale)), "Структура расходов", font=font_section, fill=color_text)

    draw_pill(
        draw,
        (left_x + left_w - int(160 * scale), expenses_card_y + int(14 * scale)),
        "Основной счёт",
        font_small,
        fill=color_accent_soft,
        text_color=color_accent,
        radius=int(12 * scale),
    )

    pie_pad = int(18 * scale)
    pie_top = expenses_card_y + int(52 * scale)
    pie_area_h = expenses_card[3] - pie_top - int(16 * scale)

    pie_size = min(int(left_w * 0.46), pie_area_h)
    pie_x0 = left_x + pie_pad
    pie_y0 = pie_top + int((pie_area_h - pie_size) / 2)
    pie_box = (pie_x0, pie_y0, pie_x0 + pie_size, pie_y0 + pie_size)

    total_expenses = float(summary.get("month_expenses_sum") or 0.0)

    palette = [
        (37, 99, 235),   # blue
        (16, 185, 129),  # green
        (245, 158, 11),  # amber
        (239, 68, 68),   # red
        (139, 92, 246),  # violet
        (14, 165, 233),  # sky
    ]

    cats_nonzero = [c for c in top_categories if float(c.get("sum") or 0.0) > 0.0]

    if total_expenses <= 0 or not cats_nonzero:
        msg = "Нет данных"
        msg_w, msg_h = text_bbox(draw, msg, font_body)
        draw.text((pie_x0 + (pie_size - msg_w) / 2, pie_y0 + (pie_size - msg_h) / 2), msg, font=font_body, fill=color_muted2)
    else:
        start_angle = -90
        for idx, item in enumerate(cats_nonzero[:6]):
            value = float(item["sum"])
            sweep = 360.0 * (value / total_expenses)
            draw.pieslice(pie_box, start=start_angle, end=start_angle + sweep, fill=palette[idx % len(palette)])
            start_angle += sweep

        # donut hole
        cx = pie_x0 + pie_size / 2
        cy = pie_y0 + pie_size / 2
        hole = pie_size * 0.66
        draw.ellipse((cx - hole / 2, cy - hole / 2, cx + hole / 2, cy + hole / 2), fill=color_card)

        # center text
        center_label = "Итого"
        center_value = fmt_money_commas(total_expenses)
        lw, lh = text_bbox(draw, center_label, font_small)
        center_font = font_small
        min_center_font_size = max(8, int(10 * scale))
        for size in range(max(int(14 * scale), min_center_font_size), min_center_font_size - 1, -1):
            candidate = load_ttf_font(ImageFont, size=size, bold=True)
            cw, ch = text_bbox(draw, center_value, candidate)
            if cw <= hole * 0.84 and ch <= hole * 0.44:
                center_font = candidate
                break
        vw, vh = text_bbox(draw, center_value, center_font)
        draw.text((cx - lw / 2, cy - (lh + vh) / 2 - int(2 * scale)), center_label, font=font_small, fill=color_muted2)
        draw.text((cx - vw / 2, cy - (lh + vh) / 2 + lh + int(1 * scale)), center_value, font=center_font, fill=color_text)

    # legend
    legend_x = pie_x0 + pie_size + int(16 * scale)
    legend_y = pie_top + int(8 * scale)
    max_legend = 6

    shown = cats_nonzero[:max_legend]
    if shown:
        for idx, item in enumerate(shown):
            val = float(item["sum"])
            pct = (val / total_expenses * 100.0) if total_expenses > 0 else 0.0
            label = f"{item.get('category','—')}: {fmt_money(val)} • {pct:.0f}%"
            color = palette[idx % len(palette)]
            y = legend_y + idx * int(24 * scale)
            draw.rounded_rectangle((legend_x, y + int(4 * scale), legend_x + int(12 * scale), y + int(16 * scale)), radius=int(4 * scale), fill=color)
            draw.text((legend_x + int(18 * scale), y), label, font=font_small, fill=color_text)

    # --- INCOME CHART ---
    chart_card = (right_x, section_y, right_x + right_w, section_y + top_h)
    draw_card(
        bg, chart_card, radius, color_card,
        shadow_alpha=38, shadow_blur=int(18 * scale), shadow_offset=(0, int(10 * scale)),
        outline=color_stroke, outline_width=max(1, int(1 * scale)),
    )
    draw.text((right_x + int(18 * scale), section_y + int(16 * scale)), "Доходы по служениям", font=font_section, fill=color_text)
    draw_pill(
        draw,
        (right_x + right_w - int(160 * scale), section_y + int(14 * scale)),
        "Основной счёт",
        font_small,
        fill=color_accent_soft,
        text_color=color_accent,
        radius=int(12 * scale),
    )

    chart_pad = int(24 * scale)
    chart_x0 = right_x + chart_pad
    chart_y0 = section_y + int(56 * scale)
    chart_w = right_w - chart_pad * 2
    sub_table_h = min(int(150 * scale), int(top_h * 0.34))
    chart_h = top_h - sub_table_h - int(92 * scale)

    base_y = chart_y0 + chart_h - int(18 * scale)
    plot_h = chart_h - int(44 * scale)

    service_items = []
    for s in services:
        total = float(s["total"] or 0.0)
        try:
            service_date = dt.date.fromisoformat(str(s["service_date"]))
        except Exception:
            continue
        weekly_min_for_service = float(s["weekly_min_needed"] or 0.0)
        service_items.append(
            {
                "date": service_date,
                "total": total,
                "status": str(s["mnsps_status"] or ""),
                "weekly_min_needed": weekly_min_for_service,
                "is_collected": weekly_min_for_service <= 0 or total >= weekly_min_for_service,
            }
        )
    max_total = max([it["total"] for it in service_items], default=0.0)

    # grid
    for p in (0.25, 0.50, 0.75):
        gy = base_y - int(plot_h * p)
        draw.line((chart_x0, gy, chart_x0 + chart_w, gy), fill=color_stroke, width=max(1, int(1 * scale)))

    if max_total <= 0:
        msg = "Нет данных"
        msg_w, msg_h = text_bbox(draw, msg, font_body)
        draw.text((chart_x0 + (chart_w - msg_w) / 2, chart_y0 + (chart_h - msg_h) / 2), msg, font=font_body, fill=color_muted2)
    else:
        n = max(1, len(service_items))
        bar_gap = max(6, int(10 * scale))
        bar_w = max(10, int((chart_w - bar_gap * (n + 1)) / n))
        bar_r = max(4, int(min(bar_w * 0.22, 12 * scale)))

        for idx, item in enumerate(service_items):
            bar_x = chart_x0 + bar_gap + idx * (bar_w + bar_gap)
            bar_h = int((item["total"] / max_total) * plot_h)

            status_ok = bool(item.get("is_collected"))
            bar_color = color_accent if status_ok else color_danger

            draw.rounded_rectangle(
                (bar_x, base_y - bar_h, bar_x + bar_w, base_y),
                radius=bar_r,
                fill=bar_color,
            )

            value_label = fmt_money(item["total"])
            val_w, val_h = text_bbox(draw, value_label, font_bar_value)
            val_y = max(chart_y0 + int(6 * scale), base_y - bar_h - val_h - int(4 * scale))
            draw.text((bar_x + (bar_w - val_w) / 2, val_y), value_label, font=font_bar_value, fill=color_text)

            label = item["date"].strftime("%d.%m")
            lw, lh = text_bbox(draw, label, font_small)
            draw.text((bar_x + (bar_w - lw) / 2, base_y + int(6 * scale)), label, font=font_small, fill=color_muted2)

        # weekly min (dashed)
        weekly_min = float(summary.get("weekly_min_needed") or 0.0)
        if weekly_min > 0:
            line_y = base_y - int((min(weekly_min, max_total) / max_total) * plot_h)
            dash = max(8, int(12 * scale))
            gap2 = max(6, int(10 * scale))
            x = chart_x0
            while x < chart_x0 + chart_w:
                x2 = min(chart_x0 + chart_w, x + dash)
                draw.line((x, line_y, x2, line_y), fill=color_warn, width=max(2, int(2 * scale)))
                x += dash + gap2
            draw.text((chart_x0 + int(4 * scale), line_y - int(18 * scale)), "МНСП", font=font_small, fill=color_warn)

    # --- SUBACCOUNTS TABLE ---
    sub_table_x0 = right_x + int(18 * scale)
    sub_table_x1 = right_x + right_w - int(18 * scale)
    sub_table_y0 = chart_y0 + chart_h + int(30 * scale)
    sub_table_y1 = section_y + top_h - int(16 * scale)

    if sub_table_y1 > sub_table_y0:
        draw.text((sub_table_x0, sub_table_y0 - int(22 * scale)), "Доп. счета", font=font_section, fill=color_text)
        header_h = int(22 * scale)
        draw.rounded_rectangle((sub_table_x0, sub_table_y0, sub_table_x1, sub_table_y0 + header_h), radius=int(10 * scale), fill=color_card2)

        col_date_w = int((sub_table_x1 - sub_table_x0) * 0.30)
        col_w = int((sub_table_x1 - sub_table_x0 - col_date_w) / 2)

        draw.text((sub_table_x0 + int(8 * scale), sub_table_y0 + int(3 * scale)), "Дата", font=font_small, fill=color_muted)
        draw.text((sub_table_x0 + col_date_w + int(8 * scale), sub_table_y0 + int(3 * scale)), "Praise +", font=font_small, fill=color_muted)
        draw.text((sub_table_x0 + col_date_w + col_w + int(8 * scale), sub_table_y0 + int(3 * scale)), "Alpha +", font=font_small, fill=color_muted)

        sub_income: Dict[dt.date, Dict[str, float]] = {}
        for row in subaccount_services:
            try:
                d = dt.date.fromisoformat(str(row["service_date"]))
            except Exception:
                continue
            acc = str(row["account"])
            total = float(row["total"] or 0.0)
            sub_income.setdefault(d, {}).setdefault(acc, 0.0)
            sub_income[d][acc] += total

        sub_spend: Dict[str, float] = {"praise": 0.0, "alpha": 0.0}
        for row in subaccount_expenses:
            acc = str(row["account"])
            if acc in sub_spend:
                sub_spend[acc] += float(row["total"] or 0.0)

        dates = sorted(sub_income.keys())
        row_h = int(20 * scale)
        y = sub_table_y0 + header_h + int(8 * scale)

        if not dates:
            draw.text((sub_table_x0 + int(8 * scale), y), "Нет данных", font=font_small, fill=color_muted2)
            y += row_h
        else:
            for i, d in enumerate(dates):
                if y + row_h > sub_table_y1 - row_h:
                    break
                if i % 2 == 0:
                    draw.rounded_rectangle((sub_table_x0, y - int(2 * scale), sub_table_x1, y + row_h - int(2 * scale)), radius=int(8 * scale), fill=(248, 250, 253))
                draw.text((sub_table_x0 + int(8 * scale), y), d.strftime("%d.%m"), font=font_small, fill=color_text)
                draw.text((sub_table_x0 + col_date_w + int(8 * scale), y), fmt_money(sub_income[d].get("praise", 0.0)), font=font_small, fill=color_text)
                draw.text((sub_table_x0 + col_date_w + col_w + int(8 * scale), y), fmt_money(sub_income[d].get("alpha", 0.0)), font=font_small, fill=color_text)
                y += row_h

        if y + row_h <= sub_table_y1:
            draw.line((sub_table_x0, y, sub_table_x1, y), fill=color_stroke, width=max(1, int(1 * scale)))
            draw.text((sub_table_x0 + int(8 * scale), y + int(2 * scale)), "Расходы", font=font_small, fill=color_muted)
            draw.text((sub_table_x0 + col_date_w + int(8 * scale), y + int(2 * scale)), fmt_money(sub_spend["praise"]), font=font_small, fill=color_danger)
            draw.text((sub_table_x0 + col_date_w + col_w + int(8 * scale), y + int(2 * scale)), fmt_money(sub_spend["alpha"]), font=font_small, fill=color_danger)
            y += row_h

        if y + row_h <= sub_table_y1:
            sub_balances = summary.get("subaccounts") if isinstance(summary, dict) else {}
            praise_balance = float((sub_balances or {}).get("praise", {}).get("balance") or 0.0)
            alpha_balance = float((sub_balances or {}).get("alpha", {}).get("balance") or 0.0)
            draw.line((sub_table_x0, y, sub_table_x1, y), fill=color_stroke, width=max(1, int(1 * scale)))
            draw.text((sub_table_x0 + int(8 * scale), y + int(2 * scale)), "Остаток", font=font_small, fill=color_muted)
            draw.text(
                (sub_table_x0 + col_date_w + int(8 * scale), y + int(2 * scale)),
                fmt_money(praise_balance),
                font=font_small,
                fill=(color_accent if praise_balance >= 0 else color_danger),
            )
            draw.text(
                (sub_table_x0 + col_date_w + col_w + int(8 * scale), y + int(2 * scale)),
                fmt_money(alpha_balance),
                font=font_small,
                fill=(color_accent if alpha_balance >= 0 else color_danger),
            )

    # --- ALL EXPENSES LIST ---
    list_card_y = section_y + top_h + gap
    list_card = (margin, list_card_y, w - margin, list_card_y + list_card_h)
    draw_card(bg, list_card, radius, color_card, shadow_alpha=38, shadow_blur=int(18 * scale), shadow_offset=(0, int(10 * scale)))
    draw.text((margin + int(18 * scale), list_card_y + int(16 * scale)), "Все расходы", font=font_section, fill=color_text)

    def truncate_text(text: str, max_width: int, font: Any) -> str:
        if text_bbox(draw, text, font)[0] <= max_width:
            return text
        ellipsis = "…"
        for i in range(len(text), 0, -1):
            candidate = text[:i] + ellipsis
            if text_bbox(draw, candidate, font)[0] <= max_width:
                return candidate
        return ellipsis

    expense_items: List[str] = []
    for row in expenses:
        try:
            date = dt.date.fromisoformat(str(row["expense_date"])).strftime("%d.%m")
        except Exception:
            date = "—"
        category = str(row["category"] or "—")
        title2 = str(row["title"] or "—")
        total2 = fmt_money(float(row["total"] or 0.0))
        expense_items.append(f"{date} • {category}: {title2} — {total2}")

    list_x0 = margin + int(18 * scale)
    list_x1 = w - margin - int(18 * scale)
    list_y0 = list_card_y + int(52 * scale)
    list_y1 = list_card[3] - int(16 * scale)
    list_header_h = int(22 * scale)

    draw.rounded_rectangle((list_x0, list_y0, list_x1, list_y0 + list_header_h), radius=int(10 * scale), fill=color_card2)
    draw.text((list_x0 + int(10 * scale), list_y0 + int(3 * scale)), "Дата • Категория • Описание • Сумма", font=font_small, fill=color_muted)

    if not expense_items:
        draw.text((list_x0 + int(10 * scale), list_y0 + list_header_h + int(10 * scale)), "Нет расходов", font=font_small, fill=color_muted2)
    else:
        line_h = int(18 * scale)
        list_area_h = list_y1 - list_y0 - list_header_h - int(8 * scale)
        max_lines = max(int(list_area_h / line_h), 1)
        columns = max(1, math.ceil(len(expense_items) / max_lines))
        column_w = (list_x1 - list_x0) / columns

        for col in range(columns):
            col_x = list_x0 + col * column_w
            if col > 0:
                draw.line((col_x, list_y0 + list_header_h, col_x, list_y1), fill=color_stroke, width=max(1, int(1 * scale)))

            for row_idx in range(max_lines):
                item_idx = col * max_lines + row_idx
                if item_idx >= len(expense_items):
                    break
                y = list_y0 + list_header_h + int(8 * scale) + row_idx * line_h

                # легкая "зебра"
                if row_idx % 2 == 0:
                    draw.rectangle((col_x + int(4 * scale), y - int(2 * scale), col_x + column_w - int(4 * scale), y + line_h - int(2 * scale)), fill=(248, 250, 253))

                text = truncate_text(expense_items[item_idx], int(column_w - int(16 * scale)), font_small)
                draw.text((col_x + int(8 * scale), y), text, font=font_small, fill=color_text)

    # footer
    tz = CFG.tzinfo()
    now = dt.datetime.now(tz)
    tz_name = now.tzname() or CFG.TZ
    footer_text = f"Сформировано: {now.strftime('%d.%m.%Y %H:%M')} ({tz_name})"
    draw.text((margin, h - footer_h), footer_text, font=font_small, fill=color_muted2)

    out = io.BytesIO()
    bg_rgb = bg.convert("RGB")
    bg_rgb.save(out, format="PNG", optimize=True, compress_level=6, dpi=(int(dpi), int(dpi)))
    return out.getvalue()



def build_month_report_png(
    month_id: int,
    preset: str = "landscape",
    pixel_ratio: int = 2,
    dpi: int = 192,
) -> Tuple[bytes, str, sqlite3.Row]:
    if preset not in PNG_PRESETS:
        raise HTTPException(status_code=400, detail="Invalid preset")

    m = get_month_by_id(month_id)
    summary = compute_month_summary(month_id, ensure_tithe=True)

    services = db_fetchall(
        "SELECT * FROM services WHERE month_id=? AND account='main' ORDER BY service_date ASC;",
        (month_id,),
    )

    rows = db_fetchall(
        """
        SELECT category, COALESCE(SUM(total),0) AS s
        FROM expenses
        WHERE month_id=? AND account='main'
        GROUP BY category
        ORDER BY s DESC, category ASC;
        """,
        (month_id,),
    )

    top_entries = [{"category": str(r["category"] or "—"), "sum": round(float(r["s"]), 2)} for r in rows[:5]]
    sum_top = sum(item["sum"] for item in top_entries)
    while len(top_entries) < 5:
        top_entries.append({"category": "—", "sum": 0.0})
    other_sum = max(float(summary["month_expenses_sum"]) - sum_top, 0.0)
    top_entries.append({"category": "Другое", "sum": round(other_sum, 2)})

    expenses = db_fetchall(
        """
        SELECT expense_date, category, title, total, account
        FROM expenses
        WHERE month_id=? AND account='main'
        ORDER BY expense_date ASC, id ASC;
        """,
        (month_id,),
    )

    sub_services = db_fetchall(
        """
        SELECT service_date, total, account
        FROM services
        WHERE month_id=? AND account!='main'
        ORDER BY service_date ASC;
        """,
        (month_id,),
    )

    sub_expenses = db_fetchall(
        """
        SELECT expense_date, total, account
        FROM expenses
        WHERE month_id=? AND account!='main'
        ORDER BY expense_date ASC, id ASC;
        """,
        (month_id,),
    )

    png_data = render_month_report_png(
        m, summary, services, top_entries, expenses, sub_services, sub_expenses,
        preset,
        pixel_ratio=pixel_ratio,
        dpi=dpi,
    )
    filename = f"report_{m['year']}_{int(m['month']):02d}_{preset}@{int(pixel_ratio)}x.png"
    return png_data, filename, m


@APP.get("/api/export/png")
def export_png(
    month_id: int = Query(...),
    preset: str = Query("landscape"),
    pixel_ratio: int = Query(2, ge=1, le=4, description="Плотность пикселей (1..4). 2 = Retina"),
    dpi: int = Query(192, ge=72, le=600, description="DPI метаданные (72..600). На экране важнее pixel_ratio"),
    u: sqlite3.Row = Depends(require_role("admin", "accountant", "viewer")),
):
    png_data, filename, _ = build_month_report_png(month_id, preset=preset, pixel_ratio=pixel_ratio, dpi=dpi)
    return Response(
        content=png_data,
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@APP.get("/api/export/csv")
def export_csv(
    month_id: int = Query(...),
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    ensure_tithe_expense(month_id, user_id=int(u["id"]))
    m = get_month_by_id(month_id)
    services = db_fetchall("SELECT * FROM services WHERE month_id=? ORDER BY service_date ASC;", (month_id,))
    expenses = db_fetchall("SELECT * FROM expenses WHERE month_id=? ORDER BY expense_date ASC, id ASC;", (month_id,))

    import io
    import csv

    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["MONTH", int(m["year"]), int(m["month"])])
    w.writerow([])
    w.writerow(["SERVICES"])
    w.writerow(
        [
            "date",
            "idx",
            "cashless",
            "cash",
            "total",
            "weekly_min_needed",
            "mnsps_status",
            "pvs_ratio",
            "income_type",
            "account",
        ]
    )
    for s in services:
        w.writerow([
            s["service_date"], s["idx"], s["cashless"], s["cash"], s["total"],
            s["weekly_min_needed"], s["mnsps_status"], s["pvs_ratio"], s["income_type"], s["account"]
        ])

    w.writerow([])
    w.writerow(["EXPENSES"])
    w.writerow(["date", "category", "title", "qty", "unit_amount", "total", "comment", "is_system", "account"])
    for e in expenses:
        w.writerow([
            e["expense_date"], e["category"], e["title"], e["qty"], e["unit_amount"],
            e["total"], e["comment"], e["is_system"], e["account"]
        ])

    data = buf.getvalue().encode("utf-8-sig")
    filename = f"month_{m['year']}_{m['month']:02d}.csv"
    return Response(
        content=data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@APP.get("/api/export/excel")
def export_excel(
    year: int = Query(...),
    u: sqlite3.Row = Depends(require_role("admin", "accountant")),
):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    months = db_fetchall("SELECT * FROM months WHERE year=? ORDER BY month ASC;", (year,))
    wb = openpyxl.Workbook()
    # remove default
    wb.remove(wb.active)

    for m in months:
        month_id = int(m["id"])
        ensure_tithe_expense(month_id, user_id=int(u["id"]))
        recalc_services_for_month(month_id)

        ws = wb.create_sheet(title=f"{int(m['month']):02d}-{int(m['year'])}")
        ws.append(["Month", int(m["year"]), int(m["month"])])
        ws.append(["monthly_min_needed", float(m["monthly_min_needed"]), "start_balance", float(m["start_balance"])])
        ws.append([])

        ws.append(["SERVICES"])
        ws.append(
            [
                "date",
                "idx",
                "cashless",
                "cash",
                "total",
                "weekly_min_needed",
                "mnsps_status",
                "pvs_ratio",
                "income_type",
                "account",
            ]
        )
        services = db_fetchall("SELECT * FROM services WHERE month_id=? ORDER BY service_date ASC;", (month_id,))
        for s in services:
            ws.append([
                s["service_date"], s["idx"], s["cashless"], s["cash"], s["total"],
                s["weekly_min_needed"], s["mnsps_status"], s["pvs_ratio"], s["income_type"], s["account"]
            ])

        ws.append([])
        ws.append(["EXPENSES"])
        ws.append(["date", "category", "title", "qty", "unit_amount", "total", "comment", "is_system", "account"])
        expenses = db_fetchall("SELECT * FROM expenses WHERE month_id=? ORDER BY expense_date ASC, id ASC;", (month_id,))
        for e in expenses:
            ws.append([
                e["expense_date"],
                e["category"],
                e["title"],
                e["qty"],
                e["unit_amount"],
                e["total"],
                e["comment"],
                e["is_system"],
                e["account"],
            ])

        # simple width
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18

    import io
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"export_{year}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------
# Run (local)
# ---------------------------

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app:APP",
        host="0.0.0.0",
        port=int(os.getenv("PORT", "8000")),
        reload=False,
        log_level="info",
    )
