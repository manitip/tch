"""cashflow_routes.py

FastAPI роуты для модуля подтверждения наличных.

ВАЖНО: модуль специально реализован с "ленивым" импортом функций из app.py,
чтобы не создать циклический импорт при подключении:

    # app.py
    from cashflow_routes import router as cashflow_router
    APP.include_router(cashflow_router)

    import cashflow_models
    # внутри init_db(): cashflow_models.init_cashflow_db(conn)

    import cashflow_bot
    # внутри lifespan после создания bot: cashflow_bot.set_bot(bot)
"""

from __future__ import annotations

import io
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException, Query, Request
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel, Field

import cashflow_models as m
import cashflow_bot as b

import datetime as dt

router = APIRouter()


# ---------------------------
# Dependencies (lazy import from app.py)
# ---------------------------


def _app_db_connect():
    from app import db_connect  # lazy
    return db_connect


def _app_get_current_user():
    from app import get_current_user  # lazy
    return get_current_user


def _app_require_role(*roles: str):
    def _dep(request: Request):
        from app import get_current_user  # lazy
        u = get_current_user(request)
        role = str(u["role"])
        if role not in roles:
            raise HTTPException(status_code=403, detail="Insufficient role")
        return u

    return _dep

def _app_finalize_cashflow_collect_request():
    from app import finalize_cashflow_collect_request  # lazy
    return finalize_cashflow_collect_request


def _get_cfg_paths() -> Dict[str, Any]:
    from app import CFG  # lazy
    return {
        "base_dir": Path(__file__).resolve().parent,
        "users_json": Path(CFG.USERS_JSON_PATH),
        "db_path": Path(CFG.DB_PATH),
    }


def _ensure_cashflow_tables() -> None:
    db_connect = _app_db_connect()
    with db_connect() as conn:
        m.init_cashflow_db(conn)


def _cash_cfg() -> m.CashflowConfig:
    paths = _get_cfg_paths()
    # переопределяем paths, чтобы совпадало с app.py
    cfg = m.load_cashflow_config(paths["base_dir"])
    return m.CashflowConfig(
        base_dir=cfg.base_dir,
        db_path=paths["db_path"],
        users_json_path=paths["users_json"],
        uploads_dir=cfg.uploads_dir,
        timezone=cfg.timezone,
    )


# ---------------------------
# Models
# ---------------------------


class CashRequestCreateIn(BaseModel):
    account: str = Field(..., description="main|praise|alpha")
    op_type: str = Field(..., description="collect|withdraw")
    amount: float
    source_kind: Optional[str] = None
    source_id: Optional[int] = None


class CashResendIn(BaseModel):
    target_telegram_ids: Optional[List[int]] = Field(None, description="Если не задано — всем отказавшим на попытке 1")
    admin_comment: Optional[str] = None


class CashSignIn(BaseModel):
    signature: str = Field(..., description="data:image/png;base64,... или base64")


class CashRefuseIn(BaseModel):
    reason: str = Field(..., min_length=1)


# ---------------------------
# HTML entry (cashapp)
# ---------------------------


@router.get("/cashapp")
def cashapp_html():
    """Отдельный интерфейс подписанта (не бухгалтерия)."""
    path = Path(__file__).resolve().parent / "cashapp.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="cashapp.html not found")
    return FileResponse(str(path), media_type="text/html")


# ---------------------------
# API: Requests
# ---------------------------


@router.post("/api/cashflow/requests")
def create_request(
    body: CashRequestCreateIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("admin", "accountant")),
):
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()

    with db_connect() as conn:
        request_id = m.create_cash_request(
            conn,
            cfg,
            account=body.account,
            op_type=body.op_type,
            amount=float(body.amount),
            created_by_telegram_id=int(u["telegram_id"]),
            source_kind=body.source_kind,
            source_id=body.source_id,
        )
        view = m.build_request_view(conn, request_id)
        req = view["request"]
        signers = [p["telegram_id"] for p in view["participants"] if not p["is_admin"]]

    # уведомляем подписантов
    bg.add_task(
        b.notify_signers_new_request,
        request_id=request_id,
        account=req["account"],
        op_type=req["op_type"],
        amount=float(req["amount"]),
        signer_ids=signers,
        is_retry=False,
        admin_comment=None,
    )
    return {"id": request_id, "item": view}


@router.get("/api/cashflow/requests/my")
def my_requests(
    account: Optional[str] = Query(None),
    only_open: bool = Query(False),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    u=Depends(_app_require_role("admin", "accountant", "viewer", "cash_signer")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        items = m.list_my_cash_requests(
            conn,
            telegram_id=int(u["telegram_id"]),
            account=account,
            only_open=bool(only_open),
            limit=limit,
            offset=offset,
        )
        return {"items": [dict(x) for x in items]}


@router.get("/api/cashflow/requests")
def list_requests(
    account: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        items = m.list_cash_requests(conn, account=account, status=status, limit=limit, offset=offset)
        return {"items": [dict(x) for x in items]}


@router.get("/api/cashflow/requests/{request_id}")
def request_detail(
    request_id: int,
    u=Depends(_app_require_role("admin", "accountant", "viewer", "cash_signer")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        view = m.build_request_view(conn, int(request_id))
        # доступ: админ или участник
        if str(u["role"]) != "admin":
            tid = int(u["telegram_id"])
            if not any(int(p["telegram_id"]) == tid for p in view["participants"]):
                raise HTTPException(status_code=403, detail="Not a participant")
        return view


# ---------------------------
# API: Sign / Refuse
# ---------------------------


@router.post("/api/cashflow/requests/{request_id}/sign")
def sign_request(
    request_id: int,
    body: CashSignIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("cash_signer", "admin")),
):
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        try:
            m.record_signature(
                conn,
                cfg,
                request_id=int(request_id),
                telegram_id=int(u["telegram_id"]),
                signature_data_url=body.signature,
                as_admin=False,
            )
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

        view = m.build_request_view(conn, int(request_id))
        req = view["request"]
        # найдём имя подписанта
        signer_name = str(u["name"])

    if req.get("status") == "FINAL":
        finalize = _app_finalize_cashflow_collect_request()
        finalize(int(request_id))


    # уведомить ответственного админа
    bg.add_task(
        b.notify_admin_about_decision,
        admin_id=int(req["admin_telegram_id"]),
        request_id=int(request_id),
        account=req["account"],
        op_type=req["op_type"],
        amount=float(req["amount"]),
        signer_name=signer_name,
        decision="SIGNED",
        reason=None,
    )
    # если финализировано — уведомить инициатора
    if req.get("created_by_telegram_id") and req.get("status") == "FINAL":
        bg.add_task(
            b.notify_initiator_final,
            initiator_id=int(req["created_by_telegram_id"]),
            request_id=int(request_id),
            account=req["account"],
            op_type=req["op_type"],
            amount=float(req["amount"]),
            status=str(req["status"]),
        )

    return {"ok": True, "item": view}


@router.post("/api/cashflow/requests/{request_id}/refuse")
def refuse_request(
    request_id: int,
    body: CashRefuseIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("cash_signer")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        try:
            m.record_refusal(
                conn,
                request_id=int(request_id),
                telegram_id=int(u["telegram_id"]),
                reason=str(body.reason).strip(),
            )
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

        view = m.build_request_view(conn, int(request_id))
        req = view["request"]
        signer_name = str(u["name"])

    bg.add_task(
        b.notify_admin_about_decision,
        admin_id=int(req["admin_telegram_id"]),
        request_id=int(request_id),
        account=req["account"],
        op_type=req["op_type"],
        amount=float(req["amount"]),
        signer_name=signer_name,
        decision="REFUSED",
        reason=str(body.reason).strip(),
    )
    return {"ok": True, "item": view}


# ---------------------------
# API: Admin actions
# ---------------------------


@router.post("/api/cashflow/requests/{request_id}/resend")
def admin_resend(
    request_id: int,
    body: CashResendIn = Body(default_factory=CashResendIn),
    bg: BackgroundTasks = None,  # type: ignore[assignment]
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    cfg = _cash_cfg()
    with db_connect() as conn:
        try:
            targets = m.resend_for_refusals(
                conn,
                request_id=int(request_id),
                admin_telegram_id=int(u["telegram_id"]),
                target_telegram_ids=body.target_telegram_ids,
                admin_comment=body.admin_comment,
            )
            view = m.build_request_view(conn, int(request_id))
            req = view["request"]
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    if bg is not None:
        bg.add_task(
            b.notify_signers_new_request,
            request_id=int(request_id),
            account=req["account"],
            op_type=req["op_type"],
            amount=float(req["amount"]),
            signer_ids=targets,
            is_retry=True,
            admin_comment=req.get("admin_comment"),
        )
    return {"ok": True, "targets": targets, "item": view}


@router.post("/api/cashflow/requests/{request_id}/admin-sign")
def admin_sign(
    request_id: int,
    body: CashSignIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("admin")),
):
    raise HTTPException(status_code=403, detail="Admin sign is disabled, use /sign")


@router.post("/api/cashflow/requests/{request_id}/cancel")
def admin_cancel(
    request_id: int,
    comment: Optional[str] = Body(None),
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        try:
            m.cancel_request(conn, request_id=int(request_id), admin_telegram_id=int(u["telegram_id"]), comment=comment)
            view = m.build_request_view(conn, int(request_id))
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "item": view}


# ---------------------------
# API: Signature image
# ---------------------------


@router.get("/api/cashflow/requests/{request_id}/participants/{telegram_id}/signature.png")
def get_signature_png(
    request_id: int,
    telegram_id: int,
    u=Depends(_app_require_role("admin", "accountant", "viewer", "cash_signer")),
):
    """Возвращает эффективную подпись участника: attempt=2 если есть, иначе attempt=1."""
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()

    with db_connect() as conn:
        view = m.build_request_view(conn, int(request_id))

        role = str(u["role"])
        caller_tid = int(u["telegram_id"])

        # доступ: админ ИЛИ тот, у кого есть доступ к акту ИЛИ участник заявки
        if role != "admin":
            can_view_act = False
            try:
                can_view_act = _user_can_view_withdraw_act(caller_tid, role)
            except Exception:
                can_view_act = False

            if not can_view_act:
                if not any(int(p["telegram_id"]) == caller_tid for p in view["participants"]):
                    raise HTTPException(status_code=403, detail="Not allowed")

        # находим нужного участника
        part = None
        for p in view["participants"]:
            if int(p["telegram_id"]) == int(telegram_id):
                part = p
                break
        if not part:
            raise HTTPException(status_code=404, detail="Participant not found")

        # attempt=2 если есть, иначе attempt=1
        sig = part.get("attempt2") or part.get("attempt1") or None
        if not sig or str(sig.get("decision")) != "SIGNED" or not sig.get("signature_path"):
            raise HTTPException(status_code=404, detail="Signature not found")


        try:
            img_path = m.get_signature_file_path(cfg, str(sig["signature_path"]))
        except ValueError:
            raise HTTPException(status_code=404, detail="Signature not found")
        if not img_path.exists():
            raise HTTPException(status_code=404, detail="Signature file missing")

        return FileResponse(str(img_path), media_type="image/png")


# ---------------------------
# API: Withdrawal Act (view + Excel)
# ---------------------------


def _user_can_view_withdraw_act(telegram_id: int, role: str) -> bool:
    if role in ("admin", "accountant"):
        return True
    cfg = _cash_cfg()
    allow = m.load_users_allowlist(cfg.users_json_path)
    u = allow.get(int(telegram_id))
    if not u or not (u.get("active") is True or u.get("active") == 1):
        return False
    ops = u.get("cash_ops") or []
    if isinstance(ops, str):
        ops = [x.strip() for x in ops.split(",") if x.strip()]
    ops_n = [str(x).lower() for x in ops]
    # Акт включает операции наличных (collect/withdraw), поэтому допускаем обе.
    return str(u.get("role")) == "cash_signer" and ("withdraw" in ops_n or "collect" in ops_n)


@router.get("/api/cashflow/withdraw-act")
def withdraw_act(
    account: Optional[str] = Query(None, description="main|praise|alpha (опционально)"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    u=Depends(_app_require_role("admin", "accountant", "cash_signer")),
):
    if not _user_can_view_withdraw_act(int(u["telegram_id"]), str(u["role"])):
        raise HTTPException(status_code=403, detail="No access to withdraw act")
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        rows = m.list_withdraw_act_rows(conn, account=account, date_from=date_from, date_to=date_to)
    return {"items": rows}

@router.get("/api/cashflow/withdraw-act.xlsx")
def withdraw_act_xlsx(
    account: Optional[str] = Query(None, description="main|praise|alpha (опционально)"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    u=Depends(_app_require_role("admin", "accountant")),
):
    """Экспорт акта наличных (сбор/изъятие) в Excel с PNG подписями."""
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()



    import importlib.util

    if importlib.util.find_spec("openpyxl") is None:
        raise HTTPException(status_code=501, detail="openpyxl is not available")

    # Pillow нужен для вставки PNG в XLSX через openpyxl (иначе подписи не будут "живыми" картинками)
    if importlib.util.find_spec("PIL") is None:
        raise HTTPException(
            status_code=501,
            detail="Pillow (PIL) is required to embed signature images into Excel",
        )

    import openpyxl
    from openpyxl.utils.units import points_to_pixels, pixels_to_EMU
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D



    with db_connect() as conn:
        rows_all = m.list_withdraw_act_rows(conn, account=account, date_from=date_from, date_to=date_to)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    temp_paths: list[str] = []

    def make_sheet(account_code: str, title: str):
        ws = wb.create_sheet(title)
        ws.append(["Дата", "Операция", "Сумма", "ФИО", "Тип пользователя", "Подпись"])
        ws.freeze_panes = "A2"

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 32
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 24  # подпись

        # стили
        thin = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F3F4F6")
        header_font = Font(bold=True)
        center = Alignment(vertical="center", horizontal="center", wrap_text=True)
        left = Alignment(vertical="center", horizontal="left", wrap_text=True)

        # header style
        for col in range(1, 7):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border

        rnum = 2
        for r in rows_all:
            if r.get("account") != account_code:
                continue

            # --- ДАТА: только дата ---
            date_cell = ws.cell(row=rnum, column=1)
            raw_date = r.get("date")  # ожидаем YYYY-MM-DD
            if raw_date:
                s = str(raw_date)[:10]
                try:
                    date_cell.value = dt.date.fromisoformat(s)
                except Exception:
                    date_cell.value = s
            else:
                date_cell.value = ""
            date_cell.number_format = "DD.MM.YYYY"
            date_cell.alignment = center
            date_cell.border = border

            # --- ОПЕРАЦИЯ ---
            op = str(r.get("op_type") or "")
            op_disp = {"collect": "Сбор наличных", "withdraw": "Изъятие наличных"}.get(op, op)
            c2 = ws.cell(row=rnum, column=2, value=op_disp)
            c2.alignment = left
            c2.border = border

            # --- СУММА: всегда .00 ---
            amt = float(r.get("amount") or 0)
            c3 = ws.cell(row=rnum, column=3, value=amt)
            c3.number_format = "#,##0.00"
            c3.alignment = center
            c3.border = border

            # --- ФИО / ТИП ---
            c4 = ws.cell(row=rnum, column=4, value=r.get("fio"))
            c4.alignment = left
            c4.border = border

            c5 = ws.cell(row=rnum, column=5, value=r.get("user_type"))
            c5.alignment = center
            c5.border = border

            # --- ПОДПИСЬ ---
            sig_cell = ws.cell(row=rnum, column=6)
            sig_cell.border = border
            sig_cell.alignment = center

            signature_path = r.get("signature_path")
            signature_value = r.get("signature_value") or ""

            # место под картинку
            ws.row_dimensions[rnum].height = 36

            # подписано -> вставляем PNG (ЖИВАЯ подпись)
            if signature_path:
                try:
                    img_path = m.get_signature_file_path(cfg, str(signature_path))
                except ValueError:
                    img_path = None
                if img_path and img_path.exists():
                    raw = m.normalize_signature_png_bytes(img_path.read_bytes())

                    # чтобы в ячейке не было текста
                    sig_cell.value = ""


                    try:
                        # openpyxl.Image надёжнее работает с путём к файлу, чем с BytesIO (зависит от версии)
                        with tempfile.NamedTemporaryFile(prefix="sig_", suffix=".png", delete=False) as tmp:
                            tmp.write(raw)
                            temp_paths.append(tmp.name)

                        img = XLImage(temp_paths[-1])
                        img.width = 120
                        img.height = 32
                        col_width = ws.column_dimensions["F"].width or 8.43
                        row_height = ws.row_dimensions[rnum].height or 15
                        cell_width_px = int(col_width * 7 + 5)
                        cell_height_px = points_to_pixels(row_height)
                        x_off = max((cell_width_px - img.width) / 2, 0)
                        y_off = max((cell_height_px - img.height) / 2, 0)
                        marker = AnchorMarker(
                            col=5,
                            colOff=pixels_to_EMU(int(x_off)),
                            row=rnum - 1,
                            rowOff=pixels_to_EMU(int(y_off)),
                        )
                        size = XDRPositiveSize2D(
                            cx=pixels_to_EMU(img.width),
                            cy=pixels_to_EMU(img.height),
                        )
                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                        ws.add_image(img)  # встраивает подпись в XLSX
                    except Exception:
                        # если по какой-то причине вставка изображения не удалась — хотя бы покажем статус
                        sig_cell.value = str(signature_value or "SIGNED")

                else:
                    sig_cell.value = str(signature_value)
            else:
                # нет файла подписи (например, ещё не подписано/отказ) — показываем статус текстом
                sig_cell.value = str(signature_value)

            rnum += 1

    if account:
        acc = account.strip().lower()
        if acc not in m.ACCOUNTS:
            raise HTTPException(status_code=400, detail="Invalid account")
        make_sheet(acc, acc.upper())
    else:
        make_sheet("main", "MAIN")
        make_sheet("praise", "PRAISE")
        make_sheet("alpha", "ALPHA")

    bio = io.BytesIO()
    wb.save(bio)
    for tmp_path in temp_paths:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
    data = bio.getvalue()

    filename = "withdraw_act.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Length": str(len(data)),
    }
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
